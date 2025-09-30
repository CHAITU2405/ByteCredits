from flask import Flask, render_template, request,flash,session,redirect,url_for, jsonify, send_file
from flask import send_from_directory, abort
from models import db, Credentials, Transactions, CollegePayments, PaymentProof, RefundRequest, LibraryBook, LibraryTransaction, MockExam, ExamAttempt, ExamResult, LearningPod, PodMembership, StudentProfile, PodTask, TaskSubmission, MeritBadge, CollaborationEvent, NfcCard, ResultDeclaration, ExamContest, ContestParticipation
import os
import re
import nfc
from datetime import datetime, timedelta
import json
import pandas as pd
import io
import openpyxl
from openpyxl.utils import get_column_letter
import google.generativeai as genai
import time
import random
import string
import uuid
from typing import Optional, Dict
import PyPDF2
import base64
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# ---------- Email Configuration ----------
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_USERNAME = os.getenv('SMTP_USERNAME', 'your-email@gmail.com')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD', 'your-app-password')


# ---------- Gemini helper (SDK-agnostic) ----------
def gemini_generate_text(prompt: str) -> str:
    """Generate text using Gemini via REST, trying multiple public models.
    Prioritizes endpoints confirmed to work in this environment.
    Returns plain text; raises RuntimeError on failure.
    """
    try:
        import requests
        api_key = GEMINI_API_KEY
        if not api_key:
            raise RuntimeError("Missing GEMINI_API_KEY")

        endpoints = [
            "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent",
            "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-preview-05-20:generateContent",
            "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-pro-preview-06-05:generateContent",
            "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-lite-preview-06-17:generateContent",
        ]

        payload = {
            "contents": [{"parts": [{"text": prompt}]}]
        }
        headers = {"Content-Type": "application/json"}

        last_err = None
        for ep in endpoints:
            try:
                resp = requests.post(f"{ep}?key={api_key}", headers=headers, data=json.dumps(payload), timeout=60)
                if resp.status_code != 200:
                    last_err = f"HTTP {resp.status_code}: {resp.text[:200]}"
                    continue
                data = resp.json()
                parts = (((data or {}).get('candidates') or [{}])[0].get('content') or {}).get('parts') or []
                text = ''.join([p.get('text', '') for p in parts])
                if text:
                    return text
                last_err = "Empty response"
            except Exception as e:
                last_err = str(e)
                continue
        raise RuntimeError(last_err or "All Gemini endpoints failed")
    except Exception as e:
        raise RuntimeError(f"Gemini generation failed: {e}")

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///byte_credits.db'
app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads')
app.config['QUESTION_PAPERS_FOLDER'] = os.path.join('static', 'question_papers')
app.config['EXAM_FOLDER'] = os.path.join('static', 'exams')
app.config['PDF_UPLOAD_FOLDER'] = os.path.join('static', 'pdf_uploads')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'app'

# Configure Gemini API
# Prefer environment variable, fallback to existing key
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', 'AIzaSyBJekYGPvQ19725xseLuN7s-noiSV0MaGc')
genai.configure(api_key=GEMINI_API_KEY)

db.init_app(app)

def extract_text_from_pdf(pdf_path: str) -> str:
    """Extract text content from PDF file with multiple fallback methods"""
    if not os.path.exists(pdf_path):
        print(f"PDF file not found: {pdf_path}")
        return ""
    
    # Method 1: Try modern pypdf library (recommended)
    try:
        import pypdf
        with open(pdf_path, 'rb') as file:
            pdf_reader = pypdf.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                try:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                except Exception as e:
                    print(f"Error extracting text from page: {e}")
                    continue
            if text.strip():
                return text.strip()
    except Exception as e:
        print(f"pypdf extraction failed: {e}")
    
    # Method 2: Try PyPDF2 (legacy)
    try:
        import PyPDF2
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                try:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                except Exception as e:
                    print(f"Error extracting text from page: {e}")
                    continue
            if text.strip():
                return text.strip()
    except Exception as e:
        print(f"PyPDF2 extraction failed: {e}")
    
    # Method 3: Try pdfminer.six (most robust)
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(pdf_path)
        if text and text.strip():
            return text.strip()
    except Exception as e:
        print(f"pdfminer extraction failed: {e}")
    
    print(f"All PDF extraction methods failed for: {pdf_path}")
    return ""

def process_pdf_for_chatbot(pdf_path: str) -> str:
    """Process PDF and create a summary for chatbot context"""
    try:
        text = extract_text_from_pdf(pdf_path)
        if not text:
            return ""
        
        # Try to use Gemini to create a summary, fallback to basic processing
        try:
            prompt = f"""
            Please analyze and summarize the following PDF content for a student chatbot context. 
            Create a clear, concise summary that covers the main topics, key concepts, and important details.
            The summary should be educational and help answer questions about the content.
            
            PDF Content:
            {text[:4000]}  # Limit to avoid token limits
            
            Please provide a structured summary with:
            1. Main topics covered
            2. Key concepts and definitions
            3. Important details and examples
            4. Any practical applications or examples
            
            Keep the summary under 1000 words and make it easy to understand for students.
            """
            
            summary = gemini_generate_text(prompt)
            if summary and summary.strip():
                return summary
        except Exception as e:
            print(f"Gemini processing failed, using fallback: {e}")
        
        # Fallback: Create a basic summary from the extracted text
        # Take first 2000 characters and clean it up
        clean_text = text.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        # Remove multiple spaces
        import re
        clean_text = re.sub(r'\s+', ' ', clean_text).strip()
        
        # Create a basic summary
        summary = f"""
        PDF Content Summary:
        
        This document contains educational content with approximately {len(text)} characters.
        Key topics appear to include: {clean_text[:500]}...
        
        The full content is available for detailed analysis and question answering.
        """
        
        return summary
        
    except Exception as e:
        print(f"Error processing PDF: {e}")
        return ""

def generate_chat_summary(chat_history: list) -> str:
    """Generate a summary of the chat conversation using AI"""
    try:
        if not chat_history:
            return "No chat history available for summary."
        
        # Format chat history for AI processing
        chat_text = ""
        for message in chat_history:
            role = "Student" if message.get('isUser', False) else "Assistant"
            content = message.get('content', '')
            timestamp = message.get('timestamp', '')
            chat_text += f"{role} ({timestamp}): {content}\n\n"
        
        # Use Gemini to create a comprehensive summary
        prompt = f"""
        Please create a comprehensive summary of the following chat conversation between a student and an AI assistant.
        The summary should be educational and useful for the student to review their learning session.
        
        Chat Conversation:
        {chat_text[:3000]}  # Limit to avoid token limits
        
        Please provide a structured summary that includes:
        1. Main topics discussed
        2. Key questions asked by the student
        3. Important answers and explanations provided
        4. Key learning points and takeaways
        5. Any concepts or topics that might need further study
        
        Make the summary clear, educational, and helpful for the student to understand what they learned.
        Keep it under 800 words and use bullet points for better readability.
        """
        
        summary = gemini_generate_text(prompt)
        return summary
    except Exception as e:
        print(f"Error generating chat summary: {e}")
        return "Failed to generate chat summary. Please try again."

def create_chat_summary_pdf(chat_history: list, username: str) -> str:
    """Create a PDF summary of the chat conversation"""
    try:
        # Generate summary
        summary = generate_chat_summary(chat_history)
        
        # Create PDF filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"chat_summary_{username}_{timestamp}.pdf"
        file_path = os.path.join(app.config['PDF_UPLOAD_FOLDER'], filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#667eea')
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#374151')
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=6,
            alignment=TA_LEFT
        )
        
        # Build PDF content
        story = []
        
        # Title
        story.append(Paragraph("Chat Conversation Summary", title_style))
        story.append(Spacer(1, 20))
        
        # Student info
        story.append(Paragraph(f"<b>Student:</b> {username}", normal_style))
        story.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", normal_style))
        story.append(Spacer(1, 20))
        
        # Summary section
        story.append(Paragraph("Summary", heading_style))
        story.append(Paragraph(summary, normal_style))
        story.append(Spacer(1, 20))
        
        # Chat history section
        story.append(Paragraph("Chat History", heading_style))
        
        # Create table for chat history
        chat_data = [['Time', 'Speaker', 'Message']]
        for message in chat_history[-10:]:  # Show last 10 messages
            role = "Student" if message.get('isUser', False) else "Assistant"
            content = message.get('content', '')[:100] + "..." if len(message.get('content', '')) > 100 else message.get('content', '')
            timestamp = message.get('timestamp', '')
            chat_data.append([timestamp, role, content])
        
        chat_table = Table(chat_data, colWidths=[1*inch, 1*inch, 4*inch])
        chat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(chat_table)
        story.append(Spacer(1, 20))
        
        # Footer
        story.append(Paragraph("Generated by ByteCredits Student Assistant", 
                              ParagraphStyle('Footer', parent=styles['Normal'], 
                                           fontSize=8, alignment=TA_CENTER, 
                                           textColor=colors.grey)))
        
        # Build PDF
        doc.build(story)
        
        return file_path
        
    except Exception as e:
        print(f"Error creating chat summary PDF: {e}")
        return None

def is_educational_query(user_text: str) -> bool:
    """Return True if the message is educational/college-app related.
    Blocks entertainment/fun-only prompts (jokes, memes, roasts, songs, etc.)."""
    if not user_text:
        return False
    text = user_text.lower()

    entertainment_keywords = [
        'joke', 'memes', 'meme', 'funny', 'roast', 'insult', 'prank', 'pickup line',
        'pickup-line', 'love letter', 'lyrics', 'song', 'poem', 'poetry', 'story',
        'horoscope', 'astrology', 'celebrity', 'movie', 'tv show', 'game idea',
        'standup', 'stand-up', 'riddle', 'dark humor', 'fun question'
    ]

    # Explicitly off-topic consumer-service queries
    off_topic_services = [
        'bookmyshow', 'book my show', 'movie ticket', 'movie tickets', 'train ticket', 'irctc',
        'flight ticket', 'flight booking', 'bus ticket', 'redbus', 'ola', 'uber', 'swiggy', 'zomato',
        'food order', 'amazon', 'flipkart', 'myntra', 'shopping', 'crypto', 'bitcoin', 'stock market',
        'trading', 'demat', 'loan apply', 'credit card', 'bank account', 'pan card', 'aadhar',
        'instagram followers', 'youtube subscribers', 'reels idea', 'whatsapp trick'
    ]

    educational_overrides = [
        # general academic / college-app related
        'study', 'assignment', 'homework', 'exam', 'mock exam', 'question paper', 'subject',
        'syllabus', 'notes', 'explain', 'definition', 'example', 'how to', 'tutorial',
        # app features
        'fees', 'payment', 'library', 'attendance', 'nfc', 'card', 'learning pod', 'pod task',
        # common subjects / cs terms
        'math', 'physics', 'chemistry', 'biology', 'english', 'computer', 'programming',
        'python', 'java', 'dbms', 'database', 'operating system', 'os', 'data structures', 'dsa', 'algorithms',
        'object oriented', 'oop', 'oops', 'inheritance', 'polymorphism', 'encapsulation', 'abstraction',
        'network', 'cn', 'wt', 'web technology', 'software engineering', 'compiler', 'ai', 'ml', 'data science'
    ]

    # Hard block explicit off-topic consumer services
    if any(k in text for k in off_topic_services):
        return False

    has_educational_signal = any(k in text for k in educational_overrides)

    # If clearly entertainment-only and no academic signal, block
    if any(k in text for k in entertainment_keywords) and not has_educational_signal:
        return False

    # Otherwise allow by default as long as it's not explicitly off-topic
    return True if (has_educational_signal or not any(k in text for k in entertainment_keywords)) else False

def migrate_database_schema():
    """Add missing columns to existing tables"""
    try:
        # Check if mock_exams table exists and add missing columns
        with db.engine.connect() as conn:
            # Check if total_marks column exists
            result = conn.execute(db.text("PRAGMA table_info(mock_exams)"))
            columns = [row[1] for row in result.fetchall()]
            
            if 'total_marks' not in columns:
                print("Adding total_marks column to mock_exams table...")
                conn.execute(db.text("ALTER TABLE mock_exams ADD COLUMN total_marks INTEGER NOT NULL DEFAULT 50"))
                conn.commit()
            
            if 'difficulty_level' not in columns:
                print("Adding difficulty_level column to mock_exams table...")
                conn.execute(db.text("ALTER TABLE mock_exams ADD COLUMN difficulty_level VARCHAR(20) DEFAULT 'medium'"))
                conn.commit()
            
            if 'format_description' not in columns:
                print("Adding format_description column to mock_exams table...")
                conn.execute(db.text("ALTER TABLE mock_exams ADD COLUMN format_description TEXT"))
                conn.commit()

            # Ensure nfc_cards table has new columns
            result = conn.execute(db.text("PRAGMA table_info(nfc_cards)"))
            nfc_columns = [row[1] for row in result.fetchall()]
            # Ensure student_profiles has resume fields
            result = conn.execute(db.text("PRAGMA table_info(student_profiles)"))
            sp_columns = [row[1] for row in result.fetchall()]
            if 'resume_json' not in sp_columns:
                print("Adding resume_json to student_profiles...")
                conn.execute(db.text("ALTER TABLE student_profiles ADD COLUMN resume_json TEXT"))
                conn.commit()
            if 'is_resume_public' not in sp_columns:
                print("Adding is_resume_public to student_profiles...")
                conn.execute(db.text("ALTER TABLE student_profiles ADD COLUMN is_resume_public BOOLEAN DEFAULT 0"))
                conn.commit()
            if 'public_slug' not in sp_columns:
                print("Adding public_slug to student_profiles...")
                conn.execute(db.text("ALTER TABLE student_profiles ADD COLUMN public_slug VARCHAR(64)"))
                conn.commit()

            # Ensure result_declarations table exists (created by models)
            # Nothing to do here; db.create_all() will create if missing

            # Add cohort fields to credentials
            result = conn.execute(db.text("PRAGMA table_info(credentials)"))
            cred_columns = [row[1] for row in result.fetchall()]
            if 'study_year' not in cred_columns:
                print("Adding study_year to credentials...")
                conn.execute(db.text("ALTER TABLE credentials ADD COLUMN study_year INTEGER"))
                conn.commit()
            if 'department' not in cred_columns:
                print("Adding department to credentials...")
                conn.execute(db.text("ALTER TABLE credentials ADD COLUMN department VARCHAR(50)"))
                conn.commit()
            if 'section' not in cred_columns:
                print("Adding section to credentials...")
                conn.execute(db.text("ALTER TABLE credentials ADD COLUMN section VARCHAR(10)"))
                conn.commit()

            # Add cohort fields to learning_pods
            result = conn.execute(db.text("PRAGMA table_info(learning_pods)"))
            pod_columns = [row[1] for row in result.fetchall()]
            # Add points column to credentials
            result = conn.execute(db.text("PRAGMA table_info(credentials)"))
            cred_cols = [row[1] for row in result.fetchall()]
            if 'points' not in cred_cols:
                print("Adding points to credentials...")
                conn.execute(db.text("ALTER TABLE credentials ADD COLUMN points INTEGER DEFAULT 0"))
                conn.commit()

            # Add grading fields to task_submissions
            result = conn.execute(db.text("PRAGMA table_info(task_submissions)"))
            ts_cols = [row[1] for row in result.fetchall()]
            if 'awarded_points' not in ts_cols:
                print("Adding awarded_points to task_submissions...")
                conn.execute(db.text("ALTER TABLE task_submissions ADD COLUMN awarded_points INTEGER DEFAULT 0"))
                conn.commit()
            if 'teacher_feedback' not in ts_cols:
                print("Adding teacher_feedback to task_submissions...")
                conn.execute(db.text("ALTER TABLE task_submissions ADD COLUMN teacher_feedback TEXT"))
                conn.commit()
            if 'evaluated_at' not in ts_cols:
                print("Adding evaluated_at to task_submissions...")
                conn.execute(db.text("ALTER TABLE task_submissions ADD COLUMN evaluated_at DATETIME"))
                conn.commit()
            if 'evaluated_by' not in ts_cols:
                print("Adding evaluated_by to task_submissions...")
                conn.execute(db.text("ALTER TABLE task_submissions ADD COLUMN evaluated_by VARCHAR(50)"))
                conn.commit()
            if 'study_year' not in pod_columns:
                print("Adding study_year to learning_pods...")
                conn.execute(db.text("ALTER TABLE learning_pods ADD COLUMN study_year INTEGER"))
                conn.commit()
            if 'department' not in pod_columns:
                print("Adding department to learning_pods...")
                conn.execute(db.text("ALTER TABLE learning_pods ADD COLUMN department VARCHAR(50)"))
                conn.commit()
            if 'section' not in pod_columns:
                print("Adding section to learning_pods...")
                conn.execute(db.text("ALTER TABLE learning_pods ADD COLUMN section VARCHAR(10)"))
                conn.commit()

            if 'email' not in nfc_columns:
                print("Adding email column to nfc_cards table...")
                conn.execute(db.text("ALTER TABLE nfc_cards ADD COLUMN email VARCHAR(120)"))
                conn.commit()

            if 'roll_no' not in nfc_columns:
                print("Adding roll_no column to nfc_cards table...")
                conn.execute(db.text("ALTER TABLE nfc_cards ADD COLUMN roll_no VARCHAR(50)"))
                conn.commit()
                
        print("Database migration completed successfully!")
        
    except Exception as e:
        print(f"Database migration error: {e}")
        # If migration fails, the app will still work with existing data

with app.app_context():
    db.create_all()
    migrate_database_schema()

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    print("Creating upload folder...")
    os.makedirs(app.config['UPLOAD_FOLDER'])

if not os.path.exists(app.config['QUESTION_PAPERS_FOLDER']):
    print("Creating question papers folder...")
    os.makedirs(app.config['QUESTION_PAPERS_FOLDER'])

if not os.path.exists(app.config['EXAM_FOLDER']):
    print("Creating exams folder...")
    os.makedirs(app.config['EXAM_FOLDER'])

# Install with: pip install nfcpy



# ---------------------- Result Declarations ----------------------
def _extract_text_with_pypdf(pdf_path: str) -> str:
    """Extract text from PDF using multiple methods with fallbacks"""
    if not os.path.exists(pdf_path):
        print(f"PDF file not found: {pdf_path}")
        return ""
    
    # Try modern 'pypdf' first (recommended)
    try:
        import pypdf
        text_parts: list[str] = []
        with open(pdf_path, 'rb') as f:
            reader = pypdf.PdfReader(f)
            print(f"PDF has {len(reader.pages)} pages")
            for page_num, page in enumerate(reader.pages):
                try:
                    page_text = page.extract_text()
                    if page_text and page_text.strip():
                        text_parts.append(page_text.strip())
                        print(f"Page {page_num + 1}: Extracted {len(page_text)} characters")
                    else:
                        print(f"Page {page_num + 1}: No text extracted")
                except Exception as e:
                    print(f"Error extracting page {page_num + 1}: {e}")
                    continue
        collected = "\n".join(text_parts).strip()
        if collected:
            print(f"Total extracted text length: {len(collected)} characters")
            return collected
        else:
            print("No text collected from any page")
    except Exception as e:
        print(f"pypdf extraction failed: {e}")

    # Fallback to legacy 'PyPDF2'
    try:
        import PyPDF2
        text_parts = []
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page_num, page in enumerate(reader.pages):
                try:
                    page_text = page.extract_text()
                    if page_text and page_text.strip():
                        text_parts.append(page_text.strip())
                except Exception as e:
                    print(f"Error extracting page {page_num + 1}: {e}")
                    continue
        collected = "\n".join(text_parts).strip()
        if collected:
            return collected
    except Exception as e:
        print(f"PyPDF2 extraction failed: {e}")

    # Fallback to pdfminer.six (most robust for complex PDFs)
    try:
        from pdfminer.high_level import extract_text
        collected = extract_text(pdf_path)
        if collected and collected.strip():
            return collected.strip()
    except Exception as e:
        print(f"pdfminer extraction failed: {e}")
    
    print(f"All PDF extraction methods failed for: {pdf_path}")
    return ''

def _validate_roll_number(roll: str) -> bool:
    """Validate if a string looks like a valid roll number"""
    if not roll or len(roll) < 4 or len(roll) > 15:
        return False
    
    # Skip if all zeros or too many zeros
    if roll.count('0') > len(roll) * 0.8:  # Skip if more than 80% zeros
        return False
    
    # Skip if too many repeated digits/characters
    if len(set(roll)) < 3:  # Skip if less than 3 unique characters
        return False
    
    # Skip common non-roll-number patterns
    invalid_patterns = [
        'HTTP', 'HTML', 'PDF', 'XML', 'CSS', 'JS', 'PHP', 'SQL',
        'WWW', 'COM', 'ORG', 'NET', 'GOV', 'EDU',
        'COPYRIGHT', 'ALLRIGHTS', 'RESERVED'
    ]
    
    if roll in invalid_patterns:
        return False
    
    # Skip if it's mostly punctuation or special characters
    if sum(1 for c in roll if c.isalnum()) < len(roll) * 0.7:
        return False
    
    return True

def _extract_roll_numbers_from_pdf(saved_pdf_path: str) -> list:
    """Extract roll numbers from PDF using multiple methods"""
    if not os.path.exists(saved_pdf_path):
        print(f"PDF file not found: {saved_pdf_path}")
        return []
    
    roll_numbers: list[str] = []
    
    # 1) Try native text extraction first (works for text-based PDFs)
    try:
        text = _extract_text_with_pypdf(saved_pdf_path)
        if text and text.strip():
            print(f"Extracted {len(text)} characters from PDF")
            print(f"First 500 characters of extracted text: {text[:500]}")
            
            try:
                # Try multiple regex patterns for different roll number formats
                patterns = [
                    # Common roll number patterns
                    r"\b[0-9]{2}[A-Z]{2}[0-9]{4}\b",  # Format: 22CS1234
                    r"\b[A-Z]{2}[0-9]{4,6}\b",  # Format: CS1234, CS123456
                    r"\b[0-9]{4}[A-Z]{2}[0-9]{2}\b",  # Format: 2024CS01
                    r"\b[A-Z]{3}[0-9]{4,6}\b",  # Format: CSE1234
                    r"\b[0-9]{2}[A-Z]{3}[0-9]{3}\b",  # Format: 22CSE123
                    
                    # More flexible patterns
                    r"\b[0-9]{6,10}\b",  # Numeric only (6-10 digits)
                    r"\b[A-Z0-9]{6,12}\b",  # Alphanumeric (6-12 chars)
                    r"\b[0-9]{2,4}[A-Z]{1,3}[0-9]{2,6}\b",  # Flexible format
                    r"\b[A-Z]{1,3}[0-9]{3,8}\b",  # Letter(s) + numbers
                    
                    # Patterns without word boundaries (in case of formatting issues)
                    r"(?<![A-Z0-9])[0-9]{2}[A-Z]{2}[0-9]{4}(?![A-Z0-9])",  # 22CS1234
                    r"(?<![A-Z0-9])[A-Z]{2}[0-9]{4,6}(?![A-Z0-9])",  # CS1234
                    r"(?<![A-Z0-9])[0-9]{6,8}(?![A-Z0-9])",  # Numeric
                ]
                
                for pattern in patterns:
                    matches = re.findall(pattern, text.upper())
                    if matches:
                        # Validate matches before adding
                        valid_matches = [m for m in matches if _validate_roll_number(m)]
                        if valid_matches:
                            roll_numbers.extend(valid_matches)
                            print(f"Found {len(valid_matches)} valid roll numbers with pattern: {pattern}")
                            print(f"Sample matches: {valid_matches[:5]}")
                            break
                        
            except Exception as e:
                print(f"Regex extraction failed: {e}")
                roll_numbers = []
                
            if roll_numbers:
                # Deduplicate preserving order
                seen = set()
                unique = []
                for r in roll_numbers:
                    if r not in seen:
                        seen.add(r)
                        unique.append(r)
                print(f"Returning {len(unique)} unique roll numbers from text extraction")
                return unique
        else:
            print("No text extracted from PDF")
    except Exception as e:
        print(f"Text extraction failed: {e}")

    # 2) Fallback to OCR using pdf2image + pytesseract (requires Poppler and Tesseract)
    try:
        from pdf2image import convert_from_path
        import pytesseract

        print("Attempting OCR extraction...")
        
        # Support setting Tesseract and Poppler paths via environment (useful on Windows)
        tesseract_path = os.environ.get('TESSERACT_PATH')
        if tesseract_path:
            try:
                pytesseract.pytesseract.tesseract_cmd = tesseract_path
                print(f"Using custom Tesseract path: {tesseract_path}")
            except Exception as e:
                print(f"Failed to set Tesseract path: {e}")
                
        poppler_path = os.environ.get('POPPLER_PATH')
        if poppler_path:
            print(f"Using custom Poppler path: {poppler_path}")

        # Convert PDF to images
        pages = convert_from_path(saved_pdf_path, 300, poppler_path=poppler_path) if poppler_path else convert_from_path(saved_pdf_path, 300)
        print(f"Converted PDF to {len(pages)} images")
        
        roll_numbers = []
        for i, page in enumerate(pages):
            try:
                text = pytesseract.image_to_string(page)
                if text.strip():
                    # Try multiple patterns for OCR text
                    patterns = [
                        # Common roll number patterns for OCR
                        r"\b[0-9]{2}[A-Z]{2}[0-9]{4}\b",  # Format: 22CS1234
                        r"\b[A-Z]{2}[0-9]{4,6}\b",  # Format: CS1234, CS123456
                        r"\b[0-9]{4}[A-Z]{2}[0-9]{2}\b",  # Format: 2024CS01
                        r"\b[A-Z]{3}[0-9]{4,6}\b",  # Format: CSE1234
                        r"\b[0-9]{2}[A-Z]{3}[0-9]{3}\b",  # Format: 22CSE123
                        
                        # More flexible patterns for OCR
                        r"\b[0-9]{6,10}\b",  # Numeric only
                        r"\b[A-Z0-9]{6,12}\b",  # Alphanumeric
                        r"\b[0-9]{2,4}[A-Z]{1,3}[0-9]{2,6}\b",  # Flexible format
                        r"\b[A-Z]{1,3}[0-9]{3,8}\b",  # Letter(s) + numbers
                        
                        # OCR-specific patterns (less strict boundaries)
                        r"(?<![A-Z0-9])[0-9]{2}[A-Z]{2}[0-9]{4}(?![A-Z0-9])",  # 22CS1234
                        r"(?<![A-Z0-9])[A-Z]{2}[0-9]{4,6}(?![A-Z0-9])",  # CS1234
                        r"(?<![A-Z0-9])[0-9]{6,8}(?![A-Z0-9])",  # Numeric
                    ]
                    
                    for pattern in patterns:
                        matches = re.findall(pattern, text.upper())
                        if matches:
                            # Validate matches before adding
                            valid_matches = [m for m in matches if _validate_roll_number(m)]
                            if valid_matches:
                                roll_numbers.extend(valid_matches)
                                print(f"Page {i+1}: Found {len(valid_matches)} valid roll numbers with pattern: {pattern}")
                                break
            except Exception as e:
                print(f"OCR failed on page {i+1}: {e}")
                continue

        if roll_numbers:
            # Deduplicate preserving order
            seen = set()
            unique = []
            for r in roll_numbers:
                if r not in seen:
                    seen.add(r)
                    unique.append(r)
            print(f"OCR extraction returned {len(unique)} unique roll numbers")
            return unique
        else:
            print("No roll numbers found via OCR")
            
    except ImportError as e:
        print(f"OCR dependencies not available: {e}")
        print("Install with: pip install pdf2image pytesseract")
    except Exception as e:
        print(f"OCR extraction failed: {e}")
        print("Ensure Poppler is installed and in PATH or set POPPLER_PATH")
        print("Ensure Tesseract is installed or set TESSERACT_PATH")
    
    # 3) Final fallback: Try to extract any alphanumeric sequences that might be roll numbers
    try:
        print("Attempting basic text extraction fallback...")
        # Try to read the PDF as a binary file and look for patterns
        with open(saved_pdf_path, 'rb') as f:
            content = f.read().decode('utf-8', errors='ignore')
            
        # Look for common roll number patterns in raw content
        patterns = [
            # Common roll number patterns
            r"[0-9]{2}[A-Z]{2}[0-9]{4}",  # 22CS1234
            r"[A-Z]{2}[0-9]{4,6}",  # CS1234, CS123456
            r"[0-9]{4}[A-Z]{2}[0-9]{2}",  # 2024CS01
            r"[A-Z]{3}[0-9]{4,6}",  # CSE1234
            r"[0-9]{2}[A-Z]{3}[0-9]{3}",  # 22CSE123
            
            # More flexible patterns
            r"(?<![0-9])[0-9]{6,8}(?![0-9])",  # Numeric (6-8 digits, not part of longer numbers)
            r"[A-Z0-9]{6,10}(?![A-Z0-9])",  # Alphanumeric (6-10 chars, not part of longer strings)
            r"[0-9]{2,4}[A-Z]{1,3}[0-9]{2,6}",  # Flexible format
            r"[A-Z]{1,3}[0-9]{3,8}",  # Letter(s) + numbers
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content.upper())
            if matches:
                # Filter out obviously invalid roll numbers
                valid_rolls = [roll for roll in matches if _validate_roll_number(roll)]
                
                if valid_rolls:
                    roll_numbers = list(set(valid_rolls))  # Remove duplicates
                    print(f"Fallback extraction found {len(roll_numbers)} valid roll numbers")
                    return roll_numbers
                
    except Exception as e:
        print(f"Fallback extraction failed: {e}")
    
    print("No roll numbers could be extracted from PDF")
    return []

@app.route('/admin/results', methods=['GET', 'POST'])
def admin_results():
    if 'username' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    if request.method == 'POST':
        file = request.files.get('pdf')
        title = request.form.get('title') or 'Exam Results'
        rolls_text = (request.form.get('rolls_text') or '').strip()
        if not file or not file.filename.lower().endswith('.pdf'):
            flash('Please upload a PDF file')
            return redirect(url_for('admin_results'))
        # Save
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], f"results_{int(time.time())}.pdf")
        file.save(save_path)
        
        # Debug: Print file info
        print(f"Uploaded PDF: {file.filename}")
        print(f"Saved to: {save_path}")
        print(f"File size: {os.path.getsize(save_path)} bytes")
        
        # Extract
        rolls = _extract_roll_numbers_from_pdf(save_path)
        
        # Manual fallback/merge from textarea
        if rolls_text:
            manual = re.findall(r"\b[0-9A-Za-z]{6,20}\b", rolls_text.upper())
            merged = []
            seen = set()
            for r in (rolls or []) + manual:
                if r not in seen:
                    seen.add(r)
                    merged.append(r)
            rolls = merged
        
        rec = ResultDeclaration(title=title, pdf_path=save_path, passed_rolls=json.dumps(rolls), uploaded_by=session['username'])
        db.session.add(rec)
        db.session.commit()
        if len(rolls) == 0:
            flash("Uploaded. 0 roll numbers extracted. For scanned PDFs, install Poppler and Tesseract, and set POPPLER_PATH/TESSERACT_PATH. For text PDFs, install 'pypdf' or 'pdfminer.six'.")
        else:
            flash(f"Uploaded. {len(rolls)} roll numbers extracted: {', '.join(rolls[:10])}{'...' if len(rolls) > 10 else ''}")
        return redirect(url_for('admin_results'))
    # GET list
    results = db.session.execute(db.select(ResultDeclaration).order_by(ResultDeclaration.uploaded_at.desc())).scalars().all()
    return render_template('admin_results.html', results=results)

@app.route('/test_pdf_extraction', methods=['GET', 'POST'])
def test_pdf_extraction():
    """Public test endpoint for PDF extraction (no auth required)"""
    if request.method == 'POST':
        file = request.files.get('pdf')
        if not file or not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Please upload a PDF file'}), 400
        
        # Save temporarily
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], f"test_{int(time.time())}.pdf")
        file.save(save_path)
        
        try:
            # Extract text first
            text = _extract_text_with_pypdf(save_path)
            
            # Extract roll numbers
            rolls = _extract_roll_numbers_from_pdf(save_path)
            
            # Clean up temp file
            os.remove(save_path)
            
            return jsonify({
                'filename': file.filename,
                'file_size': os.path.getsize(save_path) if os.path.exists(save_path) else 0,
                'extracted_text_length': len(text),
                'extracted_text_preview': text[:1000] if text else "No text extracted",
                'extracted_rolls': rolls,
                'roll_count': len(rolls)
            })
        except Exception as e:
            # Clean up temp file
            if os.path.exists(save_path):
                os.remove(save_path)
            return jsonify({'error': str(e)}), 500
    
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Test PDF Extraction</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            .form-group { margin: 10px 0; }
            label { display: block; margin-bottom: 5px; }
            input[type="file"] { margin-bottom: 10px; }
            button { padding: 10px 20px; background: #007bff; color: white; border: none; cursor: pointer; }
            .result { margin-top: 20px; padding: 15px; background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 5px; }
            .error { background: #f8d7da; border-color: #f5c6cb; color: #721c24; }
        </style>
    </head>
    <body>
        <h1>Test PDF Roll Number Extraction</h1>
        <p>Upload a PDF file to test roll number extraction. This is a public endpoint for testing purposes.</p>
        <form method="POST" enctype="multipart/form-data">
            <div class="form-group">
                <label for="pdf">Select PDF file:</label>
                <input type="file" id="pdf" name="pdf" accept="application/pdf" required>
            </div>
            <button type="submit">Test Extraction</button>
        </form>
        <div id="result"></div>
        
        <script>
            document.querySelector('form').addEventListener('submit', async function(e) {
                e.preventDefault();
                const formData = new FormData(this);
                const resultDiv = document.getElementById('result');
                resultDiv.innerHTML = '<p>Processing...</p>';
                
                try {
                    const response = await fetch('/test_pdf_extraction', {
                        method: 'POST',
                        body: formData
                    });
                    const data = await response.json();
                    
                    if (data.error) {
                        resultDiv.innerHTML = `<div class="result error"><h3>Error:</h3><p>${data.error}</p></div>`;
                    } else {
                        resultDiv.innerHTML = `
                            <div class="result">
                                <h3>Extraction Results:</h3>
                                <p><strong>Filename:</strong> ${data.filename}</p>
                                <p><strong>File Size:</strong> ${data.file_size} bytes</p>
                                <p><strong>Extracted Text Length:</strong> ${data.extracted_text_length} characters</p>
                                <p><strong>Roll Numbers Found:</strong> ${data.roll_count}</p>
                                <p><strong>Roll Numbers:</strong> ${data.extracted_rolls.join(', ')}</p>
                                <h4>Extracted Text Preview:</h4>
                                <pre style="background: white; padding: 10px; border: 1px solid #ccc; max-height: 300px; overflow-y: auto;">${data.extracted_text_preview}</pre>
                            </div>
                        `;
                    }
                } catch (error) {
                    resultDiv.innerHTML = `<div class="result error"><h3>Network Error:</h3><p>${error.message}</p></div>`;
                }
            });
        </script>
    </body>
    </html>
    '''

@app.route('/admin/debug_pdf', methods=['GET', 'POST'])
def debug_pdf():
    """Debug endpoint to test PDF extraction"""
    if 'username' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        file = request.files.get('pdf')
        if not file or not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Please upload a PDF file'}), 400
        
        # Save temporarily
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], f"debug_{int(time.time())}.pdf")
        file.save(save_path)
        
        try:
            # Extract text first
            text = _extract_text_with_pypdf(save_path)
            
            # Extract roll numbers
            rolls = _extract_roll_numbers_from_pdf(save_path)
            
            # Clean up temp file
            os.remove(save_path)
            
            return jsonify({
                'filename': file.filename,
                'file_size': os.path.getsize(save_path) if os.path.exists(save_path) else 0,
                'extracted_text_length': len(text),
                'extracted_text_preview': text[:1000] if text else "No text extracted",
                'extracted_rolls': rolls,
                'roll_count': len(rolls)
            })
        except Exception as e:
            # Clean up temp file
            if os.path.exists(save_path):
                os.remove(save_path)
            return jsonify({'error': str(e)}), 500
    
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Debug PDF Extraction</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            .form-group { margin: 10px 0; }
            label { display: block; margin-bottom: 5px; }
            input[type="file"] { margin-bottom: 10px; }
            button { padding: 10px 20px; background: #007bff; color: white; border: none; cursor: pointer; }
            .result { margin-top: 20px; padding: 15px; background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 5px; }
            .error { background: #f8d7da; border-color: #f5c6cb; color: #721c24; }
        </style>
    </head>
    <body>
        <h1>Debug PDF Extraction</h1>
        <form method="POST" enctype="multipart/form-data">
            <div class="form-group">
                <label for="pdf">Select PDF file:</label>
                <input type="file" id="pdf" name="pdf" accept="application/pdf" required>
            </div>
            <button type="submit">Test Extraction</button>
        </form>
        <div id="result"></div>
        
        <script>
            document.querySelector('form').addEventListener('submit', async function(e) {
                e.preventDefault();
                const formData = new FormData(this);
                const resultDiv = document.getElementById('result');
                resultDiv.innerHTML = '<p>Processing...</p>';
                
                try {
                    const response = await fetch('/admin/debug_pdf', {
                        method: 'POST',
                        body: formData
                    });
                    const data = await response.json();
                    
                    if (data.error) {
                        resultDiv.innerHTML = `<div class="result error"><h3>Error:</h3><p>${data.error}</p></div>`;
                    } else {
                        resultDiv.innerHTML = `
                            <div class="result">
                                <h3>Extraction Results:</h3>
                                <p><strong>Filename:</strong> ${data.filename}</p>
                                <p><strong>File Size:</strong> ${data.file_size} bytes</p>
                                <p><strong>Extracted Text Length:</strong> ${data.extracted_text_length} characters</p>
                                <p><strong>Roll Numbers Found:</strong> ${data.roll_count}</p>
                                <p><strong>Roll Numbers:</strong> ${data.extracted_rolls.join(', ')}</p>
                                <h4>Extracted Text Preview:</h4>
                                <pre style="background: white; padding: 10px; border: 1px solid #ccc; max-height: 300px; overflow-y: auto;">${data.extracted_text_preview}</pre>
                            </div>
                        `;
                    }
                } catch (error) {
                    resultDiv.innerHTML = `<div class="result error"><h3>Network Error:</h3><p>${error.message}</p></div>`;
                }
            });
        </script>
    </body>
    </html>
    '''

def _latest_passed_set() -> set:
    rec = db.session.execute(db.select(ResultDeclaration).order_by(ResultDeclaration.uploaded_at.desc())).scalars().first()
    if not rec or not rec.passed_rolls:
        return set()

@app.route('/results', methods=['GET', 'POST'])
def student_results_view():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    if session.get('role') not in ['student', 'teacher', 'admin']:
        flash("Access denied")
        return redirect(url_for('access'))

    # Fetch latest 5 declarations for dropdown/list
    recent_results = db.session.execute(
        db.select(ResultDeclaration).order_by(ResultDeclaration.uploaded_at.desc()).limit(5)
    ).scalars().all()
    latest = recent_results[0] if recent_results else None

    result = None
    checked_roll = None
    selected_id = None
    selected_result = latest
    if request.method == 'POST':
        checked_roll = (request.form.get('roll_no') or '').strip().upper()
        try:
            selected_id = int(request.form.get('result_id') or 0)
        except Exception:
            selected_id = 0
        if selected_id:
            selected_result = db.session.execute(db.select(ResultDeclaration).filter_by(id=selected_id)).scalar() or latest
        else:
            selected_result = latest

        if checked_roll and selected_result and selected_result.passed_rolls:
            try:
                passed = set(json.loads(selected_result.passed_rolls))
            except Exception:
                passed = set()
            result = 'passed' if checked_roll in passed else 'not_listed'

    return render_template(
        'student_results.html',
        latest=latest,
        recent_results=recent_results,
        selected_result=selected_result,
        selected_id=selected_id,
        checked_roll=checked_roll,
        result=result
    )

@app.route('/results/download/<int:result_id>')
def download_result_pdf(result_id: int):
    rec = db.session.execute(db.select(ResultDeclaration).filter_by(id=result_id)).scalar()
    if not rec:
        abort(404)
    try:
        return send_file(rec.pdf_path, as_attachment=True)
    except Exception:
        abort(404)
    try:
        return set(json.loads(rec.passed_rolls))
    except Exception:
        return set()

def check_file_accessibility(file_path):
    """Check if a file exists and is accessible (not locked by another application)"""
    try:
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}"
        
        # Try to open the file in read-write mode to check if it's locked
        with open(file_path, 'r+b') as f:
            pass
        return True, "File is accessible"
    except (PermissionError, OSError) as e:
        return False, f"File is locked or inaccessible. Please close Excel or any other application that might be using this file. Error: {str(e)}"
    except Exception as e:
        return False, f"Unexpected error accessing file: {str(e)}"

def read_nfc_card():
    try:
        with nfc.ContactlessFrontend('usb') as clf:
            tag = clf.connect(rdwr={'on-connect': lambda tag: False})
            if tag.ndef:
                record = tag.ndef.message.records[0]
                text = record.text
                return text  # JSON string from NFC card
            else:
                return None
    except Exception as e:
        print(f"NFC read failed: {e}")
        return None

@app.route('/validate_nfc_payload', methods=['POST'])
def validate_nfc_payload():
    try:
        data = request.get_json() or {}
        nfc_text = data.get('nfc_data')
        if not nfc_text:
            return jsonify({'success': False, 'message': 'Missing NFC data'}), 400
        try:
            payload = json.loads(nfc_text)
        except Exception:
            return jsonify({'success': False, 'message': 'Invalid NFC JSON'}), 400

        card_id = payload.get('id') or payload.get('card_id')
        username_on_card = payload.get('name')
        if not card_id:
            return jsonify({'success': False, 'message': 'Card ID missing'}), 400

        card = db.session.execute(db.select(NfcCard).filter_by(card_id=card_id)).scalar()
        if not card:
            return jsonify({'success': False, 'message': 'Card not registered'}), 403
        if card.status != 'active':
            return jsonify({'success': False, 'message': 'Card is inactive'}), 403
        if username_on_card and card.username != username_on_card:
            return jsonify({'success': False, 'message': 'Card owner mismatch'}), 403

        # Normalize roll number and return sanitized payload
        if 'roll' in payload and 'roll_no' not in payload:
            payload['roll_no'] = payload.get('roll')
        payload['status'] = 'active'
        return jsonify({'success': True, 'nfc_data': json.dumps(payload)})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Validation error: {str(e)}'}), 500

def write_nfc_card(json_text):
    try:
        with nfc.ContactlessFrontend('usb') as clf:
            def on_connect(tag):
                if tag.ndef:
                    record = nfc.ndef.TextRecord(json_text)
                    tag.ndef.message = nfc.ndef.Message(record)
                    return False
                return False
            clf.connect(rdwr={'on-connect': on_connect})
            return True
    except Exception as e:
        print(f"NFC write failed: {e}")
        return False

# Gemini API Helper Functions
def generate_exam_id():
    """Generate a unique exam ID"""
    timestamp = str(int(time.time()))
    random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"EXAM_{timestamp}_{random_str}"

def generate_questions_with_gemini(subject, syllabus, question_format, total_questions, total_marks, difficulty_level, format_description=''):
    """Generate questions using Gemini API via REST helper"""
    try:
        print(f"Generating questions for subject: {subject}")
        print(f"Syllabus: {syllabus}")
        print(f"Format: {question_format}, Questions: {total_questions}, Marks: {total_marks}")
        
        # Calculate marks per question
        marks_per_question = total_marks // total_questions
        remaining_marks = total_marks % total_questions
        
        # Create prompt based on question format and difficulty
        difficulty_instruction = {
            'easy': 'Generate questions of easy difficulty level suitable for beginners.',
            'medium': 'Generate questions of medium difficulty level suitable for intermediate learners.',
            'hard': 'Generate questions of hard difficulty level suitable for advanced learners.',
            'mixed': 'Generate questions with mixed difficulty levels (easy, medium, and hard).'
        }.get(difficulty_level, 'Generate questions of medium difficulty level.')
        
        if question_format == 'multiple_choice':
            format_instruction = f"Generate {total_questions} multiple choice questions with 4 options each (A, B, C, D) and mark the correct answer. Each question should be worth {marks_per_question} marks."
        elif question_format == 'descriptive':
            format_instruction = f"Generate {total_questions} descriptive/long answer questions that require detailed explanations. Each question should be worth {marks_per_question} marks."
        elif question_format == 'section_based':
            format_instruction = f"""Generate questions based on the following custom format description:

{format_description}

IMPORTANT INSTRUCTIONS FOR SECTION-BASED FORMAT:
1. Parse the format description carefully to understand the exact structure
2. Create sections with the exact names specified (e.g., "Section A", "Part I", etc.)
3. Generate the exact number of questions for each section as specified
4. Use the exact marks per question as specified in the format
5. For "short answer" questions, use type "short_answer" 
6. For "long answer" or "descriptive" questions, use type "descriptive"
7. For "multiple choice" questions, use type "multiple_choice"
8. Follow the format description word-for-word
9. Ensure all sections and questions match the specified structure exactly
10. Generate REAL questions based on the syllabus content, NOT sample or generic questions
11. Each question must be directly related to the syllabus topics provided
12. Questions should test actual understanding of the subject matter
13. Use specific topics from the syllabus to create meaningful questions
14. Do NOT use generic placeholders like "Question 1 about [subject]"
15. Create questions that test knowledge of the specific syllabus topics"""
        else:  # mixed
            mcq_count = int(total_questions * 0.6)
            desc_count = total_questions - mcq_count
            mcq_marks = int(total_marks * 0.4)
            desc_marks = total_marks - mcq_marks
            format_instruction = f"Generate {mcq_count} multiple choice questions (worth {mcq_marks//mcq_count} marks each) and {desc_count} descriptive questions (worth {desc_marks//desc_count} marks each)."
        
        prompt = f"""You are an expert question paper generator. Generate questions based on the following specifications:

Subject: {subject}
Syllabus: {syllabus}
Total Questions: {total_questions}
Total Marks: {total_marks}
Difficulty Level: {difficulty_level}

{format_instruction}
{difficulty_instruction}

ABSOLUTELY CRITICAL REQUIREMENTS:
1. Generate REAL, SPECIFIC questions based on the syllabus content: {syllabus}
2. NEVER use generic placeholders like "Question 1 about [subject]" or "Descriptive question 1 about WT"
3. Each question MUST be about specific topics from the syllabus
4. Use the exact syllabus topics to create meaningful, detailed questions
5. Questions should test actual knowledge of the specific syllabus topics
6. Create questions that are specific to the syllabus content, not generic placeholders
7. Use the exact topics from the syllabus to formulate questions

EXAMPLES OF WHAT TO DO:
- If syllabus mentions "HTML", create questions like "Explain the difference between HTML4 and HTML5 and list three new features in HTML5"
- If syllabus mentions "CSS", create questions like "Describe the CSS box model and explain how margin, padding, and border properties work together"
- If syllabus mentions "JavaScript", create questions like "Write a JavaScript function to validate an email address and explain how it works"

EXAMPLES OF WHAT NOT TO DO:
- "Descriptive question 1 about Web Technologies" ❌
- "Question 1 about HTML" ❌
- "Short answer question 1" ❌

IMPORTANT: You must respond with ONLY valid JSON. Do not include any explanations, markdown formatting, or additional text.

SYLLABUS TOPICS TO USE: {syllabus}

IMPORTANT: You MUST create questions based on these specific syllabus topics. Do NOT create generic questions. Use the exact topics mentioned above to formulate your questions. For example, if the syllabus mentions "HTML, CSS, JavaScript", create questions about HTML, CSS, and JavaScript specifically, not generic web development questions.

For regular formats, use this exact JSON structure:
{{
    "questions": [
        {{
            "id": 1,
            "type": "multiple_choice",
            "question": "What is HTML used for in web development?",
            "options": ["Styling web pages", "Creating web page structure", "Adding interactivity", "Database management"],
            "correct_answer": "Creating web page structure",
            "marks": {marks_per_question}
        }},
        {{
            "id": 2,
            "type": "descriptive",
            "question": "Explain the role of CSS in web development and its relationship with HTML.",
            "marks": {marks_per_question}
        }}
    ]
}}

For section-based format, use this exact JSON structure:
{{
    "sections": [
        {{
            "section_name": "Section A",
            "section_type": "short_answer",
            "questions": [
                {{
                    "id": 1,
                    "type": "short_answer",
                    "question": "What is HTML and what does it stand for?",
                    "marks": 2
                }}
            ]
        }},
        {{
            "section_name": "Section B", 
            "section_type": "descriptive",
            "questions": [
                {{
                    "id": 2,
                    "type": "descriptive",
                    "question": "Explain the difference between HTML and CSS and how they work together in web development.",
                    "marks": 5
                }}
            ]
        }}
    ]
}}

Requirements:
1. Questions must be relevant to the syllabus content
2. Difficulty level: {difficulty_level}
3. Questions should be clear and unambiguous
4. Cover different topics from the syllabus
5. Test understanding, not just memorization
6. For multiple choice questions, provide exactly 4 options (A, B, C, D format)
7. For descriptive questions, do not include options or correct_answer fields
8. For short answer questions, do not include options or correct_answer fields
9. Ensure all required fields are present in each question
10. For section-based format, follow the format description exactly
11. Use appropriate question types: "multiple_choice", "short_answer", "descriptive"
12. Generate questions that match the specified marks and difficulty
13. CRITICAL: Generate REAL questions based on the syllabus, NOT sample or generic questions
14. Each question must test actual knowledge of the subject matter provided in the syllabus
15. Questions should be specific to the topics mentioned in the syllabus content
16. Use the exact syllabus topics to create meaningful, specific questions
17. Do NOT use generic placeholders or sample questions
18. Create questions that test understanding of the specific syllabus topics provided

Respond with ONLY the JSON, no other text."""
        
        # Use REST-based helper for better compatibility
        response_text = gemini_generate_text(prompt).strip()
        
        # Remove markdown code blocks if present
        if response_text.startswith('```json'):
            response_text = response_text[7:]
        if response_text.startswith('```'):
            response_text = response_text[3:]
        if response_text.endswith('```'):
            response_text = response_text[:-3]
        
        response_text = response_text.strip()
        
        print(f"Gemini Response: {response_text[:500]}...")  # Log first 500 chars for debugging
        
        # Parse JSON
        questions_data = json.loads(response_text)
        
        # Log the parsed data for debugging
        print(f"Parsed questions data: {json.dumps(questions_data, indent=2)[:1000]}...")
        
        # Validate the response structure
        if question_format == 'section_based':
            if 'sections' not in questions_data:
                print("Error: No 'sections' key in response for section-based format")
                return None
            
            # Validate each section
            for section in questions_data['sections']:
                if 'section_name' not in section or 'questions' not in section:
                    print(f"Error: Invalid section structure: {section}")
                    return None
        else:
            if 'questions' not in questions_data:
                print("Error: No 'questions' key in response")
                return None
        
        print(f"Successfully generated questions: {len(questions_data.get('questions', questions_data.get('sections', [])))}")
        return questions_data
        
    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
        print(f"Response text: {response_text if 'response_text' in locals() else 'No response'}")
        
        # Try to fix common JSON issues
        try:
            # Remove any text before the first { and after the last }
            start_idx = response_text.find('{')
            end_idx = response_text.rfind('}')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                cleaned_text = response_text[start_idx:end_idx+1]
                questions_data = json.loads(cleaned_text)
                print("Successfully parsed after cleaning JSON")
                return questions_data
        except:
            pass
        
        return None
    except Exception as e:
        print(f"Error generating questions: {e}")
        print(f"Response text: {response_text if 'response_text' in locals() else 'No response'}")
        return None

def generate_fallback_questions(subject, total_questions, total_marks, question_format):
    """Generate simple fallback questions when AI fails"""
    try:
        marks_per_question = total_marks // total_questions
        
        if question_format == 'section_based':
            # Create a simple section-based structure
            questions_data = {
                "sections": [
                    {
                        "section_name": "Section A",
                        "section_type": "short_answer",
                        "questions": []
                    },
                    {
                        "section_name": "Section B", 
                        "section_type": "descriptive",
                        "questions": []
                    }
                ]
            }
            
            # Distribute questions between sections
            short_answer_count = total_questions // 2
            descriptive_count = total_questions - short_answer_count
            
            # Add short answer questions
            for i in range(short_answer_count):
                question = {
                    "id": i + 1,
                    "type": "short_answer",
                    "question": f"Short answer question {i + 1} about {subject}",
                    "marks": marks_per_question
                }
                questions_data["sections"][0]["questions"].append(question)
            
            # Add descriptive questions
            for i in range(descriptive_count):
                question = {
                    "id": short_answer_count + i + 1,
                    "type": "descriptive",
                    "question": f"Descriptive question {i + 1} about {subject}",
                    "marks": marks_per_question
                }
                questions_data["sections"][1]["questions"].append(question)
        else:
            # Create regular questions structure
            questions_data = {"questions": []}
            
            for i in range(total_questions):
                if question_format == 'multiple_choice':
                    question = {
                        "id": i + 1,
                        "type": "multiple_choice",
                        "question": f"Sample multiple choice question {i + 1} about {subject}",
                        "options": ["Option A", "Option B", "Option C", "Option D"],
                        "correct_answer": "Option A",
                        "marks": marks_per_question
                    }
                else:
                    question = {
                        "id": i + 1,
                        "type": "descriptive",
                        "question": f"Sample descriptive question {i + 1} about {subject}",
                        "marks": marks_per_question
                    }
                questions_data["questions"].append(question)
        
        print(f"Generated {total_questions} fallback questions")
        return questions_data
        
    except Exception as e:
        print(f"Error generating fallback questions: {e}")
        return None

def evaluate_single_question_with_gemini(question, student_answer, subject):
    """Evaluate a single question and answer using Gemini API via REST helper"""
    try:
        print(f"Evaluating question {question.get('id', 'unknown')} for subject: {subject}")
        print(f"Student answer: {student_answer[:100]}...")
        
        prompt = f"""You are an expert examiner. Score the student's answer for this question based on how well they answered it.

Subject: {subject}

Question:
{json.dumps(question, indent=2)}

Student's Answer:
{student_answer}

SCORING INSTRUCTIONS:
You need to give a score from 0 to {question.get('marks', 5)} marks based on the quality of the student's answer.

SCORING CRITERIA:
1. For multiple choice questions: Full marks if correct, 0 if incorrect
2. For short answer questions: Score based on accuracy, completeness, and relevance
3. For descriptive questions: Score based on understanding, depth of explanation, and accuracy
4. Award partial marks for partially correct answers
5. Consider the difficulty level and expected depth of answer

IMPORTANT: You must respond with ONLY valid JSON in the following format:
{{
    "question_id": {question.get('id', 1)},
    "question_type": "{question.get('type', 'unknown')}",
    "marks_awarded": 0,
    "total_marks": {question.get('marks', 5)},
    "feedback": "Detailed feedback on the answer quality and why this score was given",
    "suggestions": "Specific suggestions for improvement",
    "correct_answer": "Expected correct answer for reference",
    "is_correct": false,
    "accuracy_percentage": 0.0
}}

SCORING GUIDELINES:
- Give the exact number of marks the student deserves (0 to {question.get('marks', 5)})
- Be fair and consistent in your scoring
- Consider partial credit for partially correct answers
- Provide detailed feedback explaining your scoring decision
- Focus on what the student wrote, not what they should have written

Respond with ONLY the JSON, no other text."""
        
        # Use REST-based helper for better compatibility
        response_text = gemini_generate_text(prompt).strip()
        
        # Remove markdown code blocks if present
        if response_text.startswith('```json'):
            response_text = response_text[7:]
        if response_text.startswith('```'):
            response_text = response_text[3:]
        if response_text.endswith('```'):
            response_text = response_text[:-3]
        
        response_text = response_text.strip()
        
        print(f"Single Question Evaluation Response: {response_text[:300]}...")
        
        # Parse JSON
        evaluation_result = json.loads(response_text)
        
        print(f"Question {question.get('id', 'unknown')}: {evaluation_result.get('marks_awarded', 0)}/{evaluation_result.get('total_marks', 0)} marks")
        return evaluation_result
        
    except json.JSONDecodeError as e:
        print(f"JSON parsing error in single question evaluation: {e}")
        print(f"Response text: {response_text if 'response_text' in locals() else 'No response'}")
        
        # Try to fix common JSON issues
        try:
            start_idx = response_text.find('{')
            end_idx = response_text.rfind('}')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                cleaned_text = response_text[start_idx:end_idx+1]
                evaluation_result = json.loads(cleaned_text)
                print("Successfully parsed single question evaluation after cleaning JSON")
                return evaluation_result
        except:
            pass
        
        return None
    except Exception as e:
        print(f"Error evaluating single question: {e}")
        print(f"Response text: {response_text if 'response_text' in locals() else 'No response'}")
        return None

def evaluate_answers_with_gemini(questions, answers, subject):
    """Evaluate student answers using Gemini API - Individual question approach"""
    try:
        print(f"Starting individual question evaluation for {len(answers)} answers")
        
        # Handle different question formats
        if isinstance(questions, dict) and 'sections' in questions:
            # New format with sections - flatten questions for evaluation
            all_questions = []
            for section in questions['sections']:
                for question in section['questions']:
                    all_questions.append(question)
            questions_list = all_questions
        else:
            # Old format or direct questions
            questions_list = questions if isinstance(questions, list) else [questions]
        
        # Calculate total marks for ALL questions first
        total_marks = sum(question.get('marks', 0) for question in questions_list)
        print(f"Total marks for exam: {total_marks}")
        
        # Evaluate each question individually
        detailed_feedback = []
        obtained_marks = 0
        questions_attempted = 0
        correct_answers = 0
        
        for question in questions_list:
            question_id = str(question.get('id', ''))
            student_answer = answers.get(question_id, '')
            
            if student_answer.strip():  # Only evaluate if student provided an answer
                questions_attempted += 1
                
                # Evaluate single question
                single_result = evaluate_single_question_with_gemini(question, student_answer, subject)
                
                if single_result:
                    detailed_feedback.append(single_result)
                    obtained_marks += single_result.get('marks_awarded', 0)
                    
                    if single_result.get('is_correct', False):
                        correct_answers += 1
                else:
                    # Fallback if evaluation fails
                    detailed_feedback.append({
                        "question_id": question.get('id', 0),
                        "question_type": question.get('type', 'unknown'),
                        "marks_awarded": 0,
                        "total_marks": question.get('marks', 0),
                        "feedback": "Evaluation failed - no marks awarded",
                        "suggestions": "Please try again",
                        "correct_answer": "Not available",
                        "is_correct": False,
                        "accuracy_percentage": 0.0
                    })
            else:
                # Add entry for unanswered questions
                detailed_feedback.append({
                    "question_id": question.get('id', 0),
                    "question_type": question.get('type', 'unknown'),
                    "marks_awarded": 0,
                    "total_marks": question.get('marks', 0),
                    "feedback": "No answer provided",
                    "suggestions": "Please attempt all questions",
                    "correct_answer": "Not available",
                    "is_correct": False,
                    "accuracy_percentage": 0.0
                })
        
        # Calculate overall results
        percentage = (obtained_marks / total_marks * 100) if total_marks > 0 else 0
        total_questions = len(questions_list)
        accuracy = (correct_answers / total_questions * 100) if total_questions > 0 else 0
        
        # Determine grade
        if percentage >= 90:
            grade = "A+"
        elif percentage >= 80:
            grade = "A"
        elif percentage >= 70:
            grade = "B+"
        elif percentage >= 60:
            grade = "B"
        elif percentage >= 50:
            grade = "C+"
        elif percentage >= 40:
            grade = "C"
        elif percentage >= 30:
            grade = "D"
        else:
            grade = "F"
        
        evaluation_result = {
            "total_marks": total_marks,
            "obtained_marks": obtained_marks,
            "percentage": round(percentage, 2),
            "grade": grade,
            "total_questions": total_questions,
            "questions_attempted": questions_attempted,
            "correct_answers": correct_answers,
            "incorrect_answers": total_questions - correct_answers,
            "accuracy": round(accuracy, 2),
            "detailed_feedback": detailed_feedback
        }
        
        print(f"Individual evaluation completed: {obtained_marks}/{total_marks} marks ({percentage:.1f}%) - Grade: {grade}")
        return evaluation_result
        
    except Exception as e:
        print(f"Error in individual question evaluation: {e}")
        import traceback
        traceback.print_exc()
        return None


@app.route('/')
def index():
    return render_template('home.html')

@app.route('/signup',methods=['GET','POST'])
def signup():
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role')

    if not (username and email and password and role):
        return "All fields are required", 400

    existing_user = db.session.execute(
        db.select(Credentials).filter((Credentials.username == username) | (Credentials.email == email))
    ).scalar()

    if existing_user:
        flash("Username or email already exists")
    else:
        new_user = Credentials(username=username, email=email, password=password, role=role,credits=2000)
        db.session.add(new_user)
        db.session.commit()
        flash('Signup success')
    return render_template('login_signup.html')

@app.route('/login',methods=['GET','POST'])
def login():
    username = request.form.get('username')
    password = request.form.get('password')

    result = db.session.execute(
        db.select(Credentials).filter((Credentials.username == username) | (Credentials.email==username))
    ).first()

    if result:
        user = result[0]
        if user.password == password:
            session['username'] = user.username
            session['email']=user.email
            session['role'] = user.role
            flash("Login successful")
            if user.role=='student':
                return redirect(url_for('student'))
            elif user.role=='admin':
                return redirect(url_for('admin_dashboard'))
            elif user.role=='college':
                return redirect(url_for('college_dashboard'))
            elif user.role=='teacher':
                return redirect(url_for('teacher_dashboard'))
            elif user.role=='library':
                return redirect(url_for('library_dashboard'))
        else:
            flash("Incorrect password")
    else:
        flash("User not found")

    return redirect(url_for('access'))

@app.route('/student',methods=['GET','POST'])
def student():
    # Check if user is logged in
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    
    # Get user credentials
    user = db.session.execute(
        db.select(Credentials).filter(
            (Credentials.username == session['username']) & (Credentials.email == session['email'])
        )
    ).scalar()

    if not user:
        flash("User not found")
        return redirect(url_for('access'))

    # Get user's transaction history
    transactions = db.session.execute(
        db.select(Transactions).filter(
            (Transactions.from_user == session['username']) | 
            (Transactions.to_user == session['username'])
        ).order_by(Transactions.date.desc()).limit(10)
    ).scalars().all()

    # Get user's payment proof history
    payment_proofs = db.session.execute(
        db.select(PaymentProof).filter_by(username=session['username']).order_by(PaymentProof.submitted_at.desc())
    ).scalars().all()

    # Get user's college payment history
    college_payments = db.session.execute(
        db.select(CollegePayments).filter_by(sender_username=session['username']).order_by(CollegePayments.payment_date.desc())
    ).scalars().all()

    # Get user's refund request history
    refund_requests = db.session.execute(
        db.select(RefundRequest).filter_by(username=session['username']).order_by(RefundRequest.requested_at.desc())
    ).scalars().all()

    # Get user's NFC cards
    nfc_cards = db.session.execute(
        db.select(NfcCard).filter_by(username=session['username']).order_by(NfcCard.updated_at.desc())
    ).scalars().all()

    # Result status: check latest declaration
    latest_passed = _latest_passed_set()
    roll = None
    # Prefer NfcCard.roll_no if available
    try:
        card = db.session.execute(db.select(NfcCard).filter_by(username=session['username'])).scalars().first()
        if card and card.roll_no:
            roll = card.roll_no
    except Exception:
        pass
    # Fallback to any known roll in LibraryTransaction
    if not roll:
        lt = db.session.execute(db.select(LibraryTransaction).filter_by(student_username=session['username'])).scalars().first()
        if lt:
            roll = lt.student_roll_no

    pass_status = None
    if roll and latest_passed:
        pass_status = 'passed' if roll.upper() in latest_passed else 'not_listed'

    return render_template('student_dashboard.html', 
                         user_credits=user.credits,
                         user=user,
                         transactions=transactions,
                         payment_proofs=payment_proofs,
                         college_payments=college_payments,
                         refund_requests=refund_requests,
                         nfc_cards=nfc_cards,
                         result_status=pass_status,
                         roll_no=roll)


@app.route('/question-papers')
def question_papers_list():
    # Ensure only logged-in students can access
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    if session.get('role') != 'student':
        flash("Access denied. Students only.")
        return redirect(url_for('access'))

    allowed_extensions = {'.pdf', '.doc', '.docx', '.png', '.jpg', '.jpeg'}
    folder_path = app.config['QUESTION_PAPERS_FOLDER']
    try:
        all_files = os.listdir(folder_path)
    except FileNotFoundError:
        all_files = []

    files = []
    for name in sorted(all_files):
        _, ext = os.path.splitext(name)
        if ext.lower() in allowed_extensions and os.path.isfile(os.path.join(folder_path, name)):
            file_path = os.path.join(folder_path, name)
            size_bytes = os.path.getsize(file_path)
            files.append({
                'name': name,
                'size_kb': max(1, size_bytes // 1024)
            })

    return render_template('question_papers.html', files=files)


@app.route('/question-papers/download/<path:filename>')
def download_question_paper(filename):
    # Ensure only logged-in students can access
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    if session.get('role') != 'student':
        flash("Access denied. Students only.")
        return redirect(url_for('access'))

    folder_path = app.config['QUESTION_PAPERS_FOLDER']
    try:
        return send_from_directory(folder_path, filename, as_attachment=True)
    except FileNotFoundError:
        abort(404)



@app.route('/submit_payment', methods=['GET', 'POST'])
def submit_payment():
    # Check if user is logged in
    if 'username' not in session:
        flash("Please login first to submit payment proof")
        return redirect(url_for('access'))
    
    if request.method == 'POST':
        try:
            # Get form data
            proof_file = request.files.get('proof')
            amount = request.form.get('amount')
            note = request.form.get('note', '')
            
            # Validate required fields
            if not proof_file or not amount:
                flash("Please provide both proof file and amount")
                return redirect(url_for('student'))
            
            # Validate amount
            try:
                amount = int(amount)
                if amount <= 0:
                    flash("Amount must be greater than 0")
                    return redirect(url_for('student'))
            except (ValueError, TypeError):
                flash("Please enter a valid amount")
                return redirect(url_for('student'))
            
            # Validate file
            if proof_file.filename == '':
                flash("Please select a file")
                return redirect(url_for('student'))
            
            # Check file type (allow images and PDFs)
            allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx'}
            if not ('.' in proof_file.filename and 
                   proof_file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
                flash("Please upload a valid file type (images or PDF)")
                return redirect(url_for('student'))
            
            # Generate unique filename
            import uuid
            file_extension = proof_file.filename.rsplit('.', 1)[1].lower()
            unique_filename = f"{uuid.uuid4()}_{proof_file.filename}"
            
            # Save file to uploads directory
            upload_folder = 'static/uploads'
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)
            
            file_path = os.path.join(upload_folder, unique_filename)
            proof_file.save(file_path)
            
            # Save payment proof to database
            payment_proof = PaymentProof(
                username=session['username'],
                amount=amount,
                proof_file=unique_filename,
                note=note,
                status='pending'
            )
            
            db.session.add(payment_proof)
            db.session.commit()
            
            flash("Payment proof submitted successfully! It will be reviewed by admin.")
            return redirect(url_for('student'))
            
        except Exception as e:
            print(f"Error in submit_payment: {e}")
            flash("An error occurred while submitting payment proof")
            return redirect(url_for('student'))
    
    # GET request - redirect to student dashboard
    return redirect(url_for('student'))

@app.route('/save_nfc', methods=['POST'])
def save_nfc():
    data = request.get_json()
    nfc_text = data.get('nfc_data')
    nfc_text = json.loads(nfc_text)
    print(nfc_text['name'])
    print(f"NFC Data Received: {nfc_text}")
    return "NFC data received successfully", 200

@app.route('/proceed_payment',methods=['GET','POST'])
def proceed_payment():
    return render_template('nfc_payment.html')

@app.route('/nfc-payment', methods=['POST'])
def nfc_payment():
    try:
        data = request.get_json()
        if not data:
            return "Invalid request data", 400
            
        to_username = data.get('to_username')
        amount_str = data.get('amount')
        tx_type = data.get('type')
        
        # Validate required fields
        if not to_username or not amount_str or not tx_type:
            return "Missing required fields", 400
            
        # Validate amount
        try:
            amount = int(amount_str)
            if amount <= 0:
                return "Amount must be greater than 0", 400
        except (ValueError, TypeError):
            return "Invalid amount format", 400
        if tx_type == "mobile":
            try:
                nfc_json = json.loads(data['nfc_data'])
                sender_username = nfc_json.get('name')  # note: 'name' field as per card
                card_id = nfc_json.get('id') or nfc_json.get('card_id')
            except (json.JSONDecodeError, KeyError, TypeError):
                return "Invalid mobile NFC data", 400

        elif tx_type == "external":
            nfc_text = read_nfc_card()
            if not nfc_text:
                return "External NFC read failed", 500
            try:
                nfc_json = json.loads(nfc_text)
                sender_username = nfc_json.get('name')
                card_id = nfc_json.get('id') or nfc_json.get('card_id')
            except (json.JSONDecodeError, KeyError, TypeError):
                return "Invalid NFC content", 400
        else:
            return "Invalid reader type", 400

        if not sender_username or not to_username:
            return "Missing sender or recipient", 400

        # Validate card status
        if not card_id:
            return "Card ID missing in NFC data", 400
        card = db.session.execute(db.select(NfcCard).filter_by(card_id=card_id)).scalar()
        if not card:
            return "Card not registered", 403
        if card.username != sender_username:
            return "Card does not belong to this user", 403
        if card.status != 'active':
            return "Card is inactive. Please contact admin.", 403
        
        # Check if PIN is required for amounts over 500
        if amount > 500:
            provided_pin = data.get('pin')
            if not provided_pin:
                return "PIN required for payments over 500 credits", 400
            if not card.pin:
                return "Card PIN not set. Please set a PIN first.", 400
            if card.pin != provided_pin:
                return "Invalid PIN", 403

        # Check if user is logged in
        if 'username' not in session:
            return "Please login first", 401

        if session['username'] == to_username:
            return "invalid operation"
            
        # Fetch sender and receiver from DB
        to_username = to_username.strip()
        sender = db.session.execute(db.select(Credentials).filter_by(username=sender_username)).scalar()
        receiver = db.session.execute(db.select(Credentials).filter_by(username=to_username)).scalar()
        
        if not sender:
            return "Sender not found", 404

        if not receiver:
            return "Receiver not found", 404

        # Validate that the NFC card belongs to the logged-in user
        logged_in_username = session['username']
        if sender_username != logged_in_username:
            return f"Invalid card! You can only pay with your own card. Logged in as: {logged_in_username}, Card belongs to: {sender_username}", 403

        if sender.credits < amount:
            return "Insufficient balance", 403

        # Perform transaction
        sender.credits -= amount
        receiver.credits += amount

        # Save transaction
        transaction = Transactions(
            from_user=sender_username,
            to_user=to_username,
            amount=amount
        )
        db.session.add(transaction)
        db.session.commit()

        # Return success message
        return "Payment successful! Transaction completed."
        
    except Exception as e:
        print(f"Error in NFC payment: {e}")
        return "Internal server error", 500

@app.route('/college_payment')
def college_payment():
    # Check if user is logged in
    if 'username' not in session:
        flash("Please login first to access college payments")
        return redirect(url_for('access'))
    
    return render_template('college_payment.html')

@app.route('/college-nfc-payment', methods=['POST'])
def college_nfc_payment():
    try:
        data = request.get_json()
        if not data:
            return "Invalid request data", 400
            
        payment_data = data.get('paymentData', {})
        tx_type = data.get('type')
        
        # Validate required fields
        if not payment_data or not tx_type:
            return "Missing required fields", 400
            
        amount = payment_data.get('amount')
        fee_type = payment_data.get('feeType')
        student_name = payment_data.get('studentName')
        student_id = payment_data.get('studentId')
        
        # Validate amount
        try:
            amount = int(amount)
            if amount <= 0:
                return "Amount must be greater than 0", 400
        except (ValueError, TypeError):
            return "Invalid amount format", 400
            
        if tx_type == "mobile":
            try:
                nfc_json = json.loads(data['nfc_data'])
                sender_username = nfc_json.get('name')  # note: 'name' field as per card
                card_id = nfc_json.get('id') or nfc_json.get('card_id')
            except (json.JSONDecodeError, KeyError, TypeError):
                return "Invalid mobile NFC data", 400

        elif tx_type == "external":
            nfc_text = read_nfc_card()
            if not nfc_text:
                return "External NFC read failed", 500
            try:
                nfc_json = json.loads(nfc_text)
                sender_username = nfc_json.get('name')
                card_id = nfc_json.get('id') or nfc_json.get('card_id')
            except (json.JSONDecodeError, KeyError, TypeError):
                return "Invalid NFC content", 400
        else:
            return "Invalid reader type", 400

        if not sender_username:
            return "Missing sender information", 400

        # Validate card status
        if not card_id:
            return "Card ID missing in NFC data", 400
        card = db.session.execute(db.select(NfcCard).filter_by(card_id=card_id)).scalar()
        if not card:
            return "Card not registered", 403
        if card.username != sender_username:
            return "Card does not belong to this user", 403
        if card.status != 'active':
            return "Card is inactive. Please contact admin.", 403
        
        # Check if PIN is required for amounts over 500
        if amount > 500:
            provided_pin = data.get('pin')
            if not provided_pin:
                return "PIN required for payments over 500 credits", 400
            if not card.pin:
                return "Card PIN not set. Please set a PIN first.", 400
            if card.pin != provided_pin:
                return "Invalid PIN", 403

        # Fetch sender from DB
        sender = db.session.execute(db.select(Credentials).filter_by(username=sender_username)).scalar()

        if not sender:
            return "Sender not found", 404

        # Check if user is logged in
        if 'username' not in session:
            return "Please login first", 401

        # Validate that the NFC card belongs to the logged-in user
        logged_in_username = session['username']
        if sender_username != logged_in_username:
            return f"Invalid card! You can only pay with your own card. Logged in as: {logged_in_username}, Card belongs to: {sender_username}", 403

        if sender.credits < amount:
            return "Insufficient balance", 403

        # Fetch college account (user with role="college")
        college_account = db.session.execute(db.select(Credentials).filter_by(role="college")).scalar()
        
        if not college_account:
            return "College account not found", 404

        # Perform transaction - debit from sender, credit to college account
        sender.credits -= amount
        college_account.credits += amount

        # Save college payment to dedicated table
        college_payment = CollegePayments(
            student_name=student_name,
            student_id=student_id,
            student_email=payment_data.get('email', ''),
            student_phone=payment_data.get('phone', ''),
            department=payment_data.get('department', ''),
            fee_type=fee_type,
            amount=amount,
            semester=payment_data.get('semester', ''),
            academic_year=payment_data.get('academicYear', ''),
            payment_method='nfc',
            nfc_reader_type=tx_type,
            sender_username=sender_username,
            remarks=payment_data.get('remarks', ''),
            status='completed'
        )
        db.session.add(college_payment)
        db.session.commit()

        # Log the college payment details
        print(f"College Payment: {student_name} ({student_id}) paid ₹{amount} for {fee_type}")
        
        return "Payment successful! College fee payment completed."
        
    except Exception as e:
        print(f"Error in college NFC payment: {e}")
        return "Internal server error", 500

@app.route('/write')
def write():
    return render_template('write.html')

@app.route('/nfc/register', methods=['GET'])
def nfc_register_page():
    if 'username' not in session:
        return redirect(url_for('access'))
    return render_template('nfc_register.html')

@app.route('/nfc/cards', methods=['GET'])
def list_nfc_cards():
    if 'username' not in session:
        return redirect(url_for('access'))
    username = session['username']
    cards = db.session.execute(db.select(NfcCard).filter_by(username=username)).scalars().all()
    return render_template('nfc_cards.html', cards=cards)

@app.route('/nfc/cards/register', methods=['POST'])
def register_nfc_card():
    if 'username' not in session:
        return "Please login first", 401
    try:
        data = request.get_json() or {}
        full_name = data.get('name')
        email = data.get('email')
        roll_no = data.get('roll_no')
        tx_type = data.get('type')  # 'mobile' or 'external'

        if not (full_name and email and roll_no and tx_type):
            return "Missing required fields", 400

        # Build NFC payload
        card_id = data.get('card_id') or f"CARD-{uuid.uuid4().hex[:12].upper()}"
        nfc_payload = {
            'name': session['username'],
            'id': card_id,
            'full_name': full_name,
            'email': email,
            'roll_no': roll_no
        }
        json_text = json.dumps(nfc_payload)

        # Write to NFC depending on reader type
        if tx_type == 'mobile':
            # Expect the client to write using Web NFC and also send back nfc_data read confirmation
            # If client already read from tag, prefer that id
            client_nfc_text = data.get('nfc_data')
            if client_nfc_text:
                try:
                    client_json = json.loads(client_nfc_text)
                    card_id = client_json.get('id') or card_id
                except Exception:
                    pass
            # Server cannot write via mobile; proceed to store card with generated/confirmed id
        elif tx_type == 'external':
            ok = write_nfc_card(json_text)
            if not ok:
                return "Failed to write to NFC card", 500
        else:
            return "Invalid reader type", 400

        # Upsert card in DB
        card = db.session.execute(db.select(NfcCard).filter_by(card_id=card_id)).scalar()
        if card:
            card.username = session['username']
            card.status = 'active'
            card.email = email
            card.roll_no = roll_no
        else:
            card = NfcCard(card_id=card_id, username=session['username'], status='active', email=email, roll_no=roll_no)
            db.session.add(card)
        db.session.commit()
        return "Card registered/activated successfully"
    except Exception as e:
        print(f"Error registering card: {e}")
        return "Internal server error", 500

@app.route('/nfc/cards/<card_id>/toggle', methods=['POST'])
def toggle_nfc_card(card_id):
    if 'username' not in session:
        return "Please login first", 401
    card = db.session.execute(db.select(NfcCard).filter_by(card_id=card_id)).scalar()
    if not card:
        return "Card not found", 404
    if card.username != session['username']:
        return "Not authorized to modify this card", 403
    new_status = 'inactive' if card.status == 'active' else 'active'
    card.status = new_status
    db.session.commit()
    return f"Card status updated to {new_status}"

@app.route('/nfc/cards/<card_id>/set-pin', methods=['POST'])
def set_nfc_pin(card_id):
    if 'username' not in session:
        return "Please login first", 401
    
    data = request.get_json()
    if not data or 'pin' not in data:
        return "PIN is required", 400
    
    pin = data['pin']
    
    # Validate PIN format (4 digits)
    if not pin.isdigit() or len(pin) != 4:
        return "PIN must be exactly 4 digits", 400
    
    card = db.session.execute(db.select(NfcCard).filter_by(card_id=card_id)).scalar()
    if not card:
        return "Card not found", 404
    
    if card.username != session['username']:
        return "Not authorized to modify this card", 403
    
    # Set the PIN
    card.pin = pin
    db.session.commit()
    
    return "PIN set successfully"

@app.route('/nfc/cards/<card_id>/update-pin', methods=['POST'])
def update_nfc_pin(card_id):
    if 'username' not in session:
        return "Please login first", 401
    
    data = request.get_json()
    if not data or 'old_pin' not in data or 'new_pin' not in data:
        return "Old PIN and new PIN are required", 400
    
    old_pin = data['old_pin']
    new_pin = data['new_pin']
    
    # Validate PIN format (4 digits)
    if not old_pin.isdigit() or len(old_pin) != 4:
        return "Old PIN must be exactly 4 digits", 400
    
    if not new_pin.isdigit() or len(new_pin) != 4:
        return "New PIN must be exactly 4 digits", 400
    
    card = db.session.execute(db.select(NfcCard).filter_by(card_id=card_id)).scalar()
    if not card:
        return "Card not found", 404
    
    if card.username != session['username']:
        return "Not authorized to modify this card", 403
    
    # Check if card has a PIN set
    if not card.pin:
        return "Card PIN not set. Please set a PIN first.", 400
    
    # Verify old PIN
    if card.pin != old_pin:
        return "Invalid old PIN", 403
    
    # Update the PIN
    card.pin = new_pin
    db.session.commit()
    
    return "PIN updated successfully"

@app.route('/nfc/cards/<card_id>/verify-pin', methods=['POST'])
def verify_nfc_pin(card_id):
    if 'username' not in session:
        return "Please login first", 401
    
    data = request.get_json()
    if not data or 'pin' not in data:
        return "PIN is required", 400
    
    pin = data['pin']
    
    card = db.session.execute(db.select(NfcCard).filter_by(card_id=card_id)).scalar()
    if not card:
        return "Card not found", 404
    
    if card.username != session['username']:
        return "Not authorized to access this card", 403
    
    # Check if card has a PIN set
    if not card.pin:
        return "Card PIN not set", 400
    
    # Verify PIN
    if card.pin != pin:
        return "Invalid PIN", 403
    
    return "PIN verified successfully"

@app.route('/payments')
def payments():
    return render_template('list_payments.html')

@app.route('/nfc-pin-management')
def nfc_pin_management():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    # Get user's NFC cards
    nfc_cards = db.session.execute(
        db.select(NfcCard).filter_by(username=session['username']).order_by(NfcCard.updated_at.desc())
    ).scalars().all()
    
    return render_template('nfc_pin_management.html', nfc_cards=nfc_cards)


@app.route('/admin/analytical-dashboard')
def admin_analytical_dashboard():
    # Check if user is logged in and is admin
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    # Check if user is admin
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    # Get statistics
    total_users = db.session.execute(db.select(Credentials)).scalars().all()
    total_users_count = len(total_users)
    
    pending_payments = db.session.execute(
        db.select(PaymentProof).filter_by(status='pending')
    ).scalars().all()
    pending_payments_count = len(pending_payments)
    
    total_transactions = db.session.execute(db.select(Transactions)).scalars().all()
    total_transactions_count = len(total_transactions)
    
    all_payments = db.session.execute(db.select(PaymentProof)).scalars().all()
    total_amount = sum([p.amount for p in all_payments if p.status == 'approved'])
    
    # Get NFC cards count
    nfc_cards = db.session.execute(db.select(NfcCard)).scalars().all()
    nfc_cards_count = len(nfc_cards)
    
    # Calculate daily revenue for last 7 days
    from datetime import datetime, timedelta
    daily_revenue = []
    for i in range(7):
        date = datetime.now() - timedelta(days=i)
        day_payments = [p for p in all_payments if p.submitted_at.date() == date.date() and p.status == 'approved']
        daily_revenue.append(sum([p.amount for p in day_payments]))
    daily_revenue.reverse()
    
    # Calculate user growth for last 7 days
    daily_user_growth = []
    for i in range(7):
        date = datetime.now() - timedelta(days=i)
        day_users = [u for u in total_users if u.created_at.date() == date.date()]
        daily_user_growth.append(len(day_users))
    daily_user_growth.reverse()
    
    return render_template('admin_analytical_dashboard.html',
                         total_users=total_users_count,
                         pending_payments=pending_payments_count,
                         total_transactions=total_transactions_count,
                         total_amount=total_amount,
                         nfc_cards_count=nfc_cards_count,
                         daily_revenue=daily_revenue,
                         daily_user_growth=daily_user_growth)

# ===== ADMIN SIDEBAR ROUTES =====

@app.route('/admin/users')
def admin_users():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    users = db.session.execute(db.select(Credentials).order_by(Credentials.created_at.desc())).scalars().all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/user-profiles')
def admin_user_profiles():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    profiles = db.session.execute(db.select(StudentProfile).order_by(StudentProfile.updated_at.desc())).scalars().all()
    return render_template('admin_user_profiles.html', profiles=profiles)

@app.route('/admin/nfc-cards')
def admin_nfc_cards():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    nfc_cards = db.session.execute(db.select(NfcCard).order_by(NfcCard.updated_at.desc())).scalars().all()
    return render_template('admin_nfc_cards.html', nfc_cards=nfc_cards)

@app.route('/admin/transactions')
def admin_transactions():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    transactions = db.session.execute(db.select(Transactions).order_by(Transactions.date.desc())).scalars().all()
    return render_template('admin_transactions.html', transactions=transactions)

@app.route('/admin/payments')
def admin_payments():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    payments = db.session.execute(db.select(PaymentProof).order_by(PaymentProof.submitted_at.desc())).scalars().all()
    return render_template('admin_payments.html', payments=payments)

@app.route('/admin/college-payments')
def admin_college_payments():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    college_payments = db.session.execute(db.select(CollegePayments).order_by(CollegePayments.payment_date.desc())).scalars().all()
    return render_template('admin_college_payments.html', college_payments=college_payments)

@app.route('/admin/refunds')
def admin_refunds():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    refunds = db.session.execute(db.select(RefundRequest).order_by(RefundRequest.requested_at.desc())).scalars().all()
    return render_template('admin_refunds.html', refunds=refunds)

@app.route('/admin/learning-pods')
def admin_learning_pods():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    pods = db.session.execute(db.select(LearningPod).order_by(LearningPod.created_at.desc())).scalars().all()
    return render_template('admin_learning_pods.html', pods=pods)

@app.route('/admin/library')
def admin_library():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    books = db.session.execute(db.select(LibraryBook).order_by(LibraryBook.created_at.desc())).scalars().all()
    transactions = db.session.execute(db.select(LibraryTransaction).order_by(LibraryTransaction.issue_date.desc())).scalars().all()
    return render_template('admin_library.html', books=books, transactions=transactions)

@app.route('/admin/exams')
def admin_exams():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    mock_exams = db.session.execute(db.select(MockExam).order_by(MockExam.created_at.desc())).scalars().all()
    contests = db.session.execute(db.select(ExamContest).order_by(ExamContest.created_at.desc())).scalars().all()
    return render_template('admin_exams.html', mock_exams=mock_exams, contests=contests)

@app.route('/admin/contests')
def admin_contests():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    contests = db.session.execute(db.select(ExamContest).order_by(ExamContest.created_at.desc())).scalars().all()
    participations = db.session.execute(db.select(ContestParticipation).order_by(ContestParticipation.submitted_at.desc())).scalars().all()
    return render_template('admin_contests.html', contests=contests, participations=participations)

@app.route('/admin/settings')
def admin_settings():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    return render_template('admin_settings.html')

@app.route('/admin/logs')
def admin_logs():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    return render_template('admin_logs.html')

@app.route('/admin/security')
def admin_security():
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    return render_template('admin_security.html')

@app.route('/request_refund', methods=['POST'])
def request_refund():
    # Check if user is logged in
    if 'username' not in session:
        flash("Please login first to request refund")
        return redirect(url_for('access'))
    
    try:
        amount = request.form.get('amount')
        reason = request.form.get('reason', '')
        
        # Validate required fields
        if not amount or not reason:
            flash("Please provide both amount and reason")
            return redirect(url_for('student'))
        
        # Validate amount
        try:
            amount = int(amount)
            if amount <= 0:
                flash("Amount must be greater than 0")
                return redirect(url_for('student'))
        except (ValueError, TypeError):
            flash("Please enter a valid amount")
            return redirect(url_for('student'))
        
        # Get user's current balance
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user:
            flash("User not found")
            return redirect(url_for('student'))
        
        if user.credits < amount:
            flash(f"Cannot request refund of ₹{amount}. Your current balance is only ₹{user.credits} credits.")
            return redirect(url_for('student'))
        
        # Save refund request to database
        refund_request = RefundRequest(
            username=session['username'],
            amount=amount,
            reason=reason,
            status='pending'
        )
        
        db.session.add(refund_request)
        db.session.commit()
        
        flash("Refund request submitted successfully! It will be reviewed by admin.")
        return redirect(url_for('student'))
        
    except Exception as e:
        print(f"Error in request_refund: {e}")
        flash("An error occurred while submitting refund request")
        return redirect(url_for('student'))

@app.route('/admin')
def admin_dashboard():
    # Redirect to analytical dashboard
    return redirect(url_for('admin_analytical_dashboard'))

@app.route('/admin/old-dashboard')
def admin_old_dashboard():
    # Check if user is logged in and is admin
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    
    # Check if user is admin
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    # Get pending payments
    pending_payments = db.session.execute(
        db.select(PaymentProof).filter_by(status='pending').order_by(PaymentProof.submitted_at.desc())
    ).scalars().all()
    
    # Get all payments
    all_payments = db.session.execute(
        db.select(PaymentProof).order_by(PaymentProof.submitted_at.desc())
    ).scalars().all()
    
    # Get all users
    users = db.session.execute(
        db.select(Credentials).order_by(Credentials.created_at.desc())
    ).scalars().all()
    
    # Get pending refund requests
    pending_refunds = db.session.execute(
        db.select(RefundRequest).filter_by(status='pending').order_by(RefundRequest.requested_at.desc())
    ).scalars().all()
    
    # Get all refund requests
    all_refunds = db.session.execute(
        db.select(RefundRequest).order_by(RefundRequest.requested_at.desc())
    ).scalars().all()
    
    # Get all transactions
    transactions = db.session.execute(
        db.select(Transactions).order_by(Transactions.date.desc()).limit(50)
    ).scalars().all()
    
    # Convert transactions to JSON-serializable format
    transactions_data = []
    for t in transactions:
        transactions_data.append({
            'id': t.id,
            'from_user': t.from_user,
            'to_user': t.to_user,
            'amount': t.amount,
            'date': t.date.isoformat() if t.date else None
        })
    
    # Get all college payments
    college_payments = db.session.execute(
        db.select(CollegePayments).order_by(CollegePayments.payment_date.desc()).limit(50)
    ).scalars().all()
    
    # Convert college payments to JSON-serializable format
    college_payments_data = []
    for cp in college_payments:
        college_payments_data.append({
            'id': cp.id,
            'student_name': cp.student_name,
            'student_id': cp.student_id,
            'student_email': cp.student_email,
            'student_phone': cp.student_phone,
            'department': cp.department,
            'fee_type': cp.fee_type,
            'amount': cp.amount,
            'semester': cp.semester,
            'academic_year': cp.academic_year,
            'payment_method': cp.payment_method,
            'nfc_reader_type': cp.nfc_reader_type,
            'sender_username': cp.sender_username,
            'remarks': cp.remarks,
            'payment_date': cp.payment_date.isoformat() if cp.payment_date else None,
            'status': cp.status
        })
    
    # Calculate statistics
    total_users = len(users)
    pending_payments_count = len(pending_payments)
    total_transactions = len(transactions)
    total_amount = sum([p.amount for p in all_payments if p.status == 'approved'])
    
    # Calculate additional analytics data
    approved_payments = len([p for p in all_payments if p.status == 'approved'])
    rejected_payments = len([p for p in all_payments if p.status == 'rejected'])
    
    # Calculate daily revenue for last 7 days
    from datetime import datetime, timedelta
    daily_revenue = []
    for i in range(7):
        date = datetime.now() - timedelta(days=i)
        day_payments = [p for p in all_payments if p.submitted_at.date() == date.date() and p.status == 'approved']
        daily_revenue.append(sum([p.amount for p in day_payments]))
    daily_revenue.reverse()  # Reverse to show oldest to newest
    
    # Calculate user growth for last 7 days
    daily_user_growth = []
    for i in range(7):
        date = datetime.now() - timedelta(days=i)
        day_users = [u for u in users if u.created_at.date() == date.date()]
        daily_user_growth.append(len(day_users))
    daily_user_growth.reverse()
    
    # Calculate transaction volume for last 7 days
    daily_transactions = []
    for i in range(7):
        date = datetime.now() - timedelta(days=i)
        day_transactions = [t for t in transactions if t.date.date() == date.date()]
        daily_transactions.append(len(day_transactions))
    daily_transactions.reverse()
    
    # Debug: Print transaction data
    print(f"Total transactions: {len(transactions)}")
    print(f"Daily transactions: {daily_transactions}")
    for i, t in enumerate(transactions):
        print(f"Transaction {i+1}: {t.from_user} -> {t.to_user}, Amount: {t.amount}, Date: {t.date}")
    
    # Calculate monthly revenue for last 6 months
    monthly_revenue = []
    for i in range(6):
        month_start = datetime.now().replace(day=1) - timedelta(days=30*i)
        month_payments = [p for p in all_payments if p.submitted_at.month == month_start.month and p.submitted_at.year == month_start.year and p.status == 'approved']
        monthly_revenue.append(sum([p.amount for p in month_payments]))
    monthly_revenue.reverse()
    
    return render_template('admin_dashboard.html',
                         payment_proofs=all_payments,
                         refund_requests=all_refunds,
                         users=users,
                         transactions=transactions,  # Original objects for template display
                         transactions_json=transactions_data,  # JSON data for JavaScript
                         college_payments=college_payments,  # Original objects for template display
                         college_payments_json=college_payments_data,  # JSON data for JavaScript
                         total_users=total_users,
                         pending_payments=pending_payments_count,
                         total_transactions=total_transactions,
                         total_amount=total_amount,
                         approved_payments=approved_payments,
                         rejected_payments=rejected_payments,
                         daily_revenue=daily_revenue,
                         daily_user_growth=daily_user_growth,
                         daily_transactions=daily_transactions,
                         monthly_revenue=monthly_revenue)

@app.route('/admin/approve_payment', methods=['POST'])
def approve_payment():
    # Check if user is logged in and is admin
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    try:
        payment_id = request.form.get('payment_id')
        username = request.form.get('username')
        amount = request.form.get('amount')
        
        if not payment_id or not username or not amount:
            flash("Missing required data")
            return redirect(url_for('admin_dashboard'))
        
        # Convert amount to integer
        try:
            amount = int(amount)
        except (ValueError, TypeError):
            flash("Invalid amount")
            return redirect(url_for('admin_dashboard'))
        
        # Get the payment proof
        payment_proof = db.session.execute(db.select(PaymentProof).filter_by(id=payment_id)).scalar()
        if not payment_proof:
            flash("Payment proof not found")
            return redirect(url_for('admin_dashboard'))
        
        if payment_proof.status != 'pending':
            flash("Payment has already been processed")
            return redirect(url_for('admin_dashboard'))
        
        # Get the user to credit
        user_to_credit = db.session.execute(db.select(Credentials).filter_by(username=username)).scalar()
        if not user_to_credit:
            flash("User not found")
            return redirect(url_for('admin_dashboard'))
        
        # Update payment proof status
        payment_proof.status = 'approved'
        payment_proof.reviewed_at = datetime.utcnow()
        payment_proof.reviewed_by = session['username']
        payment_proof.review_note = f"Approved by {session['username']}"
        
        # Credit the user's wallet
        user_to_credit.credits += amount
        
        # Save changes
        db.session.commit()
        
        flash(f"Payment approved! {username} has been credited with ₹{amount}")
        return redirect(url_for('admin_dashboard'))
        
    except Exception as e:
        print(f"Error in approve_payment: {e}")
        flash("An error occurred while approving payment")
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/reject_payment', methods=['POST'])
def reject_payment():
    # Check if user is logged in and is admin
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    try:
        payment_id = request.form.get('payment_id')
        reason = request.form.get('reason', 'No reason provided')
        
        if not payment_id:
            flash("Missing payment ID")
            return redirect(url_for('admin_dashboard'))
        
        # Get the payment proof
        payment_proof = db.session.execute(db.select(PaymentProof).filter_by(id=payment_id)).scalar()
        if not payment_proof:
            flash("Payment proof not found")
            return redirect(url_for('admin_dashboard'))
        
        if payment_proof.status != 'pending':
            flash("Payment has already been processed")
            return redirect(url_for('admin_dashboard'))
        
        # Update payment proof status
        payment_proof.status = 'rejected'
        payment_proof.reviewed_at = datetime.utcnow()
        payment_proof.reviewed_by = session['username']
        payment_proof.review_note = f"Rejected by {session['username']}: {reason}"
        
        # Save changes
        db.session.commit()
        
        flash(f"Payment rejected. Reason: {reason}")
        return redirect(url_for('admin_dashboard'))
        
    except Exception as e:
        print(f"Error in reject_payment: {e}")
        flash("An error occurred while rejecting payment")
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/approve_refund', methods=['POST'])
def approve_refund():
    # Check if user is logged in and is admin
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    try:
        refund_id = request.form.get('refund_id')
        username = request.form.get('username')
        amount = request.form.get('amount')
        
        if not refund_id or not username or not amount:
            flash("Missing required data")
            return redirect(url_for('admin_dashboard'))
        
        # Convert amount to integer
        try:
            amount = int(amount)
        except (ValueError, TypeError):
            flash("Invalid amount")
            return redirect(url_for('admin_dashboard'))
        
        # Get the refund request
        refund_request = db.session.execute(db.select(RefundRequest).filter_by(id=refund_id)).scalar()
        if not refund_request:
            flash("Refund request not found")
            return redirect(url_for('admin_dashboard'))
        
        if refund_request.status != 'pending':
            flash("Refund request has already been processed")
            return redirect(url_for('admin_dashboard'))
        
        # Get the user to refund
        user_to_refund = db.session.execute(db.select(Credentials).filter_by(username=username)).scalar()
        if not user_to_refund:
            flash("User not found")
            return redirect(url_for('admin_dashboard'))
        
        # Check if user still has sufficient credits for refund
        if user_to_refund.credits < amount:
            flash(f"Insufficient balance for refund. User has ₹{user_to_refund.credits} credits, but refund amount is ₹{amount}")
            return redirect(url_for('admin_dashboard'))
        
        # Update refund request status
        refund_request.status = 'approved'
        refund_request.reviewed_at = datetime.utcnow()
        refund_request.reviewed_by = session['username']
        refund_request.review_note = f"Approved by {session['username']}"
        
        # Process the refund (deduct from user's wallet)
        user_to_refund.credits -= amount
        
        # Save changes
        db.session.commit()
        
        flash(f"Refund approved! {username} has been refunded ₹{amount}")
        return redirect(url_for('admin_dashboard'))
        
    except Exception as e:
        print(f"Error in approve_refund: {e}")
        flash("An error occurred while approving refund")
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/reject_refund', methods=['POST'])
def reject_refund():
    # Check if user is logged in and is admin
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('student'))
    
    try:
        refund_id = request.form.get('refund_id')
        reason = request.form.get('reason', 'No reason provided')
        
        if not refund_id:
            flash("Missing refund ID")
            return redirect(url_for('admin_dashboard'))
        
        # Get the refund request
        refund_request = db.session.execute(db.select(RefundRequest).filter_by(id=refund_id)).scalar()
        if not refund_request:
            flash("Refund request not found")
            return redirect(url_for('admin_dashboard'))
        
        if refund_request.status != 'pending':
            flash("Refund request has already been processed")
            return redirect(url_for('admin_dashboard'))
        
        # Update refund request status
        refund_request.status = 'rejected'
        refund_request.reviewed_at = datetime.utcnow()
        refund_request.reviewed_by = session['username']
        refund_request.review_note = f"Rejected by {session['username']}: {reason}"
        
        # Save changes
        db.session.commit()
        
        flash(f"Refund rejected. Reason: {reason}")
        return redirect(url_for('admin_dashboard'))
        
    except Exception as e:
        print(f"Error in reject_refund: {e}")
        flash("An error occurred while rejecting refund")
        return redirect(url_for('admin_dashboard'))

@app.route('/teacher_dashboard')
def teacher_dashboard():
    # Check if user is logged in and is teacher
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    
    # Check if user is teacher
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'teacher':
        flash("Access denied. Teacher privileges required.")
        return redirect(url_for('student'))
    
    return render_template('teacher_dashboard.html', user_credits=user.credits)

@app.route('/college_dashboard')
def college_dashboard():
    # Check if user is logged in and is college
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('access'))
    
    # Check if user is college
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'college':
        flash("Access denied. College privileges required.")
        return redirect(url_for('student'))
    
    # Get all college payments
    college_payments = db.session.execute(
        db.select(CollegePayments).order_by(CollegePayments.payment_date.desc())
    ).scalars().all()
    
    # Get recent payments (last 10)
    recent_payments = college_payments[:10] if college_payments else []
    
    # Calculate statistics
    stats = {
        'total_payments': len(college_payments),
        'total_amount': sum([p.amount for p in college_payments]),
        'completed_payments': len([p for p in college_payments if p.status == 'completed']),
        'current_credits': user.credits
    }
    
    # Calculate summary data
    by_fee_type = {}
    unique_students = set()
    total_amount = 0
    
    for payment in college_payments:
        # Fee type summary
        fee_type = payment.fee_type
        if fee_type not in by_fee_type:
            by_fee_type[fee_type] = {'count': 0, 'total': 0}
        by_fee_type[fee_type]['count'] += 1
        by_fee_type[fee_type]['total'] += payment.amount
        
        # Unique students
        unique_students.add(payment.student_id)
        total_amount += payment.amount
    
    summary = {
        'by_fee_type': by_fee_type,
        'unique_students': len(unique_students),
        'average_amount': round(total_amount / len(college_payments), 2) if college_payments else 0
    }
    
    return render_template('college_dashboard.html',
                         college_payments=college_payments,
                         recent_payments=recent_payments,
                         stats=stats,
                         summary=summary)

@app.route('/access', methods=['GET', 'POST'])
def access():
    return render_template('login_signup.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("You have been logged out successfully")
    return redirect(url_for('access'))

# Attendance Routes
@app.route('/read_external_nfc', methods=['GET'])
def read_external_nfc():
    """Read NFC card using external reader"""
    try:
        nfc_text = read_nfc_card()
        if nfc_text:
            # Validate card status
            try:
                payload = json.loads(nfc_text)
                card_id = payload.get('id') or payload.get('card_id')
                username_on_card = payload.get('name')
                if not card_id:
                    return jsonify({'success': False, 'message': 'Card ID missing'}), 400
                card = db.session.execute(db.select(NfcCard).filter_by(card_id=card_id)).scalar()
                if not card:
                    return jsonify({'success': False, 'message': 'Card not registered'}), 403
                if card.status != 'active':
                    return jsonify({'success': False, 'message': 'Card is inactive'}), 403
                if username_on_card and card.username != username_on_card:
                    return jsonify({'success': False, 'message': 'Card owner mismatch'}), 403
                # Normalize roll number key
                if 'roll' in payload and 'roll_no' not in payload:
                    payload['roll_no'] = payload.get('roll')
                return jsonify({'success': True, 'nfc_data': json.dumps(payload)})
            except Exception:
                return jsonify({'success': True, 'nfc_data': nfc_text})
        else:
            return jsonify({
                'success': False,
                'message': 'No NFC card detected'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error reading NFC: {str(e)}'
        })

@app.route('/save_attendance', methods=['POST'])
def save_attendance():
    """Save attendance data to Excel file - ONLY Pin Numbers"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Please login first'})
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return jsonify({'success': False, 'message': 'Admin or Teacher privileges required'})
        
        data = request.get_json()
        session_data = data.get('session')
        
        if not session_data or not session_data.get('presentStudents'):
            return jsonify({'success': False, 'message': 'No attendance data to save'})
        
        # Create classes directory if it doesn't exist
        attendance_dir = os.path.join('static', 'classes')
        if not os.path.exists(attendance_dir):
            os.makedirs(attendance_dir)
        
        # Get session details
        year = session_data.get('year', 'Unknown')
        department = session_data.get('department', 'Unknown')
        section = session_data.get('section', 'Unknown')
        subject = session_data.get('subject', 'Unknown')
        period = session_data.get('period', '1')
        numberOfPeriods = session_data.get('numberOfPeriods', 1)
        present_students = [str(student) for student in session_data.get('presentStudents', [])]  # Convert to strings
        
        # Use existing class file instead of creating new one
        class_clean = f"{year}_{department.replace(' ', '_')}_{section}"
        class_file_path = os.path.join('static', 'classes', f"{class_clean}.xlsx")
        
        if not os.path.exists(class_file_path):
            return jsonify({'success': False, 'message': f'Class file not found: {class_clean}.xlsx. Please create the class file first.'})
        
        print(f"Debug: Looking for class file: {class_file_path}")
        print(f"Debug: Present students: {present_students}")
        
        # Load existing class file
        try:
            # Check if file exists and is accessible
            if not os.path.exists(class_file_path):
                return jsonify({'success': False, 'message': f'Class file not found: {class_file_path}'})
            
            # Check if file is locked (being used by another application)
            try:
                with open(class_file_path, 'r+b') as f:
                    pass
            except (PermissionError, OSError) as e:
                return jsonify({'success': False, 'message': f'File is locked or inaccessible. Please close Excel or any other application that might be using this file. Error: {str(e)}'})
            
            wb = openpyxl.load_workbook(class_file_path)
            print(f"Debug: Loaded workbook, sheets: {wb.sheetnames}")
            
            # Get today's date for sheet name
            today_date = datetime.now().strftime('%Y-%m-%d')
            sheet_name = today_date
            print(f"Debug: Looking for sheet: {sheet_name}")
            
            # Check if sheet for today's date exists
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"Debug: Using existing sheet: {sheet_name}")
            else:
                print(f"Debug: Creating new sheet: {sheet_name}")
                # Create new sheet for today's date
                ws = wb.create_sheet(sheet_name)
                
                # Add Pin Number header in first column
                ws['A1'] = 'Pin Number'
                
                # Get all pin numbers from the original class list (date sheet)
                # Find the first sheet that contains pin numbers (should be a date sheet)
                original_ws = None
                for existing_sheet_name in wb.sheetnames:
                    if existing_sheet_name != 'Class Information':  # Skip info sheet
                        original_ws = wb[existing_sheet_name]
                        break
                
                if original_ws:
                    for row in range(2, original_ws.max_row + 1):
                        pin = original_ws[f'A{row}'].value
                        if pin:
                            ws[f'A{row}'] = pin
            
            # Calculate column positions based on period numbers
            # If start_period is 4, we need to place the subject in column 5 (4+1)
            # This leaves empty columns for periods 1, 2, 3
            
            if numberOfPeriods > 1:
                # Multiple continuous periods - use same subject name in all columns
                start_col = int(period) + 1  # +1 because column 1 is Pin Number
                
                # Add subject header in ALL continuous period columns
                for i in range(numberOfPeriods):
                    period_col = start_col + i
                    ws.cell(row=1, column=period_col, value=subject)
                    ws.cell(row=1, column=period_col).font = openpyxl.styles.Font(bold=True)
                
                # Get all pin numbers from the current sheet
                all_pins = {}
                for row in range(2, ws.max_row + 1):
                    pin = ws[f'A{row}'].value
                    if pin:
                        all_pins[str(pin)] = row  # Convert to string for comparison
                
                # Mark attendance for all continuous periods
                for pin, row_num in all_pins.items():
                    for i in range(numberOfPeriods):
                        period_col = start_col + i
                        if pin in present_students:
                            ws.cell(row=row_num, column=period_col, value='P')  # Present
                        else:
                            ws.cell(row=row_num, column=period_col, value='A')  # Absent
            else:
                # Single period - place in the correct period column
                period_col = int(period) + 1  # +1 because column 1 is Pin Number
                
                # Add subject header
                ws.cell(row=1, column=period_col, value=subject)
                ws.cell(row=1, column=period_col).font = openpyxl.styles.Font(bold=True)
                
                # Get all pin numbers from the current sheet
                all_pins = {}
                for row in range(2, ws.max_row + 1):
                    pin = ws[f'A{row}'].value
                    if pin:
                        all_pins[str(pin)] = row  # Convert to string for comparison
                
                # Mark attendance for this period only
                for pin, row_num in all_pins.items():
                    if pin in present_students:
                        ws.cell(row=row_num, column=period_col, value='P')  # Present
                    else:
                        ws.cell(row=row_num, column=period_col, value='A')  # Absent
            
            # Save the modified class file
            wb.save(class_file_path)
            
        except Exception as e:
            print(f"Debug: Error in save_attendance: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': f'Error modifying class file: {str(e)}'})
        
        return jsonify({
            'success': True,
            'message': f'Attendance saved successfully to class file: {class_clean}.xlsx (Sheet: {today_date})',
            'filename': f'{class_clean}.xlsx'
        })
        
    except Exception as e:
        print(f"Error saving attendance: {e}")
        return jsonify({'success': False, 'message': f'Error saving attendance: {str(e)}'})

@app.route('/download_attendance', methods=['POST'])
def download_attendance():
    """Download attendance as Excel file - ONLY Pin Numbers"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Please login first'})
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return jsonify({'success': False, 'message': 'Admin or Teacher privileges required'})
        
        data = request.get_json()
        session_data = data.get('session')
        
        if not session_data or not session_data.get('presentStudents'):
            return jsonify({'success': False, 'message': 'No attendance data to download'})
        
        # Get session details
        year = session_data.get('year', 'Unknown')
        department = session_data.get('department', 'Unknown')
        section = session_data.get('section', 'Unknown')
        subject = session_data.get('subject', 'Unknown')
        period = session_data.get('period', '1')
        numberOfPeriods = session_data.get('numberOfPeriods', 1)
        present_students = [str(student) for student in session_data.get('presentStudents', [])]  # Convert to strings
        
        # Use existing class file for download
        class_clean = f"{year}_{department.replace(' ', '_')}_{section}"
        class_file_path = os.path.join('static', 'classes', f"{class_clean}.xlsx")
        
        if not os.path.exists(class_file_path):
            return jsonify({'success': False, 'message': f'Class file not found: {class_clean}.xlsx. Please create the class file first.'})
        
        # Load existing class file and add attendance column
        try:
            # Check if file exists and is accessible
            if not os.path.exists(class_file_path):
                return jsonify({'success': False, 'message': f'Class file not found: {class_file_path}'})
            
            # Check if file is locked (being used by another application)
            try:
                with open(class_file_path, 'r+b') as f:
                    pass
            except (PermissionError, OSError) as e:
                return jsonify({'success': False, 'message': f'File is locked or inaccessible. Please close Excel or any other application that might be using this file. Error: {str(e)}'})
            
            wb = openpyxl.load_workbook(class_file_path)
            
            # Get today's date for sheet name
            today_date = datetime.now().strftime('%Y-%m-%d')
            sheet_name = today_date
            
            # Check if sheet for today's date exists
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                # Create new sheet for today's date
                ws = wb.create_sheet(sheet_name)
                
                # Add Pin Number header in first column
                ws['A1'] = 'Pin Number'
                
                # Get all pin numbers from the original class list (date sheet)
                # Find the first sheet that contains pin numbers (should be a date sheet)
                original_ws = None
                for existing_sheet_name in wb.sheetnames:
                    if existing_sheet_name != 'Class Information':  # Skip info sheet
                        original_ws = wb[existing_sheet_name]
                        break
                
                if original_ws:
                    for row in range(2, original_ws.max_row + 1):
                        pin = original_ws[f'A{row}'].value
                        if pin:
                            ws[f'A{row}'] = pin
            
            # Calculate column positions based on period numbers
            # If start_period is 4, we need to place the subject in column 5 (4+1)
            # This leaves empty columns for periods 1, 2, 3
            
            if numberOfPeriods > 1:
                # Multiple continuous periods - use same subject name in all columns
                start_col = int(period) + 1  # +1 because column 1 is Pin Number
                
                # Add subject header in ALL continuous period columns
                for i in range(numberOfPeriods):
                    period_col = start_col + i
                    ws.cell(row=1, column=period_col, value=subject)
                    ws.cell(row=1, column=period_col).font = openpyxl.styles.Font(bold=True)
                
                # Get all pin numbers from the current sheet
                all_pins = {}
                for row in range(2, ws.max_row + 1):
                    pin = ws[f'A{row}'].value
                    if pin:
                        all_pins[str(pin)] = row  # Convert to string for comparison
                
                # Mark attendance for all continuous periods
                for pin, row_num in all_pins.items():
                    for i in range(numberOfPeriods):
                        period_col = start_col + i
                        if pin in present_students:
                            ws.cell(row=row_num, column=period_col, value='P')  # Present
                        else:
                            ws.cell(row=row_num, column=period_col, value='A')  # Absent
            else:
                # Single period - place in the correct period column
                period_col = int(period) + 1  # +1 because column 1 is Pin Number
                
                # Add subject header
                ws.cell(row=1, column=period_col, value=subject)
                ws.cell(row=1, column=period_col).font = openpyxl.styles.Font(bold=True)
                
                # Get all pin numbers from the current sheet
                all_pins = {}
                for row in range(2, ws.max_row + 1):
                    pin = ws[f'A{row}'].value
                    if pin:
                        all_pins[str(pin)] = row  # Convert to string for comparison
                
                # Mark attendance for this period only
                for pin, row_num in all_pins.items():
                    if pin in present_students:
                        ws.cell(row=row_num, column=period_col, value='P')  # Present
                    else:
                        ws.cell(row=row_num, column=period_col, value='A')  # Absent
            
            # Save to memory
            output = io.BytesIO()
            wb.save(output)
            # Close workbook to ensure clean file stream
            wb.close()
            output.seek(0)
            
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error processing class file: {str(e)}'})
        
        # Generate filename for download
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        subject_clean = subject.replace(' ', '_').replace('/', '_')
        filename = f'{class_clean}_with_{subject_clean}_{date_str}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error downloading attendance: {e}")
        return jsonify({'success': False, 'message': f'Error downloading attendance: {str(e)}'})

@app.route('/get_available_periods', methods=['POST'])
def get_available_periods():
    """Get available periods based on existing attendance in Excel file"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Please login first'})
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return jsonify({'success': False, 'message': 'Admin or Teacher privileges required'})
        
        data = request.get_json()
        year = data.get('year', '').strip()
        department = data.get('department', '').strip()
        section = data.get('section', '').strip()
        
        if not all([year, department, section]):
            return jsonify({'success': False, 'message': 'Year, Department, and Section are required'})
        
        # Check if class file exists
        class_clean = f"{year}_{department.replace(' ', '_')}_{section}"
        class_file_path = os.path.join('static', 'classes', f"{class_clean}.xlsx")
        
        if not os.path.exists(class_file_path):
            # If no class file exists, return all periods
            return jsonify({
                'success': True,
                'available_periods': [1, 2, 3, 4, 5, 6, 7, 8, 9],
                'message': 'No class file found, all periods available'
            })
        
        # Load the class file and check today's sheet
        today_date = datetime.now().strftime('%Y-%m-%d')
        
        try:
            # Check if file exists and is accessible
            if not os.path.exists(class_file_path):
                return jsonify({'success': False, 'message': f'Class file not found: {class_file_path}'})
            
            # Check if file is locked (being used by another application)
            try:
                with open(class_file_path, 'r+b') as f:
                    pass
            except (PermissionError, OSError) as e:
                return jsonify({'success': False, 'message': f'File is locked or inaccessible. Please close Excel or any other application that might be using this file. Error: {str(e)}'})
            
            wb = openpyxl.load_workbook(class_file_path)
            
            if today_date not in wb.sheetnames:
                # If no sheet for today, return all periods
                return jsonify({
                    'success': True,
                    'available_periods': [1, 2, 3, 4, 5, 6, 7, 8, 9],
                    'message': 'No attendance taken today, all periods available'
                })
            
            # Get today's sheet
            ws = wb[today_date]
            
            # Check which periods are already taken by looking at column headers
            taken_periods = set()
            
            # Check all columns (skip first column which is Pin Number)
            for col in range(2, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    # New format: Column position determines period number
                    # Column 2 = Period 1, Column 3 = Period 2, etc.
                    period_num = col - 1  # Convert column number to period number
                    taken_periods.add(period_num)
                    
                    # For continuous periods, we need to check if the same subject
                    # appears in consecutive columns and mark all those periods as taken
                    # This is already handled by the column position logic above
            
            # Calculate available periods
            all_periods = set(range(1, 10))  # Periods 1-9
            available_periods = sorted(list(all_periods - taken_periods))
            
            return jsonify({
                'success': True,
                'available_periods': available_periods,
                'taken_periods': sorted(list(taken_periods)),
                'message': f'Found {len(available_periods)} available periods'
            })
            
        except Exception as e:
            print(f"Error reading class file: {e}")
            return jsonify({
                'success': True,
                'available_periods': [1, 2, 3, 4, 5, 6, 7, 8, 9],
                'message': 'Error reading file, showing all periods'
            })
        
    except Exception as e:
        print(f"Error getting available periods: {e}")
        return jsonify({'success': False, 'message': f'Error getting available periods: {str(e)}'})

@app.route('/list_attendance_files', methods=['GET'])
def list_attendance_files():
    """List all attendance Excel files"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Please login first'})
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return jsonify({'success': False, 'message': 'Admin or Teacher privileges required'})
        
        # Get all Excel files from classes directory
        classes_dir = os.path.join('static', 'classes')
        if not os.path.exists(classes_dir):
            return jsonify({'success': True, 'files': []})
        
        files = []
        for filename in os.listdir(classes_dir):
            if filename.endswith('.xlsx') and not filename.startswith('~'):
                file_path = os.path.join(classes_dir, filename)
                file_stat = os.stat(file_path)
                
                # Get file info
                file_info = {
                    'filename': filename,
                    'size': file_stat.st_size,
                    'created': datetime.fromtimestamp(file_stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'modified': datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # Try to get sheet names and dates
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    sheet_names = [name for name in wb.sheetnames if name != 'Class Information']
                    file_info['sheets'] = sheet_names
                    file_info['dates'] = [name for name in sheet_names if re.match(r'\d{4}-\d{2}-\d{2}', name)]
                    wb.close()
                except:
                    file_info['sheets'] = []
                    file_info['dates'] = []
                
                files.append(file_info)
        
        # Sort by modification date (newest first)
        files.sort(key=lambda x: x['modified'], reverse=True)
        
        return jsonify({'success': True, 'files': files})
        
    except Exception as e:
        print(f"Error listing attendance files: {e}")
        return jsonify({'success': False, 'message': f'Error listing files: {str(e)}'})

@app.route('/attendance_files')
def attendance_files_list():
    """Show list of all attendance files"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return redirect(url_for('access'))
        
        return render_template('attendance_files_list.html', user_role=user.role)
        
    except Exception as e:
        print(f"Error loading attendance files page: {e}")
        return redirect(url_for('admin_dashboard' if user.role == 'admin' else 'teacher_dashboard'))

@app.route('/view_attendance_data/<filename>')
def view_attendance_data(filename):
    """View attendance data for a specific file with date range filtering and pagination"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return redirect(url_for('access'))
        
        # Get parameters for filtering and pagination
        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 1))  # Show 1 table per page
        
        # Load the Excel file
        file_path = os.path.join('static', 'classes', filename)
        if not os.path.exists(file_path):
            flash('File not found', 'error')
            return redirect(url_for('admin_dashboard' if user.role == 'admin' else 'teacher_dashboard'))
        
        # Check if file is locked (being used by another application)
        try:
            with open(file_path, 'r+b') as f:
                pass
        except (PermissionError, OSError) as e:
            flash(f'File is locked or inaccessible. Please close Excel or any other application that might be using this file. Error: {str(e)}', 'error')
            return redirect(url_for('admin_dashboard' if user.role == 'admin' else 'teacher_dashboard'))
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        # Get all date sheets and filter by date range
        all_date_sheets = [name for name in wb.sheetnames if re.match(r'\d{4}-\d{2}-\d{2}', name)]
        all_date_sheets.sort(reverse=True)  # Newest first
        
        # Check if any filter is applied
        filter_applied = from_date or to_date
        
        if not filter_applied:
            # No filter applied - show only the latest date
            filtered_date_sheets = all_date_sheets[:1] if all_date_sheets else []
            paginated_dates = filtered_date_sheets
            total_dates = len(filtered_date_sheets)
            total_pages = 1
        else:
            # Filter applied - use pagination
            filtered_date_sheets = []
            for date_str in all_date_sheets:
                if from_date and date_str < from_date:
                    continue
                if to_date and date_str > to_date:
                    continue
                filtered_date_sheets.append(date_str)
            
            # Calculate pagination
            total_dates = len(filtered_date_sheets)
            total_pages = (total_dates + per_page - 1) // per_page
            start_idx = (page - 1) * per_page
            end_idx = start_idx + per_page
            paginated_dates = filtered_date_sheets[start_idx:end_idx]
        
        # Get attendance data for paginated dates
        attendance_data = {}
        all_headers = set()
        
        for date_str in paginated_dates:
            if date_str in wb.sheetnames:
                ws = wb[date_str]
                
                # Get headers (first row)
                headers = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    headers.append(cell_value if cell_value else f'Column {col}')
                all_headers.update(headers)
                
                # Get data rows
                data = []
                for row in range(2, ws.max_row + 1):
                    row_data = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(cell_value if cell_value else '')
                    if any(row_data):  # Only add non-empty rows
                        data.append(row_data)
                
                attendance_data[date_str] = {
                    'headers': headers,
                    'data': data
                }
        
        wb.close()
        
        # Convert set to sorted list for consistent ordering
        all_headers = sorted(list(all_headers))
        
        return render_template('view_attendance_data.html', 
                             filename=filename,
                             all_date_sheets=all_date_sheets,
                             filtered_date_sheets=filtered_date_sheets,
                             paginated_dates=paginated_dates,
                             attendance_data=attendance_data,
                             all_headers=all_headers,
                             from_date=from_date,
                             to_date=to_date,
                             page=page,
                             per_page=per_page,
                             total_pages=total_pages,
                             total_dates=total_dates,
                             filter_applied=filter_applied,
                             user_role=user.role)
        
    except Exception as e:
        print(f"Error viewing attendance data: {e}")
        flash('Error loading attendance data', 'error')
        return redirect(url_for('admin_dashboard' if user.role == 'admin' else 'teacher_dashboard'))

@app.route('/create_class_excel', methods=['POST'])
def create_class_excel():
    """Create class Excel file with pin number ranges"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Please login first'})
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return jsonify({'success': False, 'message': 'Admin or Teacher privileges required'})
        
        data = request.get_json()
        study_year = data.get('studyYear', '').strip()
        department = data.get('department', '').strip()
        section = data.get('section', '').strip()
        start_pin = data.get('startPin')
        end_pin = data.get('endPin')
        skip_numbers_str = data.get('skipNumbers', '').strip()
        
        # Validation
        if not all([study_year, department, section, start_pin, end_pin]):
            return jsonify({'success': False, 'message': 'All required fields are required'})
        
        if start_pin >= end_pin:
            return jsonify({'success': False, 'message': 'Starting pin must be less than ending pin'})
        
        if end_pin - start_pin > 200:
            return jsonify({'success': False, 'message': 'Pin range cannot exceed 200 students'})
        
        # Parse skip numbers
        skip_numbers = set()
        if skip_numbers_str:
            try:
                # Split by comma and handle ranges
                parts = skip_numbers_str.split(',')
                for part in parts:
                    part = part.strip()
                    if '-' in part:
                        # Handle range like "5-10"
                        start_range, end_range = map(int, part.split('-'))
                        skip_numbers.update(range(start_range, end_range + 1))
                    else:
                        # Handle single number
                        skip_numbers.add(int(part))
                
                # Validate skip numbers are within range
                invalid_skips = [num for num in skip_numbers if num < start_pin or num > end_pin]
                if invalid_skips:
                    return jsonify({'success': False, 'message': f'Skip numbers {invalid_skips} are outside the pin range ({start_pin}-{end_pin})'})
                    
            except ValueError:
                return jsonify({'success': False, 'message': 'Invalid skip numbers format. Use comma-separated numbers or ranges like "5,10,15" or "5-10,15-20"'})
        
        # Create class directory if it doesn't exist
        class_dir = os.path.join('static', 'classes')
        if not os.path.exists(class_dir):
            os.makedirs(class_dir)
        
        # Generate filename: studyyear_departmentname_section.xlsx
        department_clean = department.replace(' ', '_').replace('/', '_').replace('\\', '_')
        section_clean = section.replace(' ', '_').replace('/', '_').replace('\\', '_')
        filename = f"{study_year}_{department_clean}_{section_clean}.xlsx"
        filepath = os.path.join(class_dir, filename)
        
        # Check if file already exists
        if os.path.exists(filepath):
            return jsonify({'success': False, 'message': f'Class file already exists: {filename}'})
        
        # Create Excel file with ONLY Pin Numbers
        workbook = openpyxl.Workbook()
        
        # Get today's date for sheet name
        today_date = datetime.now().strftime('%Y-%m-%d')
        
        # Remove default sheet and create new one with date name
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(today_date)
        
        # Add ONLY Pin Number header
        worksheet['A1'] = 'Pin Number'
        worksheet['A1'].font = openpyxl.styles.Font(bold=True)
        
        # Add ONLY pin numbers (excluding skipped numbers)
        row_num = 2
        for pin_num in range(start_pin, end_pin + 1):
            if pin_num not in skip_numbers:
                worksheet.cell(row=row_num, column=1, value=pin_num)  # Pin Number only
                row_num += 1
        
        # Calculate actual total students (excluding skipped numbers)
        actual_total_students = (end_pin - start_pin + 1) - len(skip_numbers)
        
        # Add class information sheet
        info_sheet = workbook.create_sheet('Class Information')
        info_data = [
            ['Study Year', study_year],
            ['Department', department],
            ['Section', section],
            ['Starting Pin', start_pin],
            ['Ending Pin', end_pin],
            ['Skip Numbers', skip_numbers_str if skip_numbers_str else 'None'],
            ['Skipped Count', len(skip_numbers)],
            ['Total Students', actual_total_students],
            ['Created Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Created By', user.username]
        ]
        
        for row, (label, value) in enumerate(info_data, 1):
            info_sheet.cell(row=row, column=1, value=label).font = openpyxl.styles.Font(bold=True)
            info_sheet.cell(row=row, column=2, value=value)
        
        # Set column width for Pin Number column
        worksheet.column_dimensions['A'].width = 15
        
        # Save the workbook
        workbook.save(filepath)
        
        return jsonify({
            'success': True,
            'message': 'Class Excel file created successfully',
            'filename': filename,
            'total_students': actual_total_students,
            'skipped_count': len(skip_numbers),
            'skip_numbers': skip_numbers_str if skip_numbers_str else 'None'
        })
        
    except Exception as e:
        print(f"Error creating class Excel: {e}")
        return jsonify({'success': False, 'message': f'Error creating class Excel: {str(e)}'})

@app.route('/view_class_files')
def view_class_files():
    """View all created class files"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            flash("Access denied. Admin or Teacher privileges required.")
            if user and user.role == 'student':
                return redirect(url_for('student'))
            else:
                return redirect(url_for('admin_dashboard'))
        
        # Get list of class files
        class_dir = os.path.join('static', 'classes')
        class_files = []
        
        if os.path.exists(class_dir):
            for filename in os.listdir(class_dir):
                if filename.endswith('.xlsx'):
                    filepath = os.path.join(class_dir, filename)
                    file_stats = os.stat(filepath)
                    class_files.append({
                        'filename': filename,
                        'size': file_stats.st_size,
                        'created': datetime.fromtimestamp(file_stats.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
                    })
        
        # Sort by creation date (newest first)
        class_files.sort(key=lambda x: x['created'], reverse=True)
        
        return render_template('class_files.html', class_files=class_files, user_role=user.role)
        
    except Exception as e:
        print(f"Error viewing class files: {e}")
        flash("Error loading class files")
        return redirect(url_for('admin_dashboard'))

@app.route('/download_class_file/<filename>')
def download_class_file(filename):
    """Download a specific class file"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            flash("Access denied. Admin or Teacher privileges required.")
            return redirect(url_for('admin_dashboard'))
        
        # Security check - ensure filename is safe
        if not filename.endswith('.xlsx') or '..' in filename or '/' in filename or '\\' in filename:
            flash("Invalid filename")
            return redirect(url_for('view_class_files'))
        
        filepath = os.path.join('static', 'classes', filename)
        
        if not os.path.exists(filepath):
            flash("File not found")
            return redirect(url_for('view_class_files'))
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"Error downloading class file: {e}")
        flash("Error downloading file")
        return redirect(url_for('view_class_files'))

@app.route('/view_attendance_history')
def view_attendance_history():
    """View attendance history"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            flash("Admin or Teacher privileges required")
            if user and user.role == 'teacher':
                return redirect(url_for('teacher_dashboard'))
            else:
                return redirect(url_for('admin_dashboard'))
        
        # Get list of attendance files
        attendance_dir = os.path.join('static', 'classes')
        attendance_files = []
        
        if os.path.exists(attendance_dir):
            for filename in os.listdir(attendance_dir):
                if filename.endswith('.xlsx'):
                    filepath = os.path.join(attendance_dir, filename)
                    file_stats = os.stat(filepath)
                    attendance_files.append({
                        'filename': filename,
                        'size': file_stats.st_size,
                        'created': datetime.fromtimestamp(file_stats.st_ctime)
                    })
        
        # Sort by creation date (newest first)
        attendance_files.sort(key=lambda x: x['created'], reverse=True)
        
        return render_template('attendance_history.html', attendance_files=attendance_files)
        
    except Exception as e:
        print(f"Error viewing attendance history: {e}")
        flash("Error loading attendance history")
        return redirect(url_for('admin_dashboard'))

@app.route('/attendance_class_files')
def attendance_class_files():
    """View attendance class files with filtering"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            flash("Access denied. Admin or Teacher privileges required.")
            if user and user.role == 'student':
                return redirect(url_for('student'))
            else:
                return redirect(url_for('admin_dashboard'))
        
        # Get filter parameters
        class_name = request.args.get('class_name', '')
        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')
        subject = request.args.get('subject', '')
        
        # Get attendance files from classes directory
        attendance_dir = 'static/classes'
        if not os.path.exists(attendance_dir):
            os.makedirs(attendance_dir)
        
        files = []
        for filename in os.listdir(attendance_dir):
            if filename.endswith('.xlsx'):
                filepath = os.path.join(attendance_dir, filename)
                file_stat = os.stat(filepath)
                
                # Parse filename to extract class info
                # Format: attendance_ClassName_Subject_YYYYMMDD_HHMMSS.xlsx
                parts = filename.replace('.xlsx', '').split('_')
                class_name_from_file = 'General'
                subject_from_file = 'General'
                date_from_file = None
                
                if len(parts) >= 4:
                    if parts[1] != 'General':
                        class_name_from_file = parts[1]
                    if len(parts) >= 3 and parts[2] != 'Attendance':
                        subject_from_file = parts[2]
                    if len(parts) >= 4:
                        try:
                            date_str = parts[-2]  # YYYYMMDD part
                            date_from_file = datetime.strptime(date_str, '%Y%m%d').date()
                        except:
                            pass
                
                # Apply filters
                if class_name and class_name.lower() not in class_name_from_file.lower():
                    continue
                if subject and subject.lower() not in subject_from_file.lower():
                    continue
                if from_date and date_from_file and date_from_file < datetime.strptime(from_date, '%Y-%m-%d').date():
                    continue
                if to_date and date_from_file and date_from_file > datetime.strptime(to_date, '%Y-%m-%d').date():
                    continue
                
                files.append({
                    'filename': filename,
                    'class_name': class_name_from_file,
                    'subject': subject_from_file,
                    'date': date_from_file,
                    'size': f"{file_stat.st_size / 1024:.1f} KB"
                })
        
        # Sort by date (newest first)
        files.sort(key=lambda x: x['date'] or datetime.min.date(), reverse=True)
        
        return render_template('attendance_class_files.html', 
                             class_files=files,
                             class_name=class_name,
                             from_date=from_date,
                             to_date=to_date,
                             subject=subject)
        
    except Exception as e:
        print(f"Error viewing attendance class files: {e}")
        flash("Error loading attendance class files")
        return redirect(url_for('admin_dashboard'))

@app.route('/edit_attendance/<filename>')
def edit_attendance(filename):
    """Edit attendance data for a specific file"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            flash("Access denied. Admin or Teacher privileges required.")
            if user and user.role == 'student':
                return redirect(url_for('student'))
            else:
                return redirect(url_for('admin_dashboard'))
        
        # Load attendance data from Excel file
        filepath = os.path.join('static', 'classes', filename)
        if not os.path.exists(filepath):
            flash("File not found")
            return redirect(url_for('attendance_class_files'))
        
        # Read Excel file
        df = pd.read_excel(filepath)
        
        # Debug: Print column names and first few rows
        print(f"Excel columns: {df.columns.tolist()}")
        print(f"DataFrame shape: {df.shape}")
        print(f"First few rows:")
        print(df.head())
        print(f"Data types:")
        print(df.dtypes)
        
        # Process attendance data
        attendance_data = {}
        total_records = len(df)
        
        # Handle different column name variations
        name_col = None
        roll_col = None
        email_col = None
        time_col = None
        date_col = None
        
        # Find the correct column names (case insensitive)
        for col in df.columns:
            col_lower = col.lower().strip()
            if 'name' in col_lower:
                name_col = col
            elif 'roll' in col_lower:
                roll_col = col
            elif 'email' in col_lower:
                email_col = col
            elif 'time' in col_lower:
                time_col = col
            elif 'date' in col_lower:
                date_col = col
        
        print(f"Found columns - Name: {name_col}, Roll: {roll_col}, Email: {email_col}, Time: {time_col}, Date: {date_col}")
        
        for _, row in df.iterrows():
            # Get date (use first available date or default)
            date = 'Today'
            if date_col and pd.notna(row[date_col]):
                date = str(row[date_col])
            elif 'Date' in df.columns and pd.notna(row.get('Date')):
                date = str(row['Date'])
            elif 'date' in df.columns and pd.notna(row.get('date')):
                date = str(row['date'])
            
            # Clean up date format
            if date != 'Today':
                try:
                    # Try to parse and format the date
                    if isinstance(date, str) and len(date) > 10:
                        date = date[:10]  # Take only the date part
                except:
                    pass
            
            if date not in attendance_data:
                attendance_data[date] = []
            
            # Get student data with fallbacks
            name = ''
            roll_no = ''
            email = ''
            time = ''
            
            if name_col and pd.notna(row[name_col]):
                name = str(row[name_col])
            elif 'Name' in df.columns and pd.notna(row.get('Name')):
                name = str(row['Name'])
            
            if roll_col and pd.notna(row[roll_col]):
                roll_no = str(row[roll_col])
            elif 'Roll No' in df.columns and pd.notna(row.get('Roll No')):
                roll_no = str(row['Roll No'])
            
            if email_col and pd.notna(row[email_col]):
                email = str(row[email_col])
            elif 'Email' in df.columns and pd.notna(row.get('Email')):
                email = str(row['Email'])
            
            if time_col and pd.notna(row[time_col]):
                time = str(row[time_col])
            elif 'Time' in df.columns and pd.notna(row.get('Time')):
                time = str(row['Time'])
            
            # Only add if we have at least a name
            if name.strip():
                attendance_data[date].append({
                    'name': name,
                    'roll_no': roll_no,
                    'email': email,
                    'time': time,
                    'status': 'Present'  # Default status
                })
        
        # If no data was found, try alternative column names
        if not attendance_data:
            print("No data found with standard column names, trying alternative approach...")
            # Try to use the first few columns as fallback
            for _, row in df.iterrows():
                date = 'Today'
                name = str(row.iloc[0]) if len(row) > 0 and pd.notna(row.iloc[0]) else ''
                roll_no = str(row.iloc[1]) if len(row) > 1 and pd.notna(row.iloc[1]) else ''
                email = str(row.iloc[2]) if len(row) > 2 and pd.notna(row.iloc[2]) else ''
                time = str(row.iloc[3]) if len(row) > 3 and pd.notna(row.iloc[3]) else ''
                
                if name.strip() and name != 'nan':
                    if date not in attendance_data:
                        attendance_data[date] = []
                    
                    attendance_data[date].append({
                        'name': name,
                        'roll_no': roll_no,
                        'email': email,
                        'time': time,
                        'status': 'Present'
                    })
        
        # Calculate summary stats
        all_students = []
        for records in attendance_data.values():
            all_students.extend([record['name'] for record in records])
        
        unique_students = len(set(all_students))
        present_count = len(all_students)
        
        summary_stats = {
            'total_students': unique_students,
            'present_count': present_count,
            'absent_count': 0,
            'attendance_rate': 100.0 if unique_students > 0 else 0.0
        }
        
        return render_template('edit_attendance.html',
                               filename=filename,
                               attendance_data=attendance_data,
                               total_records=total_records,
                               summary_stats=summary_stats,
                               last_modified=datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%Y-%m-%d %H:%M:%S'))
        
    except Exception as e:
        print(f"Error editing attendance: {e}")
        flash("Error loading attendance data")
        return redirect(url_for('attendance_class_files'))


@app.route('/view_attendance_file/<filename>')
def view_attendance_file(filename):
    """View attendance file content without downloading"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            flash("Admin or Teacher privileges required")
            if user and user.role == 'teacher':
                return redirect(url_for('teacher_dashboard'))
            else:
                return redirect(url_for('admin_dashboard'))
        
        # Security check - ensure filename is safe
        if not filename.endswith('.xlsx') or '/' in filename or '\\' in filename:
            flash("Invalid filename")
            return redirect(url_for('view_attendance_history'))
        
        # Try both attendance and classes directories
        filepath = None
        for directory in ['attendance', 'classes']:
            potential_path = os.path.join('static', directory, filename)
            if os.path.exists(potential_path):
                filepath = potential_path
                break
        
        if not filepath:
            flash("File not found")
            return redirect(url_for('view_attendance_history'))
        
        # Check if file is locked (being used by another application)
        try:
            with open(filepath, 'r+b') as f:
                pass
        except (PermissionError, OSError) as e:
            flash(f'File is locked or inaccessible. Please close Excel or any other application that might be using this file. Error: {str(e)}')
            return redirect(url_for('view_attendance_history'))
        
        # Read Excel file and convert to HTML table
        try:
            df = pd.read_excel(filepath)
            attendance_data = df.to_dict('records')
            
            return render_template('view_attendance_file.html', 
                                 filename=filename, 
                                 attendance_data=attendance_data,
                                 total_records=len(attendance_data))
        except Exception as e:
            flash("Error reading attendance file")
            return redirect(url_for('view_attendance_history'))
        
    except Exception as e:
        print(f"Error viewing attendance file: {e}")
        flash("Error viewing file")
        return redirect(url_for('view_attendance_history'))

@app.route('/save_attendance_edits/<filename>', methods=['POST'])
def save_attendance_edits(filename):
    """Save edited attendance back to the Excel file. Admin/Teacher only."""
    try:
        # Auth guard
        if 'username' not in session:
            return jsonify({"error": "Please login first"}), 401
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return jsonify({"error": "Only admin or teacher can edit attendance"}), 403

        # Filename validation
        if not filename.endswith('.xlsx') or '/' in filename or '\\' in filename:
            return jsonify({"error": "Invalid filename"}), 400

        # Try classes directory first, then attendance directory
        filepath = None
        for directory in ['classes', 'attendance']:
            potential_path = os.path.join('static', directory, filename)
            if os.path.exists(potential_path):
                filepath = potential_path
                break
        
        if not filepath:
            return jsonify({"error": "File not found"}), 404

        # Parse payload
        payload = request.get_json(silent=True) or {}
        records = payload.get('records')
        if not isinstance(records, list) or not records:
            return jsonify({"error": "Invalid or empty data"}), 400

        # Ensure file is not locked
        try:
            with open(filepath, 'r+b') as _f:
                pass
        except (PermissionError, OSError) as e:
            return jsonify({"error": f"File is locked or in use: {str(e)}"}), 423

        # Load existing, align columns, and overwrite
        try:
            existing_df = pd.read_excel(filepath)
        except Exception:
            existing_df = None

        # Normalize incoming records to DataFrame
        try:
            new_df = pd.DataFrame.from_records(records)
        except Exception as e:
            return jsonify({"error": f"Invalid data format: {str(e)}"}), 400

        # If existing columns exist, reorder and fill missing
        if existing_df is not None and not existing_df.empty:
            for col in existing_df.columns:
                if col not in new_df.columns:
                    new_df[col] = None
            new_df = new_df[existing_df.columns]

        # Write back to Excel safely
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
                new_df.to_excel(writer, index=False)
        except Exception as e:
            return jsonify({"error": f"Failed to write file: {str(e)}"}), 500

        return jsonify({"success": True})

    except Exception as e:
        print(f"Error saving attendance edits: {e}")
        return jsonify({"error": "Unexpected error while saving"}), 500

@app.route('/download_attendance_file/<filename>')
def download_attendance_file(filename):
    """Download specific attendance file"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            flash("Admin or Teacher privileges required")
            if user and user.role == 'teacher':
                return redirect(url_for('teacher_dashboard'))
            else:
                return redirect(url_for('admin_dashboard'))
        
        # Security check - ensure filename is safe
        if not filename.endswith('.xlsx') or '/' in filename or '\\' in filename:
            flash("Invalid filename")
            return redirect(url_for('view_attendance_history'))
        
        filepath = os.path.join('static', 'classes', filename)
        
        if not os.path.exists(filepath):
            flash("File not found")
            return redirect(url_for('view_attendance_history'))
        
        return send_file(filepath, as_attachment=True)
        
    except Exception as e:
        print(f"Error downloading attendance file: {e}")
        flash("Error downloading file")
        return redirect(url_for('view_attendance_history'))

# Library Management Routes
@app.route('/library_dashboard')
def library_dashboard():
    """Library management dashboard"""
    try:
        # Check if user is logged in and is admin or library staff
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'library']:
            flash("Access denied. Admin or Library privileges required.")
            if user and user.role == 'student':
                return redirect(url_for('student'))
            else:
                return redirect(url_for('admin_dashboard'))
        
        # Get library statistics
        total_books = db.session.execute(db.select(LibraryBook)).scalars().all()
        available_books = sum([book.available_copies for book in total_books])
        total_books_count = len(total_books)
        
        # Get issued books count
        issued_transactions = db.session.execute(
            db.select(LibraryTransaction).filter_by(status='issued')
        ).scalars().all()
        issued_books_count = len(issued_transactions)
        
        # Get overdue books
        from datetime import datetime, timedelta
        today = datetime.utcnow().date()
        overdue_transactions = db.session.execute(
            db.select(LibraryTransaction).filter(
                LibraryTransaction.status == 'issued',
                LibraryTransaction.due_date < today
            )
        ).scalars().all()
        overdue_books_count = len(overdue_transactions)
        
        # Get recent books
        books = db.session.execute(
            db.select(LibraryBook).order_by(LibraryBook.created_at.desc()).limit(20)
        ).scalars().all()
        
        # Get available books for issue dropdown
        available_books = db.session.execute(
            db.select(LibraryBook).filter(LibraryBook.available_copies > 0).order_by(LibraryBook.title)
        ).scalars().all()
        
        # Get recent transactions
        transactions = db.session.execute(
            db.select(LibraryTransaction).order_by(LibraryTransaction.issue_date.desc()).limit(20)
        ).scalars().all()
        
        return render_template('library_dashboard.html',
                             total_books=total_books_count,
                             available_books=available_books,
                             issued_books=issued_books_count,
                             overdue_books=overdue_books_count,
                             books=books,
                             transactions=transactions,
                             overdue_transactions=overdue_transactions)
        
    except Exception as e:
        print(f"Error in library_dashboard: {e}")
        flash("Error loading library dashboard")
        return redirect(url_for('admin_dashboard'))

@app.route('/library/add_book', methods=['POST'])
def add_book():
    """Add new book to library"""
    try:
        # Check if user is logged in and is admin
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'library']:
            flash("Access denied. Admin or Library privileges required.")
            return redirect(url_for('library_dashboard'))
        
        # Get form data
        book_id = request.form.get('book_id')
        title = request.form.get('title')
        author = request.form.get('author')
        isbn = request.form.get('isbn')
        category = request.form.get('category')
        publisher = request.form.get('publisher')
        publication_year = request.form.get('publication_year')
        total_copies = request.form.get('total_copies')
        location = request.form.get('location')
        description = request.form.get('description')
        
        # Validate required fields
        if not all([book_id, title, author, total_copies]):
            flash("Please fill in all required fields")
            return redirect(url_for('library_dashboard'))
        
        # Check if book ID already exists
        existing_book = db.session.execute(
            db.select(LibraryBook).filter_by(book_id=book_id)
        ).scalar()
        
        if existing_book:
            flash("Book ID already exists. Please use a different ID.")
            return redirect(url_for('library_dashboard'))
        
        # Create new book
        new_book = LibraryBook(
            book_id=book_id,
            title=title,
            author=author,
            isbn=isbn,
            category=category,
            publisher=publisher,
            publication_year=int(publication_year) if publication_year else None,
            total_copies=int(total_copies),
            available_copies=int(total_copies),
            location=location,
            description=description
        )
        
        db.session.add(new_book)
        db.session.commit()
        
        flash(f"Book '{title}' added successfully!")
        return redirect(url_for('library_dashboard'))
        
    except Exception as e:
        print(f"Error adding book: {e}")
        flash("Error adding book")
        return redirect(url_for('library_dashboard'))

@app.route('/library/read_nfc', methods=['GET'])
def library_read_nfc():
    """Read NFC card for library book issue"""
    try:
        # Check if user is logged in and is admin or library staff
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Please login first'})
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'library']:
            return jsonify({'success': False, 'message': 'Admin or Library privileges required'})
        
        # Read NFC card
        try:
            clf = nfc.ContactlessFrontend('usb')
            tag = clf.connect(rdwr={'on-connect': lambda tag: False})
            
            if tag.ndef:
                # Read NDEF data
                ndef_data = tag.ndef.records[0].payload.decode('utf-8')
                student_data = json.loads(ndef_data)
                
                # Get student details from database
                student = db.session.execute(
                    db.select(Credentials).filter_by(username=student_data.get('username'))
                ).scalar()
                
                if student:
                    return jsonify({
                        'success': True,
                        'student_name': student_data.get('name', ''),
                        'student_roll_no': student_data.get('roll_no', ''),
                        'student_email': student_data.get('email', ''),
                        'student_username': student_data.get('username', '')
                    })
                else:
                    return jsonify({'success': False, 'message': 'Student not found in database'})
            else:
                return jsonify({'success': False, 'message': 'No NDEF data found on card'})
                
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error reading NFC card: {str(e)}'})
        finally:
            try:
                clf.close()
            except:
                pass
                
    except Exception as e:
        print(f"Error reading NFC for library: {e}")
        return jsonify({'success': False, 'message': 'Error reading NFC card'})

@app.route('/library/issue_book', methods=['POST'])
def issue_book():
    """Issue book to student"""
    try:
        # Check if user is logged in and is admin
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'library']:
            flash("Access denied. Admin or Library privileges required.")
            return redirect(url_for('library_dashboard'))
        
        # Get form data
        book_id = request.form.get('book_id')
        student_name = request.form.get('student_name')
        student_roll_no = request.form.get('student_roll_no')
        student_email = request.form.get('student_email')
        student_username = request.form.get('student_username')
        due_date = request.form.get('due_date')
        notes = request.form.get('notes')
        
        # Validate required fields
        if not all([book_id, student_name, student_roll_no, student_email, due_date]):
            flash("Please fill in all required fields")
            return redirect(url_for('library_dashboard'))
        
        # Check if book exists and is available
        book = db.session.execute(
            db.select(LibraryBook).filter_by(book_id=book_id)
        ).scalar()
        
        if not book:
            flash("Book not found")
            return redirect(url_for('library_dashboard'))
        
        if book.available_copies <= 0:
            flash("No copies of this book are available")
            return redirect(url_for('library_dashboard'))
        
        # Generate transaction ID
        import uuid
        transaction_id = f"LIB-{uuid.uuid4().hex[:8].upper()}"
        
        # Create transaction
        transaction = LibraryTransaction(
            transaction_id=transaction_id,
            student_username=student_username,
            student_name=student_name,
            student_roll_no=student_roll_no,
            student_email=student_email,
            book_id=book_id,
            book_title=book.title,
            book_author=book.author,
            due_date=datetime.strptime(due_date, '%Y-%m-%d'),
            issued_by=session['username'],
            notes=notes
        )
        
        # Update book availability
        book.available_copies -= 1
        
        db.session.add(transaction)
        db.session.commit()
        
        flash(f"Book '{book.title}' issued to {student_name} successfully! Transaction ID: {transaction_id}")
        return redirect(url_for('library_dashboard'))
        
    except Exception as e:
        print(f"Error issuing book: {e}")
        flash("Error issuing book")
        return redirect(url_for('library_dashboard'))

@app.route('/library/return_book', methods=['POST'])
def return_book():
    """Return book from student"""
    try:
        # Check if user is logged in and is admin
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'library']:
            flash("Access denied. Admin or Library privileges required.")
            return redirect(url_for('library_dashboard'))
        
        # Get form data
        transaction_id = request.form.get('transaction_id')
        fine_amount = request.form.get('fine_amount', 0)
        notes = request.form.get('notes')
        
        # Validate required fields
        if not transaction_id:
            flash("Please provide transaction ID")
            return redirect(url_for('library_dashboard'))
        
        # Find transaction
        transaction = db.session.execute(
            db.select(LibraryTransaction).filter_by(transaction_id=transaction_id)
        ).scalar()
        
        if not transaction:
            flash("Transaction not found")
            return redirect(url_for('library_dashboard'))
        
        if transaction.status == 'returned':
            flash("Book has already been returned")
            return redirect(url_for('library_dashboard'))
        
        # Update transaction
        transaction.status = 'returned'
        transaction.return_date = datetime.utcnow()
        transaction.returned_to = session['username']
        transaction.fine_amount = int(fine_amount) if fine_amount else 0
        transaction.notes = notes
        
        # Update book availability
        book = db.session.execute(
            db.select(LibraryBook).filter_by(book_id=transaction.book_id)
        ).scalar()
        
        if book:
            book.available_copies += 1
        
        db.session.commit()
        
        flash(f"Book '{transaction.book_title}' returned successfully!")
        return redirect(url_for('library_dashboard'))
        
    except Exception as e:
        print(f"Error returning book: {e}")
        flash("Error returning book")
        return redirect(url_for('library_dashboard'))

@app.route('/library/student_history')
def student_history():
    """View student book history"""
    try:
        # Check if user is logged in and is admin
        if 'username' not in session:
            flash("Please login first")
            return redirect(url_for('access'))
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'library']:
            flash("Access denied. Admin or Library privileges required.")
            return redirect(url_for('library_dashboard'))
        
        student_id = request.args.get('student_id')
        
        if not student_id:
            return render_template('student_history.html', student_transactions=[])
        
        # Search for student transactions
        student_transactions = db.session.execute(
            db.select(LibraryTransaction).filter(
                (LibraryTransaction.student_roll_no == student_id) |
                (LibraryTransaction.student_username == student_id)
            ).order_by(LibraryTransaction.issue_date.desc())
        ).scalars().all()
        
        return render_template('student_history.html', 
                             student_transactions=student_transactions,
                             student_id=student_id)
        
    except Exception as e:
        print(f"Error viewing student history: {e}")
        flash("Error loading student history")
        return redirect(url_for('library_dashboard'))

# Run app
@app.route('/download_combined_attendance', methods=['POST'])
def download_combined_attendance():
    """Download all date sheets from a single file combined into one sheet with date headings"""
    try:
        # Check if user is logged in and is admin or teacher
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Please login first'})
        
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return jsonify({'success': False, 'message': 'Admin or Teacher privileges required'})
        
        data = request.get_json()
        filename = data.get('filename')
        
        if not filename:
            return jsonify({'success': False, 'message': 'No filename specified'})
        
        filepath = os.path.join('static', 'classes', filename)
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'File not found'})
        
        # Check if file is locked (being used by another application)
        try:
            with open(filepath, 'r+b') as f:
                pass
        except (PermissionError, OSError) as e:
            return jsonify({'success': False, 'message': f'File is locked or inaccessible. Please close Excel or any other application that might be using this file. Error: {str(e)}'})
        
        print(f"Processing file: {filename}")
        
        # Load the workbook
        workbook = openpyxl.load_workbook(filepath)
        
        # Get all date sheets (sheets that match YYYY-MM-DD format)
        date_sheets = []
        for sheet_name in workbook.sheetnames:
            if sheet_name != 'Class Information' and re.match(r'\d{4}-\d{2}-\d{2}', sheet_name):
                date_sheets.append(sheet_name)
        
        print(f"Found date sheets: {date_sheets}")
        
        if not date_sheets:
            workbook.close()
            return jsonify({'success': False, 'message': 'No attendance data found in this file'})
        
        # Sort date sheets chronologically
        date_sheets.sort()
        print(f"Processing {len(date_sheets)} date sheets: {date_sheets}")
        
        # Create a new workbook for combined data
        combined_workbook = openpyxl.Workbook()
        combined_worksheet = combined_workbook.active
        combined_worksheet.title = "Combined Attendance"
        
        # First, add Pin Number column header
        combined_worksheet.cell(row=1, column=1, value="Pin Number")
        combined_worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        
        current_col = 2  # Start from column B
        
        # Each date will always occupy 9 columns (Period 1..9)
        PERIODS_PER_DAY = 9
        
        # Process each date sheet horizontally
        for sheet_name in date_sheets:
            worksheet = workbook[sheet_name]
            print(f"Processing sheet: {sheet_name}, rows: {worksheet.max_row}, cols: {worksheet.max_column}")
            
            # Get the date in DD-MM-YYYY format for display
            try:
                date_obj = datetime.strptime(sheet_name, '%Y-%m-%d')
                display_date = date_obj.strftime('%d-%m-%Y')
            except:
                display_date = sheet_name
            
            # Add date header merged across 9 period columns
            max_col = worksheet.max_column
            print(f"Max column for {sheet_name}: {max_col}")
            start_col = current_col
            end_col = current_col + PERIODS_PER_DAY - 1
            top_left = combined_worksheet.cell(row=1, column=start_col)
            top_left.value = display_date
            top_left.font = openpyxl.styles.Font(bold=True, size=12)
            top_left.alignment = openpyxl.styles.Alignment(horizontal='center')
            combined_worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            # Write subject headers for Period 1..9 using source sheet headers if available
            for period_index in range(PERIODS_PER_DAY):
                source_col = 2 + period_index  # Source period columns start at B (2)
                header_val = None
                if source_col <= max_col:
                    header_val = worksheet.cell(row=1, column=source_col).value
                header_text = header_val if (header_val is not None and str(header_val).strip() != "") else f"Period {period_index + 1}"
                header_cell = combined_worksheet.cell(row=2, column=start_col + period_index)
                header_cell.value = header_text
                header_cell.font = openpyxl.styles.Font(bold=True)

            current_col = end_col + 1
        
        # Now copy the data rows
        # Get the maximum number of data rows across all sheets
        max_data_rows = 0
        for sheet_name in date_sheets:
            worksheet = workbook[sheet_name]
            data_rows = worksheet.max_row - 1  # Subtract header row
            max_data_rows = max(max_data_rows, data_rows)
        
        print(f"Maximum data rows: {max_data_rows}")
        
        # Copy data for each row, always writing 9 period columns per sheet
        for data_row in range(1, max_data_rows + 1):
            current_col = 2  # Reset to column B for each row
            
            # Process each date sheet for this row
            for sheet_name in date_sheets:
                worksheet = workbook[sheet_name]
                source_row = data_row + 1  # +1 because data starts from row 2
                
                # Copy Pin Number from first sheet only
                if sheet_name == date_sheets[0]:
                    pin_value = worksheet.cell(row=source_row, column=1).value
                    if pin_value is not None:
                        combined_worksheet.cell(row=data_row + 2, column=1, value=pin_value)  # +2 for header rows
                
                # Copy data for fixed 9 periods; leave blank if no data
                max_col = worksheet.max_column
                for period_index in range(PERIODS_PER_DAY):
                    source_col = 2 + period_index
                    cell_value = None
                    if source_col <= max_col:
                        cell_value = worksheet.cell(row=source_row, column=source_col).value
                    if cell_value is not None:
                        combined_worksheet.cell(row=data_row + 2, column=current_col, value=cell_value)
                    current_col += 1
        
        workbook.close()
        
        # Check if we have any data
        if max_data_rows <= 0:
            print(f"No data found in any sheets for {filename}")
            return jsonify({'success': False, 'message': 'No attendance data found in any sheets'})
        
        print(f"Total data rows processed: {max_data_rows}")
        
        # Set basic column widths
        combined_worksheet.column_dimensions['A'].width = 15  # Pin Number column
        for col in range(2, 11):  # Set width for columns B through J
            column_letter = get_column_letter(col)
            combined_worksheet.column_dimensions[column_letter].width = 12
        
        # Create response
        try:
            output = io.BytesIO()
            print("Saving workbook to BytesIO...")
            combined_workbook.save(output)
            output.seek(0)
            
            file_content = output.getvalue()
            print(f"Successfully created combined file for {filename}, size: {len(file_content)} bytes")
            
            if file_content and len(file_content) > 0:
                return send_file(
                    output,
                    as_attachment=True,
                    download_name=f'Combined_{filename}',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                print("Error: Generated file is empty")
                return jsonify({'success': False, 'message': 'Generated file is empty'})
                
        except Exception as save_error:
            print(f"Error saving workbook: {save_error}")
            import traceback
            traceback.print_exc()
            
            # Try creating a simple fallback file
            try:
                fallback_workbook = openpyxl.Workbook()
                fallback_sheet = fallback_workbook.active
                fallback_sheet['A1'] = 'Combined Attendance'
                fallback_sheet['A2'] = f'File: {filename}'
                fallback_sheet['A3'] = f'Date sheets: {len(date_sheets)}'
                fallback_sheet['A4'] = f'Error: {str(save_error)}'
                
                fallback_output = io.BytesIO()
                fallback_workbook.save(fallback_output)
                fallback_output.seek(0)
                
                return send_file(
                    fallback_output,
                    as_attachment=True,
                    download_name=f'Error_{filename}',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except:
                return jsonify({'success': False, 'message': f'Error saving Excel file: {str(save_error)}'})
        
    except Exception as e:
        print(f"Error in download_combined_attendance: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/download_filtered_attendance', methods=['POST'])
def download_filtered_attendance():
    """Download only the selected date range from a file, combined with 9 periods per day"""
    try:
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Please login first'})
        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user or user.role not in ['admin', 'teacher']:
            return jsonify({'success': False, 'message': 'Admin or Teacher privileges required'})

        data = request.get_json() or {}
        filename = data.get('filename')
        from_date = data.get('from_date')
        to_date = data.get('to_date')

        if not filename:
            return jsonify({'success': False, 'message': 'No filename specified'})

        filepath = os.path.join('static', 'classes', filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'File not found'})

        try:
            with open(filepath, 'r+b') as f:
                pass
        except (PermissionError, OSError) as e:
            return jsonify({'success': False, 'message': f'File is locked or inaccessible. Please close Excel or any other application that might be using this file. Error: {str(e)}'})

        workbook = openpyxl.load_workbook(filepath)

        # Build candidate date sheets
        date_sheets = []
        for sheet_name in workbook.sheetnames:
            if sheet_name != 'Class Information' and re.match(r'\d{4}-\d{2}-\d{2}', sheet_name):
                # Filter by range if provided
                in_range = True
                if from_date:
                    in_range = in_range and (sheet_name >= from_date)
                if to_date:
                    in_range = in_range and (sheet_name <= to_date)
                if in_range:
                    date_sheets.append(sheet_name)

        if not date_sheets:
            workbook.close()
            return jsonify({'success': False, 'message': 'No dates found for the selected range'})

        date_sheets.sort()

        # Create output workbook identical to combined endpoint logic
        combined_workbook = openpyxl.Workbook()
        combined_worksheet = combined_workbook.active
        combined_worksheet.title = 'Combined Attendance'

        combined_worksheet.cell(row=1, column=1, value='Pin Number')
        combined_worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)

        PERIODS_PER_DAY = 9
        current_col = 2

        for sheet_name in date_sheets:
            worksheet = workbook[sheet_name]
            try:
                date_obj = datetime.strptime(sheet_name, '%Y-%m-%d')
                display_date = date_obj.strftime('%d-%m-%Y')
            except:
                display_date = sheet_name
            max_col = worksheet.max_column

            start_col = current_col
            end_col = current_col + PERIODS_PER_DAY - 1
            top_left = combined_worksheet.cell(row=1, column=start_col)
            top_left.value = display_date
            top_left.font = openpyxl.styles.Font(bold=True, size=12)
            top_left.alignment = openpyxl.styles.Alignment(horizontal='center')
            combined_worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            for period_index in range(PERIODS_PER_DAY):
                source_col = 2 + period_index
                header_val = None
                if source_col <= max_col:
                    header_val = worksheet.cell(row=1, column=source_col).value
                header_text = header_val if (header_val is not None and str(header_val).strip() != '') else f'Period {period_index + 1}'
                header_cell = combined_worksheet.cell(row=2, column=start_col + period_index)
                header_cell.value = header_text
                header_cell.font = openpyxl.styles.Font(bold=True)

            current_col = end_col + 1

        # Determine max rows
        max_data_rows = 0
        for sheet_name in date_sheets:
            worksheet = workbook[sheet_name]
            max_data_rows = max(max_data_rows, worksheet.max_row - 1)

        for data_row in range(1, max_data_rows + 1):
            current_col = 2
            for sheet_name in date_sheets:
                worksheet = workbook[sheet_name]
                source_row = data_row + 1
                if sheet_name == date_sheets[0]:
                    pin_value = worksheet.cell(row=source_row, column=1).value
                    if pin_value is not None:
                        combined_worksheet.cell(row=data_row + 2, column=1, value=pin_value)
                max_col = worksheet.max_column
                for period_index in range(PERIODS_PER_DAY):
                    source_col = 2 + period_index
                    val = None
                    if source_col <= max_col:
                        val = worksheet.cell(row=source_row, column=source_col).value
                    if val is not None:
                        combined_worksheet.cell(row=data_row + 2, column=current_col, value=val)
                    current_col += 1

        workbook.close()

        # Basic widths
        combined_worksheet.column_dimensions['A'].width = 15
        for col in range(2, 2 + 9 * len(date_sheets)):
            combined_worksheet.column_dimensions[get_column_letter(col)].width = 12

        output = io.BytesIO()
        combined_workbook.save(output)
        output.seek(0)
        file_label = f"{filename.split('.')[0]}_{from_date or 'start'}_to_{to_date or 'end'}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=f'Filtered_{file_label}',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Error in download_filtered_attendance: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# Mock Exam Routes
@app.route('/mock_exam')
def mock_exam():
    """Mock exam dashboard for students"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('student'))
    
    # Get student's exam history
    exam_history = db.session.execute(
        db.select(MockExam).filter_by(student_username=session['username'])
        .order_by(MockExam.created_at.desc())
    ).scalars().all()
    
    return render_template('mock_exam.html', user=user, exam_history=exam_history)

@app.route('/test_gemini')
def test_gemini():
    """Test Gemini API connection"""
    try:
        test_prompt = "Generate a simple multiple choice question about mathematics in JSON format: {\"question\": \"What is 2+2?\", \"options\": [\"3\", \"4\", \"5\", \"6\"], \"correct_answer\": \"4\"}"
        response = gemini_generate_text(test_prompt)
        return f"Gemini API Test: {response}"
    except Exception as e:
        return f"Gemini API Error: {str(e)}"

@app.route('/test_models')
def test_models():
    """Test available Gemini models"""
    try:
        import requests
        
        # First, try to list available models
        list_url = f"https://generativelanguage.googleapis.com/v1beta/models?key={GEMINI_API_KEY}"
        try:
            resp = requests.get(list_url, timeout=10)
            if resp.status_code == 200:
                models_data = resp.json()
                available_models = []
                if 'models' in models_data:
                    for model in models_data['models']:
                        if 'generateContent' in model.get('supportedGenerationMethods', []):
                            available_models.append(model['name'].replace('models/', ''))
                
                return f"<pre>Available Models:\n{chr(10).join(available_models[:10])}</pre>"
        except Exception as e:
            pass
        
        # Fallback: Test common model names and API versions
        test_endpoints = [
            "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent",
            "https://generativelanguage.googleapis.com/v1/models/gemini-1.5-flash:generateContent",
            "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-002:generateContent",
            "https://generativelanguage.googleapis.com/v1/models/gemini-1.5-flash-002:generateContent",
        ]
        
        results = []
        for endpoint in test_endpoints:
            try:
                test_prompt = "Say 'Hello' if you can read this."
                
                payload = {"contents": [{"parts": [{"text": test_prompt}]}]}
                headers = {"Content-Type": "application/json"}
                
                resp = requests.post(f"{endpoint}?key={GEMINI_API_KEY}", 
                                   headers=headers, data=json.dumps(payload), timeout=10)
                
                model_name = endpoint.split('/')[-1].replace(':generateContent', '')
                if resp.status_code == 200:
                    results.append(f"✓ {model_name}: Working")
                else:
                    results.append(f"✗ {model_name}: HTTP {resp.status_code}")
            except Exception as e:
                model_name = endpoint.split('/')[-1].replace(':generateContent', '')
                results.append(f"✗ {model_name}: {str(e)[:50]}")
        
        return f"<pre>Model Status:\n{chr(10).join(results)}</pre>"
    except Exception as e:
        return f"Error testing models: {str(e)}"

@app.route('/test_question_generation')
def test_question_generation():
    """Test question generation with sample data"""
    try:
        # Test with sample data
        subject = "Web Technologies"
        syllabus = "HTML, CSS, JavaScript, DOM manipulation, responsive design"
        question_format = "section_based"
        total_questions = 5
        total_marks = 25
        difficulty_level = "medium"
        format_description = "Section A: 3 short answer questions, 2 marks each\nSection B: 2 descriptive questions, 5 marks each"
        
        print(f"Testing with: Subject={subject}, Syllabus={syllabus}, Format={format_description}")
        
        result = generate_questions_with_gemini(
            subject, syllabus, question_format, total_questions, total_marks, difficulty_level, format_description
        )
        
        if result:
            return f"<pre>Question Generation Test Success:\n{json.dumps(result, indent=2)}</pre>"
        else:
            return "Question Generation Test Failed - Check console logs"
            
    except Exception as e:
        return f"Test Error: {str(e)}"

@app.route('/test_simple_generation')
def test_simple_generation():
    """Test simple question generation"""
    try:
        prompt = """Generate 2 questions about HTML and CSS in JSON format:

{
  "questions": [
    {
      "id": 1,
      "type": "short_answer",
      "question": "What is HTML?",
      "marks": 2
    },
    {
      "id": 2,
      "type": "descriptive", 
      "question": "Explain the difference between HTML and CSS",
      "marks": 5
    }
  ]
}

Respond with ONLY the JSON, no other text."""
        
        response_text = gemini_generate_text(prompt)
        return f"<pre>Simple Test Response:\n{response_text}</pre>"
        
    except Exception as e:
        return f"Simple Test Error: {str(e)}"

@app.route('/test_professional_exam')
def test_professional_exam():
    """Test professional exam generation"""
    try:
        subject_name = "Web Technologies"
        syllabus_content = "HTML, CSS, JavaScript, DOM manipulation, responsive design, web accessibility"
        question_format = "mixed"
        number_of_questions = 9
        total_marks = 50
        question_paper_format = "Section A: 5 Short Answer Questions (2 marks each), Section B: 4 Long Answer Questions (10 marks each)"
        time_limit = 120
        difficulty_level = "medium"
        
        result = generate_exam_with_custom_prompt(
            subject_name, syllabus_content, question_format, number_of_questions, 
            total_marks, question_paper_format, time_limit, difficulty_level
        )
        
        if result:
            return f"<pre>Professional Exam Test Success:\n{json.dumps(result, indent=2)}</pre>"
        else:
            return "Professional Exam Test Failed - Check console logs"
            
    except Exception as e:
        return f"Professional Exam Test Error: {str(e)}"

@app.route('/test_real_questions')
def test_real_questions():
    """Test if we can generate real questions"""
    try:
        prompt = """Generate 2 REAL questions about HTML and CSS. Do NOT use generic placeholders.

Syllabus: HTML, CSS, JavaScript, DOM manipulation, responsive design

ABSOLUTELY CRITICAL:
- DO NOT generate questions like "Question 1 about HTML" or "Descriptive question 1 about Web Technologies"
- Create REAL, SPECIFIC questions about HTML and CSS
- Use the syllabus topics to create meaningful questions

Examples of what to do:
- "Explain the difference between HTML4 and HTML5 and list three new semantic elements"
- "Describe the CSS box model and explain how margin, padding, and border work together"

Examples of what NOT to do:
- "Question 1 about HTML" ❌
- "Descriptive question 1 about Web Technologies" ❌

Respond with ONLY this JSON format:
{
  "questions": [
    {
      "id": 1,
      "type": "short_answer",
      "question": "REAL question about HTML here",
      "marks": 2
    },
    {
      "id": 2,
      "type": "descriptive",
      "question": "REAL question about CSS here", 
      "marks": 5
    }
  ]
}

Respond with ONLY the JSON, no other text."""
        
        response_text = gemini_generate_text(prompt)
        return f"<pre>Real Questions Test:\n{response_text}</pre>"
        
    except Exception as e:
        return f"Real Questions Test Error: {str(e)}"

@app.route('/debug_gemini', methods=['GET', 'POST'])
def debug_gemini():
    """Debug interface to test Gemini API with custom inputs"""
    if request.method == 'POST':
        try:
            subject = request.form.get('subject', 'Web Technologies')
            syllabus = request.form.get('syllabus', 'HTML, CSS, JavaScript')
            question_format = request.form.get('question_format', 'section_based')
            total_questions = int(request.form.get('total_questions', 5))
            total_marks = int(request.form.get('total_marks', 25))
            difficulty_level = request.form.get('difficulty_level', 'medium')
            format_description = request.form.get('format_description', 'Section A: 3 short answer questions, 2 marks each\nSection B: 2 descriptive questions, 5 marks each')
            
            # Generate the prompt that will be sent to Gemini
            # Use the REST-based helper for compatibility
            
            # Calculate marks per question
            marks_per_question = total_marks // total_questions
            remaining_marks = total_marks % total_questions
            
            # Create prompt based on question format and difficulty
            if question_format == 'multiple_choice':
                format_instruction = f"Generate {total_questions} multiple choice questions with 4 options each (A, B, C, D) and mark the correct answer. Each question should be worth {marks_per_question} marks."
            elif question_format == 'descriptive':
                format_instruction = f"Generate {total_questions} descriptive/long answer questions that require detailed explanations. Each question should be worth {marks_per_question} marks."
            elif question_format == 'section_based':
                format_instruction = f"""Generate questions based on the following custom format description:

{format_description}

IMPORTANT INSTRUCTIONS FOR SECTION-BASED FORMAT:
1. Parse the format description carefully to understand the exact structure
2. Create sections with the exact names specified (e.g., "Section A", "Part I", etc.)
3. Generate the exact number of questions for each section as specified
4. Use the exact marks per question as specified in the format
5. For "short answer" questions, use type "short_answer" 
6. For "long answer" or "descriptive" questions, use type "descriptive"
7. For "multiple choice" questions, use type "multiple_choice"
8. Follow the format description word-for-word
9. Ensure all sections and questions match the specified structure exactly
10. Generate REAL questions based on the syllabus content, NOT sample or generic questions
11. Each question must be directly related to the syllabus topics provided
12. Questions should test actual understanding of the subject matter"""
            else:  # mixed
                mcq_count = int(total_questions * 0.6)
                desc_count = total_questions - mcq_count
                mcq_marks = int(total_marks * 0.4)
                desc_marks = total_marks - mcq_marks
                format_instruction = f"Generate {mcq_count} multiple choice questions (worth {mcq_marks//mcq_count} marks each) and {desc_count} descriptive questions (worth {desc_marks//desc_count} marks each)."
            
            # Difficulty instruction
            if difficulty_level == 'easy':
                difficulty_instruction = "Questions should be straightforward and test basic understanding."
            elif difficulty_level == 'medium':
                difficulty_instruction = "Questions should be moderately challenging and test application of concepts."
            else:  # hard
                difficulty_instruction = "Questions should be challenging and test deep understanding and analysis."
            
            prompt = f"""You are an expert question paper generator. Generate questions based on the following specifications:

Subject: {subject}
Syllabus: {syllabus}
Total Questions: {total_questions}
Total Marks: {total_marks}
Difficulty Level: {difficulty_level}

{format_instruction}
{difficulty_instruction}

CRITICAL REQUIREMENTS:
1. Generate REAL questions based on the syllabus content provided above
2. Do NOT generate generic or sample questions
3. Questions must be directly related to the syllabus topics
4. Use the actual subject matter from the syllabus to create meaningful questions
5. Each question should test understanding of the specific topics mentioned in the syllabus

IMPORTANT: You must respond with ONLY valid JSON. Do not include any explanations, markdown formatting, or additional text.

For section-based format, use this exact JSON structure:
{{
    "sections": [
        {{
            "section_name": "Section A",
            "section_type": "short_answer",
            "questions": [
                {{
                    "id": 1,
                    "type": "short_answer",
                    "question": "What is the capital of France?",
                    "marks": 2
                }}
            ]
        }},
        {{
            "section_name": "Section B", 
            "section_type": "descriptive",
            "questions": [
                {{
                    "id": 2,
                    "type": "descriptive",
                    "question": "Explain the process of photosynthesis in detail.",
                    "marks": 5
                }}
            ]
        }}
    ]
}}

Requirements:
1. Questions must be relevant to the syllabus content
2. Difficulty level: {difficulty_level}
3. Questions should be clear and unambiguous
4. Cover different topics from the syllabus
5. Test understanding, not just memorization
6. For multiple choice questions, provide exactly 4 options (A, B, C, D format)
7. For descriptive questions, do not include options or correct_answer fields
8. For short answer questions, do not include options or correct_answer fields
9. Ensure all required fields are present in each question
10. For section-based format, follow the format description exactly
11. Use appropriate question types: "multiple_choice", "short_answer", "descriptive"
12. Generate questions that match the specified marks and difficulty
13. CRITICAL: Generate REAL questions based on the syllabus, NOT sample or generic questions
14. Each question must test actual knowledge of the subject matter provided in the syllabus
15. Questions should be specific to the topics mentioned in the syllabus content
16. Use the exact syllabus topics to create meaningful, specific questions
17. Do NOT use generic placeholders or sample questions
18. Create questions that test understanding of the specific syllabus topics provided

Respond with ONLY the JSON, no other text."""
            
            # Send to Gemini and get response
            response_text = gemini_generate_text(prompt)
            
            return render_template('debug_gemini_result.html', 
                                 prompt=prompt, 
                                 response=response_text,
                                 subject=subject,
                                 syllabus=syllabus,
                                 question_format=question_format,
                                 total_questions=total_questions,
                                 total_marks=total_marks,
                                 difficulty_level=difficulty_level,
                                 format_description=format_description)
            
        except Exception as e:
            return f"Error: {str(e)}"
    
    return render_template('debug_gemini.html')

@app.route('/debug_evaluation', methods=['GET', 'POST'])
def debug_evaluation():
    """Debug interface to test answer evaluation with Gemini"""
    if request.method == 'POST':
        try:
            subject = request.form.get('subject', 'Web Technologies')
            questions_text = request.form.get('questions', '')
            answers_text = request.form.get('answers', '')
            
            # Parse questions and answers
            questions = json.loads(questions_text) if questions_text else {}
            answers = json.loads(answers_text) if answers_text else {}
            
            # Generate the evaluation prompt
            # Use the REST-based helper for compatibility
            
            evaluation_data = {
                "subject": subject,
                "questions": questions,
                "student_answers": answers
            }
            
            prompt = f"""You are an expert examiner. Evaluate the student's answers for the following questions and provide comprehensive feedback.

Subject: {subject}

Questions and Student Answers:
{json.dumps(evaluation_data, indent=2)}

EVALUATION CRITERIA:
1. For multiple choice questions: Check if the selected answer is correct
2. For short answer questions: Evaluate accuracy, completeness, and relevance
3. For descriptive questions: Assess understanding, depth of explanation, and accuracy
4. Award partial marks for partially correct answers
5. Provide constructive feedback for improvement

For each question, provide:
1. Marks awarded (out of total marks)
2. Detailed feedback on the answer quality
3. Specific suggestions for improvement
4. Correct answer (for reference)

Return the evaluation in the following JSON format:
{{
    "total_marks": 100,
    "obtained_marks": 85,
    "percentage": 85.0,
    "grade": "A",
    "questions_attempted": 9,
    "correct_answers": 7,
    "incorrect_answers": 2,
    "accuracy": 77.8,
    "time_taken": 45,
    "time_efficiency": 75.0,
    "avg_time_per_question": 5.0,
    "difficulty_level": "medium",
    "detailed_feedback": [
        {{
            "question_id": 1,
            "question_type": "short_answer",
            "marks_awarded": 4,
            "total_marks": 5,
            "feedback": "Good understanding of the concept, but missing some key details.",
            "suggestions": "Review the specific topic mentioned in the syllabus.",
            "correct_answer": "Expected correct answer for reference"
        }}
    ]
}}

Grading Scale:
- A+: 90-100%
- A: 80-89%
- B+: 70-79%
- B: 60-69%
- C+: 50-59%
- C: 40-49%
- D: 30-39%
- F: Below 30%

Be fair, thorough, and constructive in your evaluation. Consider partial credit for partially correct answers."""
            
            # Send to Gemini and get response
            response_text = gemini_generate_text(prompt)
            
            return render_template('debug_evaluation_result.html', 
                                 prompt=prompt, 
                                 response=response_text,
                                 subject=subject,
                                 questions=questions,
                                 answers=answers)
            
        except Exception as e:
            return f"Error: {str(e)}"
    
    return render_template('debug_evaluation.html')

# ---------------------- Student Chatbot ----------------------
@app.route('/student_chat')
def student_chat():
    if 'username' not in session:
        return redirect(url_for('login_signup'))
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    return render_template('student_chat.html', user=user)

@app.route('/api/student_chat', methods=['POST'])
def api_student_chat():
    try:
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401

        data = request.get_json() or {}
        user_message = (data.get('message') or '').strip()
        subject = (data.get('subject') or '').strip()
        if not user_message:
            return jsonify({'success': False, 'message': 'Message is required'}), 400

        # Check for summary request
        if any(keyword in user_message.lower() for keyword in ['summary', 'summarize', 'generate pdf', 'create pdf', 'download chat']):
            return handle_chat_summary_request()

        # Reject non-educational prompts
        if not is_educational_query(user_message):
            return jsonify({
                'success': False,
                'message': 'This chatbot is for education and college app help only. Please ask study or app-related questions.'
            }), 400

        # Check if user has uploaded PDFs
        pdf_context = ""
        uploaded_pdfs = session.get('uploaded_pdfs', [])
        
        if uploaded_pdfs:
            pdf_context = "\n\nIMPORTANT: The student has uploaded PDF documents. You MUST answer based ONLY on the content from these PDFs. Here are the PDF summaries:\n\n"
            for i, pdf_info in enumerate(uploaded_pdfs, 1):
                pdf_context += f"PDF {i}: {pdf_info['original_name']}\n"
                pdf_context += f"Content Summary: {pdf_info['summary']}\n\n"
            
            pdf_context += "IMPORTANT RULES:\n"
            pdf_context += "1. Answer the student's question based ONLY on the information provided in these PDFs.\n"
            pdf_context += "2. If the question is about ByteCredits app features (payments, NFC, exams, learning pods), you can answer normally.\n"
            pdf_context += "3. If the question is academic and cannot be answered from the PDF content, politely explain that the answer is not available in the uploaded documents and suggest they ask a more specific question about the PDF content.\n"
            pdf_context += "4. Always prioritize PDF content for academic questions over general knowledge."

        system_preamble = (
            "You are ByteCredits Student Assistant for EDUCATIONAL use only. "
            "Strictly refuse entertainment or non-educational requests (jokes, roasts, songs, memes, poems, stories). "
            "Answer clearly and concisely. If the question is about this college app (payments, NFC, exams, learning pods), provide actionable steps. "
            "If the user asks academic doubts, give an explanation with examples. Keep answers under 200 words unless asked to elaborate."
        )

        prompt = f"""{system_preamble}
{pdf_context}

Student: {user_message}
{('Subject context: ' + subject) if subject else ''}

Provide a helpful answer suitable for a student. If calculations or code are involved, show steps briefly.
"""

        # Use REST-based helper for compatibility with older SDK environment
        answer_text = gemini_generate_text(prompt).strip()
        
        # Store chat history
        store_chat_message(user_message, True, subject)
        store_chat_message(answer_text, False, subject)
        
        return jsonify({'success': True, 'answer': answer_text})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Chat error: {str(e)}'}), 500

def store_chat_message(content: str, is_user: bool, subject: str = ""):
    """Store a chat message in session"""
    if 'chat_history' not in session:
        session['chat_history'] = []
    
    message = {
        'content': content,
        'isUser': is_user,
        'subject': subject,
        'timestamp': datetime.now().strftime('%H:%M:%S')
    }
    
    session['chat_history'].append(message)
    # Keep only last 50 messages to avoid session bloat
    if len(session['chat_history']) > 50:
        session['chat_history'] = session['chat_history'][-50:]
    
    session.modified = True

def handle_chat_summary_request():
    """Handle chat summary generation request"""
    try:
        chat_history = session.get('chat_history', [])
        
        if not chat_history:
            return jsonify({
                'success': False, 
                'message': 'No chat history available to summarize. Please have a conversation first.'
            }), 400
        
        # Generate PDF summary
        pdf_path = create_chat_summary_pdf(chat_history, session['username'])
        
        if not pdf_path:
            return jsonify({
                'success': False, 
                'message': 'Failed to generate PDF summary. Please try again.'
            }), 500
        
        # Generate text summary for immediate response
        text_summary = generate_chat_summary(chat_history)
        
        return jsonify({
            'success': True, 
            'answer': f"📄 Chat summary generated successfully!\n\n{text_summary}\n\n✅ PDF has been created and saved. You can download it from the PDF section above.",
            'pdf_generated': True,
            'pdf_filename': os.path.basename(pdf_path)
        })
        
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'Error generating summary: {str(e)}'
        }), 500

@app.route('/api/upload_pdf', methods=['POST'])
def api_upload_pdf():
    """Upload PDF for chatbot context"""
    try:
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401

        if 'pdf_file' not in request.files:
            return jsonify({'success': False, 'message': 'No PDF file provided'}), 400

        file = request.files['pdf_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400

        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'message': 'Only PDF files are allowed'}), 400

        # Create unique filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{session['username']}_{timestamp}_{file.filename}"
        file_path = os.path.join(app.config['PDF_UPLOAD_FOLDER'], filename)

        # Save file
        file.save(file_path)

        # Process PDF and create summary
        pdf_summary = process_pdf_for_chatbot(file_path)
        
        if not pdf_summary:
            # Remove file if processing failed
            os.remove(file_path)
            return jsonify({'success': False, 'message': 'Failed to process PDF content'}), 400

        # Store PDF info in session
        if 'uploaded_pdfs' not in session:
            session['uploaded_pdfs'] = []
        
        pdf_info = {
            'filename': filename,
            'original_name': file.filename,
            'file_path': file_path,
            'summary': pdf_summary,
            'uploaded_at': datetime.now().isoformat()
        }
        
        session['uploaded_pdfs'].append(pdf_info)
        session.modified = True

        return jsonify({
            'success': True, 
            'message': f'PDF "{file.filename}" uploaded successfully',
            'summary': pdf_summary[:200] + '...' if len(pdf_summary) > 200 else pdf_summary
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Upload error: {str(e)}'}), 500

@app.route('/api/get_pdf_info', methods=['GET'])
def api_get_pdf_info():
    """Get information about uploaded PDFs"""
    try:
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401

        pdfs = session.get('uploaded_pdfs', [])
        
        # Return simplified PDF info for frontend
        pdf_info = []
        for pdf in pdfs:
            pdf_info.append({
                'original_name': pdf.get('original_name', 'Unknown'),
                'uploaded_at': pdf.get('uploaded_at', ''),
                'summary_preview': pdf.get('summary', '')[:100] + '...' if len(pdf.get('summary', '')) > 100 else pdf.get('summary', '')
            })
        
        return jsonify({
            'success': True,
            'pdfs': pdf_info
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error getting PDF info: {str(e)}'}), 500

@app.route('/api/clear_pdfs', methods=['POST'])
def api_clear_pdfs():
    """Clear all uploaded PDFs from session"""
    try:
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401

        # Remove PDF files from disk
        if 'uploaded_pdfs' in session:
            for pdf_info in session['uploaded_pdfs']:
                try:
                    if os.path.exists(pdf_info['file_path']):
                        os.remove(pdf_info['file_path'])
                except Exception as e:
                    print(f"Error removing PDF file: {e}")
            
            session['uploaded_pdfs'] = []
            session.modified = True

        return jsonify({'success': True, 'message': 'All PDFs cleared successfully'})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Clear error: {str(e)}'}), 500


@app.route('/api/download_chat_summary/<filename>')
def download_chat_summary(filename):
    """Download generated chat summary PDF"""
    try:
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401

        # Verify filename belongs to current user
        if not filename.startswith(f"chat_summary_{session['username']}_"):
            return jsonify({'success': False, 'message': 'Access denied'}), 403

        file_path = os.path.join(app.config['PDF_UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'File not found'}), 404

        return send_file(file_path, as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'success': False, 'message': f'Download error: {str(e)}'}), 500

# ---------------------- Student Profiles (Digital Resume) ----------------------

def _ensure_student_profile(username: str) -> StudentProfile:
    profile = db.session.execute(db.select(StudentProfile).filter_by(username=username)).scalar()
    if not profile:
        profile = StudentProfile(username=username)
        db.session.add(profile)
        db.session.commit()
    return profile

@app.route('/student/profile', methods=['GET', 'POST'])
def student_profile_edit():
    if 'username' not in session:
        return redirect(url_for('login'))
    # Runtime safety: ensure columns exist before ORM selects
    try:
        with db.engine.connect() as conn:
            result = conn.execute(db.text("PRAGMA table_info(student_profiles)"))
            sp_columns = [row[1] for row in result.fetchall()]
            if 'resume_json' not in sp_columns:
                conn.execute(db.text("ALTER TABLE student_profiles ADD COLUMN resume_json TEXT"))
                conn.commit()
            if 'is_resume_public' not in sp_columns:
                conn.execute(db.text("ALTER TABLE student_profiles ADD COLUMN is_resume_public BOOLEAN DEFAULT 0"))
                conn.commit()
            if 'public_slug' not in sp_columns:
                conn.execute(db.text("ALTER TABLE student_profiles ADD COLUMN public_slug VARCHAR(64)"))
                conn.commit()
    except Exception as e:
        print(f"Student profile column ensure error: {e}")
    profile = _ensure_student_profile(session['username'])
    # Load current credentials for cohort fields
    user_creds = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if request.method == 'POST':
        try:
            data = request.get_json() or {}
            # Store raw JSON blob from client editor
            profile.resume_json = json.dumps(data.get('resume') or {})
            profile.is_resume_public = bool(data.get('is_public', False))
            if profile.is_resume_public and not profile.public_slug:
                profile.public_slug = uuid.uuid4().hex[:16]
            db.session.commit()
            return jsonify({'success': True, 'public_url': url_for('student_profile_view', slug=profile.public_slug, _external=True) if profile.is_resume_public and profile.public_slug else None})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 400
    # GET
    resume = {}
    try:
        if profile.resume_json:
            resume = json.loads(profile.resume_json)
            # Normalize nested collections for the editor (keep as raw strings if originally strings)
            for key in ['education', 'experience', 'projects']:
                val = resume.get(key)
                if isinstance(val, list):
                    # keep as pretty JSON string for textarea
                    try:
                        resume[key] = json.dumps(val, ensure_ascii=False, indent=0)
                    except Exception:
                        pass
    except Exception:
        resume = {}
    return render_template('student_profile_edit.html', profile=profile, resume=resume, user=user_creds)

@app.route('/student/update_cohort', methods=['POST'])
def student_update_cohort():
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Please login first'}), 401
    try:
        data = request.get_json() or {}
        year = data.get('study_year')
        department = (data.get('department') or '').strip() or None
        section = (data.get('section') or '').strip() or None

        user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404

        # Normalize and save
        try:
            user.study_year = int(year) if year not in (None, '') else None
        except Exception:
            user.study_year = None
        user.department = department.upper() if department else None
        user.section = section.upper() if section else None
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/p/<slug>')
def student_profile_view(slug):
    if not slug:
        abort(404)
    profile = db.session.execute(db.select(StudentProfile).filter_by(public_slug=slug)).scalar()
    if not profile or not profile.is_resume_public:
        abort(404)
    resume = {}
    try:
        if profile.resume_json:
            resume = json.loads(profile.resume_json)
            # Normalize for rendering: ensure lists are lists of dicts
            def _to_list(val):
                if isinstance(val, str):
                    try:
                        parsed = json.loads(val)
                        if isinstance(parsed, list):
                            return parsed
                        return []
                    except Exception:
                        # Try to parse multiple JSON arrays/objects line by line
                        items: list = []
                        for part in [p.strip() for p in re.split(r"\n+", val) if p.strip()]:
                            try:
                                piece = json.loads(part)
                                if isinstance(piece, list):
                                    items.extend(piece)
                                elif isinstance(piece, dict):
                                    items.append(piece)
                                else:
                                    items.append({"summary": str(piece)})
                            except Exception:
                                # Last resort: treat the line as plain text summary
                                items.append({"summary": part})
                        return items
                if isinstance(val, list):
                    # If list of strings, map to dicts
                    if all(isinstance(x, str) for x in val):
                        return [{"summary": x} for x in val if x.strip()]
                    return val
                return []
            resume['education'] = _to_list(resume.get('education'))
            resume['experience'] = _to_list(resume.get('experience'))
            resume['projects'] = _to_list(resume.get('projects'))
            resume['certificates'] = _to_list(resume.get('certificates'))
            resume['achievements'] = _to_list(resume.get('achievements'))
    except Exception:
        resume = {}
    return render_template('student_profile_public.html', profile=profile, resume=resume)

def generate_exam_with_custom_prompt(subject_name, syllabus_content, question_format, number_of_questions, total_marks, question_paper_format, time_limit, difficulty_level):
    """Generate exam using the custom prompt provided by user via REST helper."""
    try:
        custom_prompt = f"""You are an expert university exam paper setter. Create a COMPLETE question paper strictly based on the syllabus and format below.

SUBJECT: {subject_name}
SYLLABUS: {syllabus_content}
TOTAL MARKS: {total_marks}
TIME LIMIT: {time_limit} minutes
DIFFICULTY: {difficulty_level}
QUESTION FORMAT: {question_format}
PAPER STRUCTURE: {question_paper_format}

HARD CONSTRAINTS:
- Use ONLY the topics present in the syllabus above; each question must reference a concrete topic from the syllabus.
- Follow the PAPER STRUCTURE exactly (section names, marks breakup, counts).
- Total number of questions across all sections should be close to {number_of_questions} (respect structure where provided).
- Sum of marks across all questions MUST equal {total_marks}.
- Output MUST be STRICT JSON only, no markdown.

ABSOLUTELY CRITICAL - READ THIS CAREFULLY:
1. You MUST generate REAL, SPECIFIC questions based on the syllabus content: {syllabus_content}
2. NEVER use generic placeholders like "Question 1 about [subject]" or "Descriptive question 1 about WT"
3. Each question MUST be about specific topics from the syllabus
4. Use the exact syllabus topics to create meaningful, detailed questions
5. Questions should test actual knowledge of the specific syllabus topics
6. Create professional exam paper format with proper sections
7. Include clear instructions for each section

EXAMPLES OF WHAT TO DO:
- If syllabus mentions "HTML", create questions like "Explain the difference between HTML4 and HTML5 and list three new features in HTML5"
- If syllabus mentions "CSS", create questions like "Describe the CSS box model and explain how margin, padding, and border properties work together"
- If syllabus mentions "JavaScript", create questions like "Write a JavaScript function to validate an email address and explain how it works"

EXAMPLES OF WHAT NOT TO DO:
- "Descriptive question 1 about Web Technologies" ❌
- "Question 1 about HTML" ❌
- "Short answer question 1" ❌

IMPORTANT: Respond with ONLY valid JSON in the following format (no markdown fences):

{{
    "exam_title": "Mock Exam - {subject_name}",
    "instructions": "General exam instructions here",
    "time_limit": {time_limit},
    "total_marks": {total_marks},
    "sections": [
        {{
            "section_name": "SECTION A: SHORT ANSWER QUESTIONS",
            "section_type": "short_answer",
            "section_info": "(5 x 2 = 10 Marks)",
            "instructions": "Answer all questions of this section together. Write clear, well-commented answers.",
            "questions": [
                {{
                    "id": 1,
                    "type": "short_answer",
                    "question": "Explain the difference between HTML4 and HTML5 and list three new semantic elements introduced in HTML5.",
                    "marks": 2
                }},
                {{
                    "id": 2,
                    "type": "short_answer", 
                    "question": "Describe the CSS box model and explain how margin, padding, and border properties work together to create element spacing.",
                    "marks": 2
                }}
            ]
        }},
        {{
            "section_name": "SECTION B: LONG ANSWER QUESTIONS",
            "section_type": "descriptive",
            "section_info": "(4 x 10 = 40 Marks)",
            "instructions": "Answer all questions of this section together. Provide detailed explanations with examples.",
            "questions": [
                {{
                    "id": 3,
                    "type": "descriptive",
                    "question": "Design and write complete HTML5 code for a responsive web page that includes: a) Proper semantic HTML5 structure with header, nav, main, and footer elements, b) A navigation menu with 5 links using CSS flexbox, c) Embedded CSS for responsive design with media queries, d) A contact form with HTML5 validation for name, email, and message fields. Explain how each part contributes to modern web development.",
                    "marks": 10
                }}
            ]
        }}
    ]
}}

SYLLABUS TOPICS TO USE: {syllabus_content}

FINAL WARNING: 
- DO NOT generate questions like "Descriptive question 1 about [subject]"
- DO NOT use generic placeholders
- DO NOT create sample questions
- YOU MUST create REAL questions about the specific syllabus topics: {syllabus_content}

Generate questions that are:
- Based on the specific syllabus topics mentioned above
- Professional and exam-appropriate
- Clear and unambiguous
- Testing actual knowledge and understanding
- Following the specified format and structure
- REAL questions about the syllabus content, not generic placeholders

FINAL WARNING: 
- DO NOT generate questions like "Descriptive question 1 about [subject]"
- DO NOT use generic placeholders
- DO NOT create sample questions
- YOU MUST create REAL questions about the specific syllabus topics: {syllabus_content}

Remember: If the syllabus mentions "HTML, CSS, JavaScript", create questions about HTML, CSS, and JavaScript specifically. Do NOT create generic questions.

Respond with ONLY the JSON, no other text."""
        
        print(f"Generating exam with custom prompt for subject: {subject_name}")
        print(f"Syllabus: {syllabus_content}")
        print(f"Format: {question_paper_format}")
        
        # Generate exam using REST helper
        response_text = gemini_generate_text(custom_prompt).strip()
        
        # Remove markdown code blocks if present
        if response_text.startswith('```json'):
            response_text = response_text[7:]
        if response_text.startswith('```'):
            response_text = response_text[3:]
        if response_text.endswith('```'):
            response_text = response_text[:-3]
        
        response_text = response_text.strip()
        
        print(f"Raw Gemini Response: {response_text[:500]}...")
        
        # Parse JSON
        exam_data = json.loads(response_text)
        
        print(f"Successfully parsed exam data with {len(exam_data.get('sections', []))} sections")
        return exam_data
        
    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
        print(f"Response text: {response_text[:500] if 'response_text' in locals() else 'No response'}")
        
        # Try to fix common JSON issues
        try:
            start_idx = response_text.find('{')
            end_idx = response_text.rfind('}')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                cleaned_text = response_text[start_idx:end_idx+1]
                exam_data = json.loads(cleaned_text)
                print("Successfully parsed after cleaning JSON")
                return exam_data
        except:
            pass
        
        return None
    except Exception as e:
        print(f"Error generating exam: {e}")
        print(f"Response text: {response_text[:500] if 'response_text' in locals() else 'No response'}")
        return None

@app.route('/create_custom_exam', methods=['GET', 'POST'])
def create_custom_exam():
    """Create exam using the custom Gemini prompt"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('student'))
    
    if request.method == 'POST':
        try:
            subject_name = request.form.get('subject_name')
            syllabus_content = request.form.get('syllabus_content')
            question_format = request.form.get('question_format', 'mixed')
            number_of_questions = int(request.form.get('number_of_questions', 10))
            total_marks = int(request.form.get('total_marks', 50))
            question_paper_format = request.form.get('question_paper_format', '')
            time_limit = int(request.form.get('time_limit', 60))
            difficulty_level = request.form.get('difficulty_level', 'medium')
            
            if not all([subject_name, syllabus_content]):
                flash('Subject name and syllabus content are required.', 'error')
                return redirect(url_for('create_custom_exam'))
            
            # Generate exam using custom prompt
            exam_data = generate_exam_with_custom_prompt(
                subject_name, syllabus_content, question_format, number_of_questions, 
                total_marks, question_paper_format, time_limit, difficulty_level
            )
            
            if not exam_data:
                flash('Failed to generate exam. Please check your inputs and try again.', 'error')
                return redirect(url_for('create_custom_exam'))
            
            # Generate unique exam ID
            exam_id = generate_exam_id()
            
            # Save exam to database
            exam = MockExam(
                exam_id=exam_id,
                student_username=session['username'],
                subject_name=subject_name,
                syllabus=syllabus_content,
                question_format=question_format,
                total_questions=number_of_questions,
                total_marks=total_marks,
                time_limit=time_limit,
                difficulty_level=difficulty_level,
                format_description=question_paper_format,
                questions=json.dumps(exam_data),
                status='active'
            )
            
            db.session.add(exam)
            db.session.commit()
            
            flash('Exam generated successfully!', 'success')
            return redirect(url_for('take_mock_exam', exam_id=exam_id))
            
        except Exception as e:
            print(f"Error creating custom exam: {e}")
            flash('Error creating exam. Please try again.', 'error')
            return redirect(url_for('create_custom_exam'))
    
    return render_template('create_custom_exam.html', user=user)

@app.route('/create_mock_exam', methods=['GET', 'POST'])
def create_mock_exam():
    """Create a new mock exam"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('student'))
    
    if request.method == 'POST':
        try:
            subject_name = request.form.get('subject_name')
            syllabus = request.form.get('syllabus')
            question_format = request.form.get('question_format')
            total_questions = int(request.form.get('total_questions', 10))
            total_marks = int(request.form.get('total_marks', 50))
            time_limit = int(request.form.get('time_limit', 60))
            difficulty_level = request.form.get('difficulty_level', 'medium')
            format_description = request.form.get('format_description', '')
            
            if not all([subject_name, syllabus, question_format]):
                flash('Please fill in all required fields.', 'error')
                return redirect(url_for('create_mock_exam'))
            
            # Validate marks distribution
            if total_marks < total_questions:
                flash('Total marks should be at least equal to the number of questions.', 'error')
                return redirect(url_for('create_mock_exam'))
            
            # Generate unique exam ID
            exam_id = generate_exam_id()
            
            # Generate questions using Gemini API with enhanced parameters
            questions_data = generate_questions_with_gemini(
                subject_name, syllabus, question_format, total_questions, total_marks, difficulty_level, format_description
            )
            
            if not questions_data:
                # Try with a simpler prompt as fallback
                print("Primary generation failed, trying fallback...")
                questions_data = generate_fallback_questions(subject_name, total_questions, total_marks, question_format)
                
                if not questions_data:
                    flash('Failed to generate questions. This could be due to API issues or invalid format. Please check your format description and try again.', 'error')
                    return redirect(url_for('create_mock_exam'))
            
            # Create exam record
            exam = MockExam(
                exam_id=exam_id,
                student_username=session['username'],
                subject_name=subject_name,
                syllabus=syllabus,
                question_format=question_format,
                total_questions=total_questions,
                total_marks=total_marks,
                time_limit=time_limit,
                difficulty_level=difficulty_level,
                format_description=format_description,
                questions=json.dumps(questions_data),
                status='created'
            )
            
            db.session.add(exam)
            db.session.commit()
            
            flash('Mock exam created successfully!', 'success')
            return redirect(url_for('take_mock_exam', exam_id=exam_id))
            
        except Exception as e:
            print(f"Error creating mock exam: {e}")
            flash('Error creating mock exam. Please try again.', 'error')
            return redirect(url_for('create_mock_exam'))
    
    return render_template('create_mock_exam.html', user=user)

@app.route('/take_mock_exam/<exam_id>')
def take_mock_exam(exam_id):
    """Take a mock exam"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('student'))
    
    # Get exam details
    exam = db.session.execute(
        db.select(MockExam).filter_by(exam_id=exam_id, student_username=session['username'])
    ).scalar()
    
    if not exam:
        flash('Exam not found.', 'error')
        return redirect(url_for('mock_exam'))
    
    if exam.status == 'completed':
        flash('This exam has already been completed.', 'info')
        return redirect(url_for('exam_result', exam_id=exam_id))
    
    # Parse questions
    questions = json.loads(exam.questions)
    
    # Handle different question formats
    if 'sections' in questions:
        return render_template('take_mock_exam.html', exam=exam, questions=questions, user=user)
    elif 'questions' in questions:
        return render_template('take_mock_exam.html', exam=exam, questions=questions['questions'], user=user)
    else:
        return render_template('take_mock_exam.html', exam=exam, questions=questions, user=user)

@app.route('/submit_mock_exam/<exam_id>', methods=['POST'])
def submit_mock_exam(exam_id):
    """Submit mock exam answers"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('student'))
    
    try:
        # Get exam details
        exam = db.session.execute(
            db.select(MockExam).filter_by(exam_id=exam_id, student_username=session['username'])
        ).scalar()
        
        if not exam:
            flash('Exam not found.', 'error')
            return redirect(url_for('mock_exam'))
        
        if exam.status == 'completed':
            flash('Exam already completed.', 'error')
            return redirect(url_for('mock_exam'))
        
        # Get answers from form
        answers = {}
        for key, value in request.form.items():
            if key.startswith('answer_'):
                question_id = key.replace('answer_', '')
                answers[question_id] = value
        
        # Calculate time taken (you might want to track this in the frontend)
        time_taken = int(request.form.get('time_taken', 0))
        
        # Create exam attempt
        attempt = ExamAttempt(
            exam_id=exam_id,
            student_username=session['username'],
            answers=json.dumps(answers),
            time_taken=time_taken,
            status='submitted'
        )
        
        db.session.add(attempt)
        
        # Update exam status
        exam.status = 'completed'
        
        db.session.commit()
        
        # Start evaluation in background (you might want to use a task queue for this)
        # For now, we'll evaluate immediately
        questions_data = json.loads(exam.questions)
        print(f"Questions data structure: {list(questions_data.keys()) if isinstance(questions_data, dict) else 'Not a dict'}")
        
        # Handle different question formats
        if 'sections' in questions_data:
            # New format with sections
            questions_for_evaluation = questions_data
            print("Using sections format for evaluation")
        elif 'questions' in questions_data:
            # Old format with direct questions
            questions_for_evaluation = questions_data['questions']
            print("Using questions format for evaluation")
        else:
            # Fallback
            questions_for_evaluation = questions_data
            print("Using fallback format for evaluation")
        
        evaluation_result = evaluate_answers_with_gemini(
            questions_for_evaluation, answers, exam.subject_name
        )
        
        if evaluation_result:
            # Create exam result
            result = ExamResult(
                exam_id=exam_id,
                attempt_id=attempt.id,
                student_username=session['username'],
                total_marks=evaluation_result['total_marks'],
                obtained_marks=evaluation_result['obtained_marks'],
                percentage=evaluation_result['percentage'],
                grade=evaluation_result['grade'],
                detailed_feedback=json.dumps(evaluation_result['detailed_feedback'])
            )
            
            db.session.add(result)
            db.session.commit()
            
            # Redirect to results page instead of returning JSON
            flash('Exam submitted and evaluated successfully!', 'success')
            return redirect(url_for('exam_result', exam_id=exam_id))
        else:
            flash('Failed to evaluate exam. Please try again.', 'error')
            return redirect(url_for('mock_exam'))
            
    except Exception as e:
        print(f"Error submitting mock exam: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error submitting exam: {str(e)}', 'error')
        return redirect(url_for('mock_exam'))

@app.route('/exam_result/<exam_id>')
def exam_result(exam_id):
    """View exam result"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('student'))
    
    # Get exam details
    exam = db.session.execute(
        db.select(MockExam).filter_by(exam_id=exam_id, student_username=session['username'])
    ).scalar()
    
    if not exam:
        flash('Exam not found.', 'error')
        return redirect(url_for('mock_exam'))
    
    # Get exam result
    result = db.session.execute(
        db.select(ExamResult).filter_by(exam_id=exam_id, student_username=session['username'])
    ).scalar()
    
    if not result:
        flash('Exam result not found. Please try again later.', 'error')
        return redirect(url_for('mock_exam'))
    
    # Get exam attempt to retrieve student answers
    attempt = db.session.execute(
        db.select(ExamAttempt).filter_by(exam_id=exam_id, student_username=session['username'])
    ).scalar()
    
    student_answers = {}
    time_taken = 0
    if attempt:
        student_answers = json.loads(attempt.answers)
        time_taken = attempt.time_taken
    
    # Parse detailed feedback
    detailed_feedback = json.loads(result.detailed_feedback)
    
    # Parse questions for display
    questions = json.loads(exam.questions)
    
    # Handle different question formats for display
    if 'sections' in questions:
        # New format with sections - flatten questions for display
        all_questions = []
        for section in questions['sections']:
            for question in section['questions']:
                all_questions.append(question)
        questions_for_display = all_questions
    elif 'questions' in questions:
        # Old format with direct questions
        questions_for_display = questions['questions']
    else:
        # Fallback
        questions_for_display = questions if isinstance(questions, list) else [questions]
    
    # Merge detailed feedback with questions for display
    questions_with_feedback = []
    for question in questions_for_display:
        question_id = str(question.get('id', ''))
        
        # Find corresponding feedback
        feedback_data = None
        for feedback in detailed_feedback:
            if str(feedback.get('question_id', '')) == question_id:
                feedback_data = feedback
                break
        
        # Create enhanced question object with feedback
        enhanced_question = question.copy()
        
        # Add student answer
        enhanced_question['student_answer'] = student_answers.get(question_id, 'No answer provided')
        
        if feedback_data:
            enhanced_question.update({
                'marks_awarded': feedback_data.get('marks_awarded', 0),
                'feedback': feedback_data.get('feedback', ''),
                'suggestions': feedback_data.get('suggestions', ''),
                'correct_answer': feedback_data.get('correct_answer', ''),
                'is_correct': feedback_data.get('is_correct', False),
                'accuracy_percentage': feedback_data.get('accuracy_percentage', 0.0)
            })
        else:
            # Fallback if no feedback found
            enhanced_question.update({
                'marks_awarded': 0,
                'feedback': 'No feedback available',
                'suggestions': 'Please contact support',
                'correct_answer': 'Not available',
                'is_correct': False,
                'accuracy_percentage': 0.0
            })
        
        questions_with_feedback.append(enhanced_question)
    
    # Add time information to result object
    result.time_taken = time_taken
    result.time_efficiency = (time_taken / exam.time_limit * 100) if exam.time_limit > 0 else 0
    result.avg_time_per_question = (time_taken / len(questions_with_feedback)) if len(questions_with_feedback) > 0 else 0
    
    # Compute performance summary metrics from feedback and answers
    try:
        total_questions = len(questions_with_feedback)
        attempted_ids = set([qid for qid, ans in student_answers.items() if str(ans).strip()])
        questions_attempted = len(attempted_ids)
        correct_answers = 0
        partially_correct = 0
        incorrect_answers = 0
        for fb in detailed_feedback:
            fb_total = fb.get('total_marks', 0) or 0
            fb_awarded = fb.get('marks_awarded', 0) or 0
            # Classification: full marks -> correct; between 0 and full -> partial; 0 -> incorrect
            if fb_total > 0 and fb_awarded >= fb_total:
                correct_answers += 1
            elif fb_total > 0 and fb_awarded > 0:
                partially_correct += 1
            else:
                incorrect_answers += 1
        # Accuracy: correct counts as 1, partially correct counts as 0.5
        accuracy = (((correct_answers) + (0.5 * partially_correct)) / total_questions * 100) if total_questions > 0 else 0
        
        # Attach to result object so the template can read them
        result.total_questions = total_questions
        result.questions_attempted = questions_attempted
        result.correct_answers = correct_answers
        result.partially_correct_answers = partially_correct
        result.incorrect_answers = incorrect_answers
        result.accuracy = round(accuracy, 2)
    except Exception as _e:
        # Fallback defaults if anything goes wrong
        result.total_questions = len(questions_with_feedback)
        result.questions_attempted = len([qid for qid, ans in student_answers.items() if str(ans).strip()])
        result.correct_answers = 0
        result.partially_correct_answers = 0
        result.incorrect_answers = result.total_questions
        result.accuracy = 0.0
    
    return render_template('exam_result.html', 
                         exam=exam, 
                         result=result, 
                         detailed_feedback=detailed_feedback,
                         questions=questions_with_feedback,
                         user=user)

@app.route('/exam_history')
def exam_history():
    """View exam history"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('student'))
    
    # Get all exams with results
    exams_with_results = db.session.execute(
        db.select(MockExam, ExamResult)
        .join(ExamResult, MockExam.exam_id == ExamResult.exam_id)
        .filter(MockExam.student_username == session['username'])
        .order_by(MockExam.created_at.desc())
    ).all()
    
    return render_template('exam_history.html', exams_with_results=exams_with_results, user=user)

@app.route('/test_single_evaluation')
def test_single_evaluation():
    """Test individual question scoring - AI gives direct marks"""
    try:
        # Sample question and answer
        sample_question = {
            "id": 1,
            "type": "short_answer",
            "question": "What is the difference between HTML and CSS?",
            "marks": 5
        }
        
        sample_answer = "HTML is used for structure and content of web pages, while CSS is used for styling and presentation. HTML defines what content appears on the page, and CSS defines how that content looks."
        
        result = evaluate_single_question_with_gemini(sample_question, sample_answer, "Web Technologies")
        
        if result:
            return f"""
            <h1>Single Question Scoring Test</h1>
            <h2>Question:</h2>
            <p><strong>{sample_question['question']}</strong> ({sample_question['marks']} marks)</p>
            
            <h2>Student Answer:</h2>
            <p>{sample_answer}</p>
            
            <h2>AI Scoring Result:</h2>
            <pre>{json.dumps(result, indent=2)}</pre>
            
            <h2>Summary:</h2>
            <p><strong>Marks Awarded by AI:</strong> {result.get('marks_awarded', 0)}/{result.get('total_marks', 0)}</p>
            <p><strong>AI Feedback:</strong> {result.get('feedback', 'No feedback')}</p>
            <p><strong>Reference Answer:</strong> {result.get('correct_answer', 'Not provided')}</p>
            <p><strong>Scoring Approach:</strong> AI directly scores the answer quality (0 to {sample_question['marks']} marks)</p>
            """
        else:
            return "<h1>Evaluation failed</h1>"
            
    except Exception as e:
        return f"<h1>Error: {e}</h1>"

# Learning Pods AI Clustering Functions
def clean_gemini_response(response_text):
    """Clean Gemini response by removing markdown code blocks"""
    if not response_text:
        return response_text
    
    text = response_text.strip()
    # Remove markdown code blocks
    if text.startswith('```json'):
        text = text[7:]
    elif text.startswith('```'):
        text = text[3:]
    
    if text.endswith('```'):
        text = text[:-3]
    
    return text.strip()

def get_student_learning_profile(username):
    """Get or create student learning profile with AI analysis"""
    profile = db.session.execute(db.select(StudentProfile).filter_by(username=username)).scalar()
    
    if not profile:
        # Get student's exam history and performance data
        exam_results = db.session.execute(
            db.select(ExamResult).filter_by(student_username=username)
        ).scalars().all()
        
        # Calculate average performance
        avg_score = 0.0
        if exam_results:
            avg_score = sum([r.percentage for r in exam_results]) / len(exam_results)
        
        # Use Gemini to analyze learning style and preferences
        try:
            # Use the REST-based helper for compatibility
            
            # Create a simple assessment prompt
            assessment_prompt = f"""
            Analyze this student's learning profile based on their academic performance:
            - Average exam score: {avg_score}%
            - Number of exams taken: {len(exam_results)}
            
            Based on this limited data, suggest:
            1. Learning style (visual, auditory, kinesthetic, reading)
            2. Preferred pace (fast, medium, slow)
            3. Collaboration comfort (high, medium, low)
            4. Leadership tendency (high, medium, low)
            5. Communication style (assertive, passive, balanced)
            
            Return a JSON response with these fields.
            """
            
            response_text = gemini_generate_text(assessment_prompt)
            cleaned_response = clean_gemini_response(response_text)
            result = json.loads(cleaned_response)
            
            # Create new profile
            profile = StudentProfile(
                username=username,
                learning_style=result.get('learning_style', 'balanced'),
                preferred_pace=result.get('preferred_pace', 'medium'),
                collaboration_comfort=result.get('collaboration_comfort', 'medium'),
                leadership_tendency=result.get('leadership_tendency', 'medium'),
                communication_style=result.get('communication_style', 'balanced'),
                average_exam_score=avg_score
            )
            
            db.session.add(profile)
            db.session.commit()
            
        except Exception as e:
            print(f"Error creating AI profile for {username}: {e}")
            # Create default profile
            profile = StudentProfile(
                username=username,
                learning_style='balanced',
                preferred_pace='medium',
                collaboration_comfort='medium',
                leadership_tendency='medium',
                communication_style='balanced',
                average_exam_score=avg_score
            )
            db.session.add(profile)
            db.session.commit()
    
    return profile

def create_ai_learning_pods(subject, max_pods=None, cohort: Optional[Dict] = None, teacher_username: Optional[str] = None):
    """Use AI to create optimal learning pods for a subject"""
    try:
        # Get all students and group by cohort (year/department/section)
        students = db.session.execute(
            db.select(Credentials).filter_by(role='student')
        ).scalars().all()

        # Optional cohort filter
        if cohort:
            year_f = cohort.get('study_year')
            dept_f = (cohort.get('department') or '').upper() or None
            sect_f = (cohort.get('section') or '').upper() or None
            def matches(u: Credentials):
                if year_f is not None and (u.study_year or None) != year_f:
                    return False
                if dept_f is not None and (u.department or '').upper() != dept_f:
                    return False
                if sect_f is not None and (u.section or '').upper() != sect_f:
                    return False
                return True
            students = [u for u in students if matches(u)]

        def cohort_key(u: Credentials):
            return (u.study_year or 0, (u.department or '').upper(), (u.section or '').upper())
        
        if len(students) < 3:
            return {"error": "Need at least 3 students to create pods"}
        
        # Get learning profiles for all students
        profiles = []
        for student in students:
            profile = get_student_learning_profile(student.username)
            profiles.append({
                'username': student.username,
                'learning_style': profile.learning_style,
                'preferred_pace': profile.preferred_pace,
                'collaboration_comfort': profile.collaboration_comfort,
                'leadership_tendency': profile.leadership_tendency,
                'communication_style': profile.communication_style,
                'average_exam_score': profile.average_exam_score
            })
        
        # Random groupings per cohort without cross mixing
        import random
        created_pods = []
        cohort_to_usernames: dict[tuple, list[str]] = {}
        for s in students:
            key = cohort_key(s)
            cohort_to_usernames.setdefault(key, []).append(s.username)

        for (year, dept, sect), usernames in cohort_to_usernames.items():
            if len(usernames) < 3:
                continue
            random.shuffle(usernames)
            pod_index = 1
            i = 0
            while i < len(usernames):
                remaining = len(usernames) - i
                if remaining >= 4:
                    size = 4
                elif remaining == 3:
                    size = 3
                elif remaining == 2 and created_pods:
                    # merge last two into previous pod if capacity allows
                    last = created_pods[-1]
                    if len(last['members']) < 4:
                        pod_id = last['id']
                        for username in usernames[i:i+2]:
                            db.session.add(PodMembership(pod_id=pod_id, student_username=username, role='member'))
                            last['members'].append(username)
                        i += 2
                        break
                    else:
                        size = 3
                else:
                    size = remaining

                members = usernames[i:i+size]
                i += size
                pod = LearningPod(
                    pod_name=f"{(dept or 'POD')}-{(sect or 'X')}-{(year or 0)}-#{pod_index}",
                    subject=subject,
                    max_members=len(members),
                    study_year=year if year != 0 else None,
                    department=dept or None,
                    section=sect or None,
                    teacher_username=teacher_username
                )
                db.session.add(pod)
                db.session.flush()
                for j, username in enumerate(members):
                    role = 'leader' if j == 0 else 'member'
                    db.session.add(PodMembership(pod_id=pod.id, student_username=username, role=role))
                created_pods.append({'id': pod.id, 'name': pod.pod_name, 'members': members})
                pod_index += 1
        
        db.session.commit()
        return {"success": True, "pods": created_pods}
        
    except Exception as e:
        print(f"Error creating AI pods: {e}")
        return {"error": f"Failed to create pods: {str(e)}"}

def create_fallback_pod_tasks(pod_id, subject, member_usernames):
    """Create fallback tasks when AI generation fails"""
    try:
        # Create 2 default tasks based on the subject
        fallback_tasks_data = [
            {
                "type": "micro_lesson",
                "title": f"Teach a Key Concept in {subject}",
                "description": f"Each member should prepare and teach one important concept from {subject} to the rest of the pod. Focus on explaining it clearly with examples.",
                "learning_objectives": [
                    f"Understand key concepts in {subject}",
                    "Develop teaching and explanation skills",
                    "Practice peer-to-peer learning"
                ],
                "estimated_duration": 25,
                "content": {
                    "instructions": f"1. Choose one important topic from {subject}\n2. Prepare a 5-minute explanation with examples\n3. Teach it to your pod members\n4. Answer questions and clarify doubts",
                    "materials": ["Course notes", "Textbook", "Online resources"]
                }
            },
            {
                "type": "peer_challenge",
                "title": f"Collaborative Problem Solving in {subject}",
                "description": f"Work together as a team to solve a challenging problem related to {subject}. Discuss different approaches and find the best solution.",
                "learning_objectives": [
                    f"Apply {subject} knowledge to solve problems",
                    "Develop collaborative problem-solving skills",
                    "Learn from different perspectives"
                ],
                "estimated_duration": 30,
                "content": {
                    "instructions": f"1. Identify a challenging problem in {subject}\n2. Brainstorm different solution approaches\n3. Work together to implement the best solution\n4. Document your process and findings",
                    "materials": ["Problem statement", "Solution framework", "Documentation template"]
                }
            }
        ]
        
        # Create tasks in database
        created_tasks = []
        for task_data in fallback_tasks_data:
            task = PodTask(
                pod_id=pod_id,
                task_type=task_data['type'],
                title=task_data['title'],
                description=task_data['description'],
                subject=subject,
                learning_objectives=json.dumps(task_data.get('learning_objectives', [])),
                content=json.dumps(task_data.get('content', {})),
                estimated_duration=task_data.get('estimated_duration', 20)
            )
            db.session.add(task)
            created_tasks.append(task)
        
        db.session.commit()
        return [{"id": t.id, "title": t.title} for t in created_tasks]
        
    except Exception as e:
        print(f"Error creating fallback tasks: {e}")
        return None

def generate_pod_tasks(pod_id, subject):
    """Generate personalized micro-curriculum and peer challenges for a pod"""
    try:
        pod = db.session.execute(db.select(LearningPod).filter_by(id=pod_id)).scalar()
        if not pod:
            return {"error": "Pod not found"}
        
        # Get pod members
        memberships = db.session.execute(
            db.select(PodMembership).filter_by(pod_id=pod_id, is_active=True)
        ).scalars().all()
        
        member_usernames = [m.student_username for m in memberships]
        
        # Check if we already have tasks for this pod to avoid quota issues
        existing_tasks = db.session.execute(
            db.select(PodTask).filter_by(pod_id=pod_id)
        ).scalars().all()
        
        if len(existing_tasks) >= 3:
            return {"error": "This pod already has enough tasks. Please wait before generating more to avoid API quota limits."}
        
        try:
            # Use Gemini/PaLM to create personalized tasks (compatible across SDK versions)
            task_prompt = f"""
            Create 2 learning tasks for a pod of students studying {subject}:
            Pod members: {', '.join(member_usernames)}
            
            Create tasks that:
            1. Encourage peer teaching and collaboration
            2. Are suitable for different learning styles
            3. Include both individual and group components
            4. Take 15-30 minutes each
            5. Have clear learning objectives
            
            Task types:
            - micro_lesson: Student teaches a concept to others
            - peer_challenge: Collaborative problem-solving
            - reflection: Group discussion and reflection
            
            Return JSON format:
            {{
                "tasks": [
                    {{
                        "type": "micro_lesson",
                        "title": "Task Title",
                        "description": "Detailed description",
                        "learning_objectives": ["objective1", "objective2"],
                        "estimated_duration": 20,
                        "content": {{"instructions": "...", "materials": [...]}}
                    }}
                ]
            }}
            """

            # Generate text with SDK compatibility handling
            response_text = None
            try:
                # Use REST API directly since the old SDK doesn't have GenerativeModel
                import requests
                api_key = GEMINI_API_KEY
                if not api_key:
                    raise RuntimeError("Missing GEMINI_API_KEY for REST call")
                
                # Try different model endpoints (prioritize working ones)
                model_endpoints = [
                    "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent",
                    "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent",
                    "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-pro:generateContent"
                ]
                
                for endpoint in model_endpoints:
                    try:
                        url = f"{endpoint}?key={api_key}"
                        payload = {
                            "contents": [
                                {
                                    "parts": [
                                        {"text": task_prompt}
                                    ]
                                }
                            ]
                        }
                        headers = {"Content-Type": "application/json"}
                        rest_resp = requests.post(url, headers=headers, data=json.dumps(payload), timeout=45)
                        if rest_resp.status_code == 200:
                            data = rest_resp.json()
                            try:
                                parts = (((data or {}).get('candidates') or [{}])[0].get('content') or {}).get('parts') or []
                                response_text = ''.join([p.get('text', '') for p in parts])
                                print(f"Successfully used endpoint: {endpoint}")
                                break
                            except Exception as pe:
                                print(f"Failed to parse response from {endpoint}: {pe}")
                                continue
                        else:
                            print(f"HTTP {rest_resp.status_code} from {endpoint}: {rest_resp.text[:100]}")
                            continue
                    except Exception as e:
                        print(f"Failed to use endpoint {endpoint}: {e}")
                        continue
                
                if not response_text:
                    raise RuntimeError("All Gemini endpoints failed")
            except Exception as gen_err:
                raise RuntimeError(f"Gemini generation failed: {gen_err}")

            cleaned_response = clean_gemini_response(response_text or '')
            result = json.loads(cleaned_response)
            
            # Create tasks in database
            created_tasks = []
            for task_data in result.get('tasks', []):
                task = PodTask(
                    pod_id=pod_id,
                    task_type=task_data['type'],
                    title=task_data['title'],
                    description=task_data['description'],
                    subject=subject,
                    learning_objectives=json.dumps(task_data.get('learning_objectives', [])),
                    content=json.dumps(task_data.get('content', {})),
                    estimated_duration=task_data.get('estimated_duration', 20)
                )
                db.session.add(task)
                created_tasks.append(task)
            
            db.session.commit()
            return {"success": True, "tasks": [{"id": t.id, "title": t.title} for t in created_tasks]}
            
        except Exception as api_error:
            error_msg = str(api_error)
            print(f"Gemini API error: {error_msg}")
            
            # Create fallback tasks when AI generation fails
            try:
                fallback_tasks = create_fallback_pod_tasks(pod_id, subject, member_usernames)
                if fallback_tasks:
                    return {"success": True, "tasks": fallback_tasks, "fallback": True}
            except Exception as fallback_error:
                print(f"Fallback task creation failed: {fallback_error}")
            
            if "quota" in error_msg.lower() or "429" in error_msg:
                return {"error": "API quota exceeded. Please wait a few minutes before generating more tasks. You can create manual tasks instead."}
            elif "rate" in error_msg.lower():
                return {"error": "API rate limit reached. Please wait 30 seconds before trying again."}
            else:
                return {"error": f"API error: {error_msg}. You can create manual tasks instead."}
        
    except Exception as e:
        print(f"Error generating pod tasks: {e}")
        return {"error": f"Failed to generate tasks: {str(e)}"}

# Learning Pods Routes
@app.route('/learning_pods')
def learning_pods_dashboard():
    """Main learning pods dashboard"""
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    
    if user.role == 'student':
        # Get student's pod memberships
        memberships = db.session.execute(
            db.select(PodMembership).filter_by(student_username=session['username'], is_active=True)
        ).scalars().all()
        
        pods = []
        for membership in memberships:
            pod = db.session.execute(db.select(LearningPod).filter_by(id=membership.pod_id)).scalar()
            if pod:
                # Get other members
                other_members = db.session.execute(
                    db.select(PodMembership).filter_by(pod_id=pod.id, is_active=True)
                ).scalars().all()
                
                pods.append({
                    'pod': pod,
                    'membership': membership,
                    'members': [m.student_username for m in other_members if m.student_username != session['username']]
                })
        
        return render_template('student_pods_dashboard.html', user=user, pods=pods)
    
    elif user.role in ['teacher', 'admin']:
        # Get all pods
        all_pods = db.session.execute(db.select(LearningPod).filter_by(is_active=True)).scalars().all()
        
        pods_with_details = []
        for pod in all_pods:
            memberships = db.session.execute(
                db.select(PodMembership).filter_by(pod_id=pod.id, is_active=True)
            ).scalars().all()
            
            pods_with_details.append({
                'pod': pod,
                'member_count': len(memberships),
                'members': [m.student_username for m in memberships],
                'cohort': {
                    'year': pod.study_year,
                    'department': pod.department,
                    'section': pod.section
                }
            })
        
        return render_template('teacher_pods_dashboard.html', user=user, pods=pods_with_details)
    
    return redirect(url_for('student'))

@app.route('/create_pods', methods=['POST'])
def create_pods():
    """Create AI-powered learning pods"""
    if 'username' not in session:
        return jsonify({"error": "Please login first"}), 401
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Only teachers and admins can create pods"}), 403
    
    data = request.get_json()
    subject = data.get('subject', 'General')
    # Optional cohort selection by teacher
    cohort = {
        'study_year': data.get('study_year') if data.get('study_year') not in (None, '') else None,
        'department': data.get('department') or None,
        'section': data.get('section') or None,
    }
    # Normalize year to int if provided
    if cohort['study_year'] is not None:
        try:
            cohort['study_year'] = int(cohort['study_year'])
        except Exception:
            cohort['study_year'] = None
    result = create_ai_learning_pods(subject, cohort=cohort, teacher_username=session['username'])
    return jsonify(result)

@app.route('/pod/<int:pod_id>')
def pod_detail(pod_id):
    """View pod details and tasks"""
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    pod = db.session.execute(db.select(LearningPod).filter_by(id=pod_id)).scalar()
    if not pod:
        flash("Pod not found")
        return redirect(url_for('learning_pods_dashboard'))
    
    # Get pod members
    memberships = db.session.execute(
        db.select(PodMembership).filter_by(pod_id=pod_id, is_active=True)
    ).scalars().all()
    
    # Get pod tasks
    tasks = db.session.execute(
        db.select(PodTask).filter_by(pod_id=pod_id).order_by(PodTask.assigned_date.desc())
    ).scalars().all()
    
    # Get collaboration events
    events = db.session.execute(
        db.select(CollaborationEvent).filter_by(pod_id=pod_id).order_by(CollaborationEvent.timestamp.desc()).limit(20)
    ).scalars().all()
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    
    return render_template('pod_detail.html', 
                         pod=pod, 
                         memberships=memberships,
                         tasks=tasks,
                         events=events,
                         user=user)

@app.route('/generate_pod_tasks/<int:pod_id>', methods=['POST'])
def generate_pod_tasks_route(pod_id):
    """Generate tasks for a specific pod"""
    if 'username' not in session:
        return jsonify({"error": "Please login first"}), 401
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Only teachers and admins can generate tasks"}), 403
    
    pod = db.session.execute(db.select(LearningPod).filter_by(id=pod_id)).scalar()
    if not pod:
        return jsonify({"error": "Pod not found"}), 404
    
    result = generate_pod_tasks(pod_id, pod.subject)
    return jsonify(result)

@app.route('/create_manual_task', methods=['POST'])
def create_manual_task():
    """Create a manual task when AI generation fails"""
    if 'username' not in session:
        return jsonify({"error": "Please login first"}), 401
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Only teachers and admins can create tasks"}), 403
    
    data = request.get_json()
    pod_id = data.get('pod_id')
    task_type = data.get('task_type', 'micro_lesson')
    title = data.get('title', 'Manual Task')
    description = data.get('description', 'A manually created learning task')
    
    if not pod_id:
        return jsonify({"error": "Pod ID is required"}), 400
    
    try:
        pod = db.session.execute(db.select(LearningPod).filter_by(id=pod_id)).scalar()
        if not pod:
            return jsonify({"error": "Pod not found"}), 404
        
        # Create manual task
        task = PodTask(
            pod_id=pod_id,
            task_type=task_type,
            title=title,
            description=description,
            subject=pod.subject,
            learning_objectives=json.dumps(["Complete the assigned task", "Collaborate with pod members"]),
            content=json.dumps({
                "instructions": description,
                "materials": ["Collaboration with pod members", "Basic materials as needed"]
            }),
            estimated_duration=20
        )
        db.session.add(task)
        db.session.commit()
        
        return jsonify({"success": True, "task": {"id": task.id, "title": task.title}})
        
    except Exception as e:
        print(f"Error creating manual task: {e}")
        return jsonify({"error": f"Failed to create task: {str(e)}"}), 500

@app.route('/take_pod_task/<int:task_id>')
def take_pod_task(task_id):
    """Take a pod task"""
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    task = db.session.execute(db.select(PodTask).filter_by(id=task_id)).scalar()
    if not task:
        flash("Task not found")
        return redirect(url_for('learning_pods_dashboard'))
    
    # Check if user is member of this pod
    membership = db.session.execute(
        db.select(PodMembership).filter_by(pod_id=task.pod_id, student_username=session['username'], is_active=True)
    ).scalar()
    
    if not membership:
        flash("You are not a member of this pod")
        return redirect(url_for('learning_pods_dashboard'))
    
    # Ensure JSON fields are parsed to Python objects for the template
    try:
        if task.learning_objectives:
            if isinstance(task.learning_objectives, str):
                task.learning_objectives = json.loads(task.learning_objectives)
        else:
            task.learning_objectives = []
    except Exception:
        # Fallback to empty list if malformed
        task.learning_objectives = []

    try:
        if task.content:
            if isinstance(task.content, str):
                task.content = json.loads(task.content)
        else:
            task.content = {}
    except Exception:
        # Fallback to empty dict if malformed
        task.content = {}

    # Get existing submission
    submission = db.session.execute(
        db.select(TaskSubmission).filter_by(task_id=task_id, student_username=session['username'])
    ).scalar()
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    
    return render_template('take_pod_task.html', task=task, submission=submission, user=user)

@app.route('/submit_pod_task', methods=['POST'])
def submit_pod_task():
    """Submit pod task response"""
    if 'username' not in session:
        return jsonify({"error": "Please login first"}), 401
    
    data = request.get_json()
    task_id = data.get('task_id')
    submission_content = data.get('submission_content')
    reflection_notes = data.get('reflection_notes', '')
    time_spent = data.get('time_spent', 0)
    
    if not task_id or not submission_content:
        return jsonify({"error": "Missing required data"}), 400
    
    try:
        # Create or update submission
        submission = db.session.execute(
            db.select(TaskSubmission).filter_by(task_id=task_id, student_username=session['username'])
        ).scalar()
        
        if submission:
            submission.submission_content = json.dumps(submission_content)
            submission.reflection_notes = reflection_notes
            submission.time_spent = time_spent
            submission.status = 'submitted'
        else:
            submission = TaskSubmission(
                task_id=task_id,
                student_username=session['username'],
                submission_content=json.dumps(submission_content),
                reflection_notes=reflection_notes,
                time_spent=time_spent,
                status='submitted'
            )
            db.session.add(submission)
        
        db.session.commit()
        
        # Removed auto-award on submission; points will be given after teacher evaluation
        
        return jsonify({"success": True, "message": "Task submitted successfully!"})
        
    except Exception as e:
        print(f"Error submitting pod task: {e}")
        return jsonify({"error": f"Failed to submit task: {str(e)}"}), 500

@app.route('/pod_analytics')
def pod_analytics():
    """Teacher analytics dashboard for pods"""
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if user.role not in ['teacher', 'admin']:
        flash("Access denied")
        return redirect(url_for('learning_pods_dashboard'))
    
    # Get analytics data
    all_pods = db.session.execute(db.select(LearningPod).filter_by(is_active=True)).scalars().all()
    
    analytics_data = []
    for pod in all_pods:
        # Get pod members
        memberships = db.session.execute(
            db.select(PodMembership).filter_by(pod_id=pod.id, is_active=True)
        ).scalars().all()
        
        # Get task completion rates
        tasks = db.session.execute(db.select(PodTask).filter_by(pod_id=pod.id)).scalars().all()
        total_tasks = len(tasks)
        completed_tasks = 0
        
        for task in tasks:
            submissions = db.session.execute(
                db.select(TaskSubmission).filter_by(task_id=task.id)
            ).scalars().all()
            if len(submissions) >= len(memberships) * 0.8:  # 80% completion threshold
                completed_tasks += 1
        
        # Get collaboration events
        events = db.session.execute(
            db.select(CollaborationEvent).filter_by(pod_id=pod.id)
        ).scalars().all()
        
        # Get merit badges earned
        badges = db.session.execute(
            db.select(MeritBadge).filter(MeritBadge.pod_id == pod.id)
        ).scalars().all()
        
        analytics_data.append({
            'pod': pod,
            'member_count': len(memberships),
            'members': [m.student_username for m in memberships],
            'task_completion_rate': (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0,
            'collaboration_events': len(events),
            'badges_earned': len(badges),
            'leader': next((m.student_username for m in memberships if m.role == 'leader'), None)
        })
    
    return render_template('pod_analytics.html', user=user, analytics=analytics_data)

@app.route('/review_submission/<int:task_id>')
def review_submission(task_id):
    """Teacher view to review a task submissions list"""
    if 'username' not in session:
        flash("Please login first")
        return redirect(url_for('login'))
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if user.role not in ['teacher', 'admin']:
        flash("Access denied")
        return redirect(url_for('learning_pods_dashboard'))

    task = db.session.execute(db.select(PodTask).filter_by(id=task_id)).scalar()
    if not task:
        flash("Task not found")
        return redirect(url_for('learning_pods_dashboard'))

    submissions = db.session.execute(db.select(TaskSubmission).filter_by(task_id=task_id)).scalars().all()
    pod = db.session.execute(db.select(LearningPod).filter_by(id=task.pod_id)).scalar()
    return render_template('review_pod_submissions.html', task=task, submissions=submissions, pod=pod, user=user)

@app.route('/grade_submission', methods=['POST'])
def grade_submission():
    if 'username' not in session:
        return jsonify({"error": "Please login first"}), 401
    teacher = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if teacher.role not in ['teacher', 'admin']:
        return jsonify({"error": "Access denied"}), 403

    data = request.get_json() or {}
    submission_id = data.get('submission_id')
    points = data.get('points')
    feedback = data.get('feedback', '')
    try:
        points = int(points)
    except Exception:
        points = 0

    submission = db.session.execute(db.select(TaskSubmission).filter_by(id=submission_id)).scalar()
    if not submission:
        return jsonify({"error": "Submission not found"}), 404

    # Update submission grading fields
    submission.awarded_points = max(points, 0)
    submission.teacher_feedback = feedback
    submission.evaluated_at = datetime.utcnow()
    submission.evaluated_by = teacher.username
    submission.status = 'graded'

    # Add points to student account (not credits)
    student = db.session.execute(db.select(Credentials).filter_by(username=submission.student_username)).scalar()
    student.points = (student.points or 0) + submission.awarded_points

    db.session.commit()
    return jsonify({"success": True, "message": "Points awarded"})

# ---------------------- Exam Contest Feature ----------------------

def generate_contest_id():
    """Generate a unique contest ID"""
    timestamp = str(int(time.time()))
    random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"CONTEST_{timestamp}_{random_str}"

@app.route('/create_contest', methods=['GET', 'POST'])
def create_contest():
    """Create a new exam contest (Teacher only)"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role not in ['teacher', 'admin']:
        flash('Access denied. Teachers and admins only.', 'error')
        return redirect(url_for('teacher_dashboard'))
    
    if request.method == 'POST':
        try:
            title = request.form.get('title')
            description = request.form.get('description', '')
            subject_name = request.form.get('subject_name')
            syllabus = request.form.get('syllabus')
            question_format = request.form.get('question_format')
            total_questions = int(request.form.get('total_questions', 10))
            total_marks = int(request.form.get('total_marks', 50))
            time_limit = int(request.form.get('time_limit', 60))
            difficulty_level = request.form.get('difficulty_level', 'medium')
            format_description = request.form.get('format_description', '')
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            
            if not all([title, subject_name, syllabus, question_format, start_date_str, end_date_str]):
                flash('Please fill in all required fields.', 'error')
                return redirect(url_for('create_contest'))
            
            # Parse dates
            start_date = datetime.strptime(start_date_str, '%Y-%m-%dT%H:%M')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M')
            
            if start_date >= end_date:
                flash('End date must be after start date.', 'error')
                return redirect(url_for('create_contest'))
            
            # Validate marks distribution
            if total_marks < total_questions:
                flash('Total marks should be at least equal to the number of questions.', 'error')
                return redirect(url_for('create_contest'))
            
            # Generate unique contest ID
            contest_id = generate_contest_id()
            
            # Generate questions using Gemini API
            questions_data = generate_questions_with_gemini(
                subject_name, syllabus, question_format, total_questions, total_marks, difficulty_level, format_description
            )
            
            if not questions_data:
                flash('Failed to generate questions. Please try again.', 'error')
                return redirect(url_for('create_contest'))
            
            # Create contest
            contest = ExamContest(
                contest_id=contest_id,
                title=title,
                description=description,
                subject_name=subject_name,
                syllabus=syllabus,
                question_format=question_format,
                total_questions=total_questions,
                total_marks=total_marks,
                time_limit=time_limit,
                difficulty_level=difficulty_level,
                format_description=format_description,
                questions=json.dumps(questions_data),
                created_by=session['username'],
                start_date=start_date,
                end_date=end_date,
                is_active=True
            )
            
            db.session.add(contest)
            db.session.commit()
            
            flash('Contest created successfully!', 'success')
            return redirect(url_for('contest_leaderboard', contest_id=contest_id))
            
        except Exception as e:
            print(f"Error creating contest: {e}")
            flash('Error creating contest. Please try again.', 'error')
            return redirect(url_for('create_contest'))
    
    return render_template('create_contest.html', user=user)

@app.route('/contests')
def contests_list():
    """List all available contests"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user:
        flash('User not found.', 'error')
        return redirect(url_for('login'))
    
    # Get all active contests
    contests = db.session.execute(
        db.select(ExamContest).filter_by(is_active=True).order_by(ExamContest.start_date.desc())
    ).scalars().all()
    
    # Get current datetime for template
    now = datetime.now()
    return render_template('contests_list.html', contests=contests, user=user, now=now)

@app.route('/take_contest/<contest_id>')
def take_contest(contest_id):
    """Take a contest exam (Student only)"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('student'))
    
    # Get contest details
    contest = db.session.execute(
        db.select(ExamContest).filter_by(contest_id=contest_id, is_active=True)
    ).scalar()
    
    if not contest:
        flash('Contest not found.', 'error')
        return redirect(url_for('contests_list'))
    
    # Check if contest is currently active
    now = datetime.now()
    print(f"DEBUG: Current time: {now}")
    print(f"DEBUG: Contest start: {contest.start_date}")
    print(f"DEBUG: Contest end: {contest.end_date}")
    print(f"DEBUG: Can start: {now >= contest.start_date}")
    print(f"DEBUG: Can end: {now <= contest.end_date}")
    
    if now < contest.start_date:
        flash(f'Contest has not started yet. Starts at {contest.start_date.strftime("%Y-%m-%d %H:%M")}', 'info')
        return redirect(url_for('contests_list'))
    if now > contest.end_date:
        flash(f'Contest has ended. Ended at {contest.end_date.strftime("%Y-%m-%d %H:%M")}', 'info')
        return redirect(url_for('contests_list'))
    
    # Check if student has already participated
    existing_participation = db.session.execute(
        db.select(ContestParticipation).filter_by(
            contest_id=contest_id, 
            student_username=session['username']
        )
    ).scalar()
    
    if existing_participation:
        flash('You have already participated in this contest.', 'info')
        return redirect(url_for('contest_leaderboard', contest_id=contest_id))
    
    # Parse questions
    questions = json.loads(contest.questions)
    
    # Handle different question formats
    if 'sections' in questions:
        return render_template('take_contest.html', contest=contest, questions=questions, user=user)
    elif 'questions' in questions:
        return render_template('take_contest.html', contest=contest, questions=questions['questions'], user=user)
    else:
        return render_template('take_contest.html', contest=contest, questions=questions, user=user)

@app.route('/submit_contest/<contest_id>', methods=['POST'])
def submit_contest(contest_id):
    """Submit contest exam answers"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role != 'student':
        flash('Access denied. Students only.', 'error')
        return redirect(url_for('student'))
    
    try:
        # Get contest details
        contest = db.session.execute(
            db.select(ExamContest).filter_by(contest_id=contest_id, is_active=True)
        ).scalar()
        
        if not contest:
            flash('Contest not found.', 'error')
            return redirect(url_for('contests_list'))
        
        # Check if student has already participated
        existing_participation = db.session.execute(
            db.select(ContestParticipation).filter_by(
                contest_id=contest_id, 
                student_username=session['username']
            )
        ).scalar()
        
        if existing_participation:
            flash('You have already participated in this contest.', 'info')
            return redirect(url_for('contest_leaderboard', contest_id=contest_id))
        
        # Get answers from form
        answers = {}
        for key, value in request.form.items():
            if key.startswith('answer_'):
                question_id = key.replace('answer_', '')
                answers[question_id] = value
        
        # Calculate time taken
        time_taken = int(request.form.get('time_taken', 0))
        
        # Parse questions for evaluation
        questions_data = json.loads(contest.questions)
        
        # Handle different question formats
        if 'sections' in questions_data:
            questions_for_evaluation = questions_data
        elif 'questions' in questions_data:
            questions_for_evaluation = questions_data['questions']
        else:
            questions_for_evaluation = questions_data
        
        # Evaluate answers
        evaluation_result = evaluate_answers_with_gemini(
            questions_for_evaluation, answers, contest.subject_name
        )
        
        if evaluation_result:
            # Create contest participation record
            participation = ContestParticipation(
                contest_id=contest_id,
                student_username=session['username'],
                student_name=user.username,  # Using username as name for now
                student_roll_no=getattr(user, 'roll_no', ''),
                student_email=user.email,
                department=getattr(user, 'department', ''),
                section=getattr(user, 'section', ''),
                answers=json.dumps(answers),
                time_taken=time_taken,
                total_marks=evaluation_result['total_marks'],
                obtained_marks=evaluation_result['obtained_marks'],
                percentage=evaluation_result['percentage'],
                grade=evaluation_result['grade'],
                detailed_feedback=json.dumps(evaluation_result['detailed_feedback'])
            )
            
            db.session.add(participation)
            db.session.commit()
            
            # Update ranks for this contest
            update_contest_ranks(contest_id)
            
            flash('Contest submitted successfully!', 'success')
            return redirect(url_for('contest_leaderboard', contest_id=contest_id))
        else:
            flash('Failed to evaluate contest. Please try again.', 'error')
            return redirect(url_for('contests_list'))
            
    except Exception as e:
        print(f"Error submitting contest: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error submitting contest: {str(e)}', 'error')
        return redirect(url_for('contests_list'))

def update_contest_ranks(contest_id):
    """Update ranks for all participants in a contest"""
    try:
        # Get all participations for this contest, ordered by percentage (descending)
        participations = db.session.execute(
            db.select(ContestParticipation).filter_by(contest_id=contest_id)
            .order_by(ContestParticipation.percentage.desc())
        ).scalars().all()
        
        # Update ranks
        for rank, participation in enumerate(participations, 1):
            participation.rank = rank
        
        db.session.commit()
        print(f"Updated ranks for contest {contest_id}")
        
    except Exception as e:
        print(f"Error updating contest ranks: {e}")

@app.route('/contest_leaderboard/<contest_id>')
def contest_leaderboard(contest_id):
    """View contest leaderboard with bar graph"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user:
        flash('User not found.', 'error')
        return redirect(url_for('login'))
    
    # Get contest details
    contest = db.session.execute(
        db.select(ExamContest).filter_by(contest_id=contest_id)
    ).scalar()
    
    if not contest:
        flash('Contest not found.', 'error')
        return redirect(url_for('contests_list'))
    
    # Get all participations for this contest, ordered by rank
    participations = db.session.execute(
        db.select(ContestParticipation).filter_by(contest_id=contest_id)
        .order_by(ContestParticipation.rank.asc())
    ).scalars().all()
    
    # Prepare data for bar graph
    leaderboard_data = []
    for participation in participations:
        leaderboard_data.append({
            'rank': participation.rank,
            'student_name': participation.student_name,
            'percentage': participation.percentage,
            'marks': f"{participation.obtained_marks}/{participation.total_marks}",
            'grade': participation.grade,
            'time_taken': participation.time_taken,
            'department': participation.department or 'N/A',
            'section': participation.section or 'N/A'
        })
    
    return render_template('contest_leaderboard.html', 
                         contest=contest, 
                         leaderboard_data=leaderboard_data, 
                         user=user)

@app.route('/teacher_contests')
def teacher_contests():
    """View contests created by teacher"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    user = db.session.execute(db.select(Credentials).filter_by(username=session['username'])).scalar()
    if not user or user.role not in ['teacher', 'admin']:
        flash('Access denied. Teachers and admins only.', 'error')
        return redirect(url_for('teacher_dashboard'))
    
    # Get contests created by this teacher
    contests = db.session.execute(
        db.select(ExamContest).filter_by(created_by=session['username'])
        .order_by(ExamContest.created_at.desc())
    ).scalars().all()
    
    # Get current datetime for template
    now = datetime.now()
    return render_template('teacher_contests.html', contests=contests, user=user, now=now)

if __name__ == '__main__':
    app.run(debug=True,host='0.0.0.0')
